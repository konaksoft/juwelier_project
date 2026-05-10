"""
==============================================================================
 FAZ B.3 — Excel Raporlama Exporter
==============================================================================

Tarih: 2026-05-05
Amaç:
    Tam firma yedeğini insan-okunabilir, denormalize edilmiş Excel raporu
    olarak üretir. Muhasebe denetimi, stok envanter dökümü ve cari ekstre
    için kullanılır.

⚠ ÖNEMLİ:
    Bu format SADECE OKUMA AMAÇLIDIR. Geri yükleme YAPILMAZ.
    Excel dosyası DB'ye geri import edilemez (bu kasıtlı bir tasarım).
    Eğer bir mağazaya geri yükleme gerekiyorsa JSON / ZIP yedeği kullanılmalı.

Sayfa Yapısı:
    1. Genel Bilgi      — yedek meta + kayıt sayıları
    2. Barkodlu Ürünler — GoldPurchases + Products JOIN
    3. Tedarikçiler     — Suppliers
    4. Tedarikçi Hareketleri — SupplierLedger
    5. Müşteriler       — Customers
    6. Müşteri Hareketleri   — CustomerLedger (TÜM satırlar — yasal)
    7. Satışlar         — Process (transaction_type='SALE')
    8. Kasalar          — BankAccount
    9. Stok Özeti       — StockSnapshot son hali

Stil:
    - Header: bold, beyaz yazı, lacivert dolgu
    - Sayısal kolonlar: sağa hizalı + decimal format
    - Tüm sayfalarda 1. satır freeze (donuk)
==============================================================================
"""

import io
from decimal import Decimal

from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---- Stiller -----------------------------------------------------------------
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
NUMBER_FORMAT = '#,##0.000'
TL_FORMAT = '#,##0.00 ₺'


# ==============================================================================
#  YARDIMCI
# ==============================================================================

def _write_header(ws, headers, row=1):
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize_columns(ws, max_width=50):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        try:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=10,
            )
        except Exception:
            max_len = 12
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), max_width)


def _safe_str(val):
    if val is None:
        return ''
    return str(val)


def _safe_decimal(val):
    if val is None or val == '':
        return None
    try:
        return float(Decimal(str(val)))
    except Exception:
        return None


# ==============================================================================
#  ANA SERVİS
# ==============================================================================

class XlsxExportService:
    """
    Tam firma yedeğini Excel raporu olarak üretir.

    Public API:
        export() → bytes (Excel dosyası içeriği)
    """

    def __init__(self, company_id):
        from apps.stores.models import Company, Stores
        self.company = Company.objects.get(id=company_id)
        self.stores = Stores.objects.filter(company=self.company, is_deleted=False)
        self.store_ids = list(self.stores.values_list('id', flat=True))

    def export(self):
        """
        Excel dosyasını bellekte oluşturup binary content olarak döner.

        Returns:
            bytes — .xlsx dosyası içeriği.
        """
        wb = openpyxl.Workbook()
        # Default sheet'i kaldır
        if wb.active is not None:
            wb.remove(wb.active)

        self._sheet_summary(wb)
        self._sheet_gold_purchases(wb)
        self._sheet_suppliers(wb)
        self._sheet_supplier_ledger(wb)
        self._sheet_customers(wb)
        self._sheet_customer_ledger(wb)
        self._sheet_sales(wb)
        self._sheet_bank_accounts(wb)
        self._sheet_stock_summary(wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # --------------------------------------------------------------------------
    #  Sayfa 1 — Genel Bilgi (özet metrikleri)
    # --------------------------------------------------------------------------
    def _sheet_summary(self, wb):
        from apps.products.models import Products
        from apps.gold_purchases.models import GoldPurchases
        from apps.customers.models import Customers, CustomerLedger
        from apps.suppliers.models import Suppliers, SupplierLedger
        from apps.process.models import Process
        from apps.banking.models import BankAccount, CashboxLedger

        ws = wb.create_sheet('Genel Bilgi')
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 35

        title_cell = ws.cell(row=1, column=1, value='YEDEK ÖZET RAPORU')
        title_cell.font = Font(name='Calibri', size=14, bold=True, color='0D47A1')
        ws.merge_cells('A1:B1')

        rows = [
            ('Firma', self.company.title or '-'),
            ('Firma ID', str(self.company.id)),
            ('Mağaza Sayısı', self.stores.count()),
            ('Yedek Tarihi', timezone.now().strftime('%d.%m.%Y %H:%M')),
            ('', ''),
            ('=== KAYIT SAYILARI ===', ''),
            ('Barkodlu Ürün (GoldPurchases)', GoldPurchases.objects.filter(store__in=self.store_ids, is_deleted=False).count()),
            ('Toplam Ürün (Products)', Products.objects.filter(store__in=self.store_ids, is_deleted=False).count()),
            ('Tedarikçi', Suppliers.objects.filter(store__in=self.store_ids, is_deleted=False).count()),
            ('Tedarikçi Hareketi', SupplierLedger.objects.filter(supplier__store__in=self.store_ids).count()),
            ('Müşteri', Customers.objects.filter(store__in=self.store_ids, is_deleted=False).distinct().count()),
            ('Müşteri Hareketi', CustomerLedger.objects.filter(store__in=self.store_ids).count()),
            ('İşlem (Process)', Process.objects.filter(store__in=self.store_ids, is_deleted=False).count()),
            ('Satış (SALE)', Process.objects.filter(store__in=self.store_ids, is_deleted=False, transaction_type='SALE').count()),
            ('Kasa/Banka Hesabı', BankAccount.objects.filter(store__in=self.store_ids, is_deleted=False).count()),
            ('Kasa Hareketi', CashboxLedger.objects.filter(store__in=self.store_ids).count()),
        ]

        for r_idx, (label, value) in enumerate(rows, start=2):
            label_cell = ws.cell(row=r_idx, column=1, value=label)
            value_cell = ws.cell(row=r_idx, column=2, value=value)
            if label.startswith('==='):
                label_cell.font = Font(bold=True, color='0D47A1')
            elif label:
                label_cell.font = Font(bold=True)

    # --------------------------------------------------------------------------
    #  Sayfa 2 — Barkodlu Ürünler
    # --------------------------------------------------------------------------
    def _sheet_gold_purchases(self, wb):
        from apps.gold_purchases.models import GoldPurchases

        ws = wb.create_sheet('Barkodlu Ürünler')
        headers = [
            'Barkod', 'Ürün Adı', 'Tip', 'Materyal', 'Ayar', 'Gram',
            'Alış HS', 'Alış TL', 'Tedarikçi', 'Mağaza', 'Durum',
            'Eklenme', 'Eklenme Tarihi',
        ]
        _write_header(ws, headers)

        qs = GoldPurchases.objects.filter(
            store__in=self.store_ids, is_deleted=False
        ).select_related('product', 'supplier', 'store', 'created_by')

        for r_idx, gp in enumerate(qs.iterator(chunk_size=500), start=2):
            p = gp.product
            ws.cell(row=r_idx, column=1, value=_safe_str(getattr(p, 'barcode', '')))
            ws.cell(row=r_idx, column=2, value=_safe_str(getattr(p, 'name', '')))
            ws.cell(row=r_idx, column=3, value=_safe_str(getattr(p, 'jewelry_type', '')))
            ws.cell(row=r_idx, column=4, value=_safe_str(getattr(p, 'material_type', '')))
            ws.cell(row=r_idx, column=5, value=_safe_str(getattr(p, 'gold_rate', '')))
            ws.cell(row=r_idx, column=6, value=_safe_decimal(getattr(p, 'gram', None))).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=7, value=_safe_decimal(getattr(p, 'buy_price_hs', None))).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=8, value=_safe_decimal(getattr(p, 'buy_price_eur', None))).number_format = TL_FORMAT
            sup = gp.supplier
            ws.cell(row=r_idx, column=9, value=(
                getattr(sup, 'company_name', '') or
                f"{getattr(sup, 'person_name', '')} {getattr(sup, 'person_surname', '')}".strip()
            ) if sup else '-')
            ws.cell(row=r_idx, column=10, value=_safe_str(getattr(gp.store, 'title', '')))
            ws.cell(row=r_idx, column=11, value='Satıldı' if getattr(p, 'is_completed', False) else 'Stokta')
            ws.cell(row=r_idx, column=12, value=_safe_str(gp.created_by) if gp.created_by else '-')
            created_on = getattr(gp, 'created_on', None) or getattr(gp, 'created_at', None)
            ws.cell(row=r_idx, column=13, value=created_on.strftime('%d.%m.%Y %H:%M') if created_on else '-')

        _autosize_columns(ws)

    # --------------------------------------------------------------------------
    #  Sayfa 3 — Tedarikçiler
    # --------------------------------------------------------------------------
    def _sheet_suppliers(self, wb):
        from apps.suppliers.models import Suppliers

        ws = wb.create_sheet('Tedarikçiler')
        headers = ['Şirket', 'Ad', 'Soyad', 'Telefon', 'E-Posta', 'Vergi No',
                   'Vergi Dairesi', 'Şehir', 'Hesap Tipi', 'Mağaza']
        _write_header(ws, headers)

        qs = Suppliers.objects.filter(
            store__in=self.store_ids, is_deleted=False
        ).select_related('store', 'tax_office', 'city')

        for r_idx, s in enumerate(qs.iterator(chunk_size=500), start=2):
            ws.cell(row=r_idx, column=1, value=_safe_str(s.company_name))
            ws.cell(row=r_idx, column=2, value=_safe_str(s.person_name))
            ws.cell(row=r_idx, column=3, value=_safe_str(s.person_surname))
            ws.cell(row=r_idx, column=4, value=_safe_str(s.phone))
            ws.cell(row=r_idx, column=5, value=_safe_str(s.email))
            ws.cell(row=r_idx, column=6, value=_safe_str(s.tax_number))
            ws.cell(row=r_idx, column=7, value=_safe_str(getattr(s.tax_office, 'name', '') if s.tax_office else ''))
            ws.cell(row=r_idx, column=8, value=_safe_str(getattr(s.city, 'name', '') if s.city else ''))
            ws.cell(row=r_idx, column=9, value=_safe_str(s.account_type))
            ws.cell(row=r_idx, column=10, value=_safe_str(getattr(s.store, 'title', '')))

        _autosize_columns(ws)

    # --------------------------------------------------------------------------
    #  Sayfa 4 — Tedarikçi Hareketleri
    # --------------------------------------------------------------------------
    def _sheet_supplier_ledger(self, wb):
        from apps.suppliers.models import SupplierLedger

        ws = wb.create_sheet('Tedarikçi Hareketleri')
        headers = ['Tarih', 'Tedarikçi', 'İşlem No', 'Tip', 'Adet', 'Gram',
                   'Tutar', 'Para Birimi', 'Kur (TL)', 'Aktif']
        _write_header(ws, headers)

        qs = SupplierLedger.objects.filter(
            supplier__store__in=self.store_ids
        ).select_related('supplier').order_by('-created_on')

        for r_idx, sl in enumerate(qs.iterator(chunk_size=1000), start=2):
            ws.cell(row=r_idx, column=1, value=sl.created_on.strftime('%d.%m.%Y %H:%M') if sl.created_on else '-')
            sup = sl.supplier
            ws.cell(row=r_idx, column=2, value=(
                getattr(sup, 'company_name', '') or
                f"{getattr(sup, 'person_name', '')} {getattr(sup, 'person_surname', '')}".strip()
            ) if sup else '-')
            ws.cell(row=r_idx, column=3, value=_safe_str(sl.process_no))
            ws.cell(row=r_idx, column=4, value=_safe_str(sl.transaction_type))
            ws.cell(row=r_idx, column=5, value=_safe_decimal(sl.quantity_piece)).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=6, value=_safe_decimal(sl.quantity_gram)).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=7, value=_safe_decimal(sl.amount_value)).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=8, value=_safe_str(sl.currency))
            ws.cell(row=r_idx, column=9, value=_safe_decimal(getattr(sl, 'exchange_rate_eur', None))).number_format = TL_FORMAT
            ws.cell(row=r_idx, column=10, value='Evet' if getattr(sl, 'is_active', True) else 'Hayır')

        _autosize_columns(ws)

    # --------------------------------------------------------------------------
    #  Sayfa 5 — Müşteriler
    # --------------------------------------------------------------------------
    def _sheet_customers(self, wb):
        from apps.customers.models import Customers

        ws = wb.create_sheet('Müşteriler')
        headers = ['Ad', 'Soyad', 'TCKN/Pasaport', 'Müşteri No', 'Telefon',
                   'E-Posta', 'Şehir', 'Adres', 'Telefon Onaylı', 'Email Onaylı']
        _write_header(ws, headers)

        qs = Customers.objects.filter(
            store__in=self.store_ids, is_deleted=False
        ).distinct().select_related('city')

        for r_idx, c in enumerate(qs.iterator(chunk_size=500), start=2):
            ws.cell(row=r_idx, column=1, value=_safe_str(c.first_name))
            ws.cell(row=r_idx, column=2, value=_safe_str(c.last_name))
            ws.cell(row=r_idx, column=3, value=_safe_str(c.identification_number))
            ws.cell(row=r_idx, column=4, value=_safe_str(c.customer_number))
            ws.cell(row=r_idx, column=5, value=_safe_str(c.phone))
            ws.cell(row=r_idx, column=6, value=_safe_str(c.email))
            ws.cell(row=r_idx, column=7, value=_safe_str(getattr(c.city, 'name', '') if c.city else ''))
            ws.cell(row=r_idx, column=8, value=_safe_str(c.address))
            ws.cell(row=r_idx, column=9, value='Evet' if getattr(c, 'is_phone_verified', False) else 'Hayır')
            ws.cell(row=r_idx, column=10, value='Evet' if getattr(c, 'is_email_verified', False) else 'Hayır')

        _autosize_columns(ws)

    # --------------------------------------------------------------------------
    #  Sayfa 6 — Müşteri Hareketleri (Cari)
    # --------------------------------------------------------------------------
    def _sheet_customer_ledger(self, wb):
        from apps.customers.models import CustomerLedger

        ws = wb.create_sheet('Müşteri Hareketleri')
        headers = ['Tarih', 'Müşteri', 'İşlem No', 'Tip', 'Tutar HS', 'Tutar TL',
                   'Para Birimi', 'Kur (TL)', 'Açıklama', 'Aktif']
        _write_header(ws, headers)

        qs = CustomerLedger.objects.filter(
            store__in=self.store_ids
        ).select_related('customer').order_by('-created_on')

        for r_idx, cl in enumerate(qs.iterator(chunk_size=1000), start=2):
            ws.cell(row=r_idx, column=1, value=cl.created_on.strftime('%d.%m.%Y %H:%M') if cl.created_on else '-')
            cust = cl.customer
            ws.cell(row=r_idx, column=2, value=(
                f"{getattr(cust, 'first_name', '')} {getattr(cust, 'last_name', '')}".strip()
            ) if cust else '-')
            ws.cell(row=r_idx, column=3, value=_safe_str(cl.process_no))
            ws.cell(row=r_idx, column=4, value=_safe_str(cl.transaction_type))
            ws.cell(row=r_idx, column=5, value=_safe_decimal(cl.amount_hs)).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=6, value=_safe_decimal(cl.amount_eur)).number_format = TL_FORMAT
            ws.cell(row=r_idx, column=7, value=_safe_str(cl.currency))
            ws.cell(row=r_idx, column=8, value=_safe_decimal(getattr(cl, 'exchange_rate_eur', None))).number_format = TL_FORMAT
            ws.cell(row=r_idx, column=9, value=_safe_str(cl.description))
            ws.cell(row=r_idx, column=10, value='Evet' if getattr(cl, 'is_active', True) else 'Hayır')

        _autosize_columns(ws)

    # --------------------------------------------------------------------------
    #  Sayfa 7 — Satışlar (Process: SALE)
    # --------------------------------------------------------------------------
    def _sheet_sales(self, wb):
        from apps.process.models import Process

        ws = wb.create_sheet('Satışlar')
        headers = ['Tarih', 'İşlem No', 'Tip', 'Müşteri', 'Ürün', 'Adet',
                   'Gram', 'Birim Fiyat', 'Tutar TL', 'HS Tutar', 'Personel', 'Mağaza']
        _write_header(ws, headers)

        qs = Process.objects.filter(
            store__in=self.store_ids, is_deleted=False, transaction_type='SALE'
        ).select_related('customer', 'product', 'employee', 'store').order_by('-date')

        for r_idx, p in enumerate(qs.iterator(chunk_size=500), start=2):
            ws.cell(row=r_idx, column=1, value=p.date.strftime('%d.%m.%Y %H:%M') if p.date else '-')
            ws.cell(row=r_idx, column=2, value=_safe_str(p.process_no))
            ws.cell(row=r_idx, column=3, value=_safe_str(p.process_type))
            cust = p.customer
            ws.cell(row=r_idx, column=4, value=(
                f"{getattr(cust, 'first_name', '')} {getattr(cust, 'last_name', '')}".strip()
            ) if cust else '-')
            ws.cell(row=r_idx, column=5, value=_safe_str(getattr(p.product, 'name', '') if p.product else ''))
            ws.cell(row=r_idx, column=6, value=_safe_decimal(p.piece)).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=7, value=_safe_decimal(p.gram)).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=8, value=_safe_decimal(getattr(p, 'unit_price', None))).number_format = TL_FORMAT
            ws.cell(row=r_idx, column=9, value=_safe_decimal(getattr(p, 'amount', None))).number_format = TL_FORMAT
            ws.cell(row=r_idx, column=10, value=_safe_decimal(getattr(p, 'price_hs', None))).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=11, value=_safe_str(p.employee) if p.employee else '-')
            ws.cell(row=r_idx, column=12, value=_safe_str(getattr(p.store, 'title', '')))

        _autosize_columns(ws)

    # --------------------------------------------------------------------------
    #  Sayfa 8 — Kasalar
    # --------------------------------------------------------------------------
    def _sheet_bank_accounts(self, wb):
        from apps.banking.models import BankAccount

        ws = wb.create_sheet('Kasalar')
        headers = ['Ad', 'Banka', 'IBAN', 'Para Birimi', 'Hesap Tipi',
                   'Mutabakat Tolerans', 'Mağaza', 'Şube Transit', 'Aktif']
        _write_header(ws, headers)

        qs = BankAccount.objects.filter(
            store__in=self.store_ids, is_deleted=False
        ).select_related('store')

        for r_idx, ba in enumerate(qs.iterator(chunk_size=500), start=2):
            ws.cell(row=r_idx, column=1, value=_safe_str(ba.name))
            ws.cell(row=r_idx, column=2, value=_safe_str(ba.bank_name))
            ws.cell(row=r_idx, column=3, value=_safe_str(ba.iban))
            ws.cell(row=r_idx, column=4, value=_safe_str(ba.currency))
            ws.cell(row=r_idx, column=5, value=_safe_str(ba.account_type))
            ws.cell(row=r_idx, column=6, value=_safe_decimal(getattr(ba, 'reconciliation_tolerance', None))).number_format = TL_FORMAT
            ws.cell(row=r_idx, column=7, value=_safe_str(getattr(ba.store, 'title', '')))
            ws.cell(row=r_idx, column=8, value='Evet' if getattr(ba, 'is_inter_branch_transit_account', False) else 'Hayır')
            ws.cell(row=r_idx, column=9, value='Evet' if getattr(ba, 'is_active', True) else 'Hayır')

        _autosize_columns(ws)

    # --------------------------------------------------------------------------
    #  Sayfa 9 — Stok Özeti (StockSnapshot son hali)
    # --------------------------------------------------------------------------
    def _sheet_stock_summary(self, wb):
        from apps.stock_management.models import StockSnapshot

        ws = wb.create_sheet('Stok Özeti')
        headers = ['Mağaza', 'Ürün', 'Adet', 'Gram', 'WAC HS', 'WAC TL', 'Güncellenme']
        _write_header(ws, headers)

        qs = StockSnapshot.objects.filter(
            store__in=self.store_ids
        ).select_related('store', 'product').order_by('-updated_on' if hasattr(StockSnapshot, 'updated_on') else '-id')

        for r_idx, ss in enumerate(qs.iterator(chunk_size=1000), start=2):
            ws.cell(row=r_idx, column=1, value=_safe_str(getattr(ss.store, 'title', '')))
            ws.cell(row=r_idx, column=2, value=_safe_str(getattr(ss.product, 'name', '') if ss.product else '-'))
            ws.cell(row=r_idx, column=3, value=_safe_decimal(getattr(ss, 'piece', None))).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=4, value=_safe_decimal(getattr(ss, 'gram', None))).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=5, value=_safe_decimal(getattr(ss, 'wac_hs', None))).number_format = NUMBER_FORMAT
            ws.cell(row=r_idx, column=6, value=_safe_decimal(getattr(ss, 'wac_tl', None))).number_format = TL_FORMAT
            updated = getattr(ss, 'updated_on', None) or getattr(ss, 'created_on', None)
            ws.cell(row=r_idx, column=7, value=updated.strftime('%d.%m.%Y %H:%M') if updated else '-')

        _autosize_columns(ws)
