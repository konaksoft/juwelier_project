# ============================================================================
# DOSYA: apps/banking/bank_views.py
# KONUM: Kuyum Plus (jewelery_project)
# VERSİYON: v6 — Kasa Yönetimi (İç Kasa) View'ları
#
# Bu dosya /banks/ prefix'i altında çalışan "Kasa Yönetimi" modülünü besler.
# Açık Bankacılık (Mutabakat) ile karıştırılmamalıdır.
#
# View'lar:
#   1. bank_management_index     — Ana liste sayfası
#   2. bank_management_get_all   — DataTables JSON (bakiye annotate'li)
#   3. bank_management_detail    — Tek hesap detay sayfası
#   4. bank_management_payments  — Detay DataTables JSON
#   5. bank_management_export    — PDF / Excel export
# ============================================================================

import io
import json
import logging
from datetime import datetime as dt_datetime
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import (
    Case, DecimalField, F, Q, Sum, Value, When,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.utils import timezone

from apps.banking.models import BankAccount
from apps.banking.services import (
    FX_SENTINEL_MAP,
    FX_SENTINEL_REVERSE_MAP,
    SUPPORTED_FX_CURRENCIES,
)
from apps.process.models import Payment

log = logging.getLogger(__name__)

# Mevcut view'lardan save/delete'i yeniden-export (bank_urls.py'den erişim için)
from apps.banking.views import save_bank_account, delete_bank_account  # noqa: F401


# ──────────────────────────────────────────────────────
# YARDIMCILAR
# ──────────────────────────────────────────────────────

def _get_store(request):
    return getattr(request.user, 'store', None)


def _get_store_primary_currency(store, default='EUR'):
    """
    FAZ 20.x — StoreConfiguration.primary_currency okur.

    StoreConfig yoksa veya alan boşsa `default` döner. Asla istisna fırlatmaz;
    rapor/özet akışları her durumda çalışmaya devam etmelidir.

    SSOT: apps.settings.currency.get_store_primary_currency
    (Üçüncü bir kopya çıkmasın diye kanonik modüle delege edildi.)
    """
    from apps.settings.currency import get_store_primary_currency
    return get_store_primary_currency(store, default=default)


def _require_store(request):
    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Kullaniciya bagli magaza bulunamadi.'})
    return None


_ZERO = Value(Decimal('0'), output_field=DecimalField(max_digits=15, decimal_places=2))


def _guess_fx_code_from_rate(exchange_rate):
    """
    FAZ 20.4 / Faz 13: Exchange rate'ten döviz türünü tespit et (FALLBACK YOL).

    İki mod:
    1. Sentinel rate (düzeltme fişleri): services.FX_SENTINEL_MAP üzerinden okunur.
       Yazma ve okuma yolları aynı SSOT haritasını kullanır.
    2. Gerçek kur (döviz bozma): Kur aralığından tahmin (yalnızca reference'sız
       eski kayıtlar için son savunma hattı; yeni kayıtlar her zaman reference taşır).

    NOT: Bu fonksiyon yalnızca Payment.reference boşsa devreye girer.
         Yeni eklenen tüm FX yazma yolları reference=[KOD] formatını zorunlu kılar.
    """
    r = float(exchange_rate)

    # Sentinel rate kontrolü (services.FX_SENTINEL_MAP'ten)
    for sentinel, code in FX_SENTINEL_REVERSE_MAP.items():
        if abs(r - sentinel) < 0.001:
            return code

    # Gerçek kur aralığı tahmini (Nisan 2026 döviz kurları, TL bazında):
    # QAR ~9 TL, CAD ~25 TL, USD ~34 TL, AUD ~22 TL, CHF ~39 TL, SAR ~9 TL,
    # EUR ~37 TL, GBP ~44 TL.
    # Aralıklar fallback amaçlıdır; yeni FX kayıtları reference üzerinden okunur.
    if r < 15:
        return 'QAR'
    elif r < 30:
        return 'CAD'
    elif r < 36:
        return 'USD'
    elif r < 41:
        return 'EUR'
    elif r < 50:
        return 'GBP'
    return 'Döviz'


def _extract_fx_code_from_reference(reference):
    """
    FAZ 20.4 / Faz 13: Payment.reference alanından döviz kodunu çıkar.
    Format: "[EUR] Açıklama notları"

    Whitelist services.SUPPORTED_FX_CURRENCIES SSOT'undan dinamik okunur;
    yeni döviz eklemek yalnızca services.CURRENCY_FROM_PRODUCT_NAME güncellemesi
    gerektirir.
    """
    if not reference:
        return None
    ref = str(reference).strip()
    if ref.startswith('[') and ']' in ref:
        code = ref[1:ref.index(']')].strip().upper()
        if code in SUPPORTED_FX_CURRENCIES:
            return code
    return None


def _get_fx_breakdown(account):
    """
    FAZ 20.4: Merkez Döviz Kasası (currency='FX') için döviz bazlı bakiye kırılımı.
    Her döviz türünün (USD, EUR, GBP) ayrı bakiyesini hesaplar.

    Döviz kodu tespit sırası:
    1. Payment.reference'dan [EUR] prefix'i oku (en güvenilir)
    2. exchange_rate sentinel değerinden eşle (0.01=USD, 0.02=EUR...)
    3. exchange_rate kur aralığından tahmin et (fallback)

    Returns:
        dict — {'USD': '100.00', 'EUR': '50.00'} veya None
    """
    if not account or account.currency != 'FX':
        return None

    fx_payments = Payment.objects.filter(
        bank_account=account,
        is_cancelled=False,
        is_approved=True,
    ).exclude(
        currency_amount__isnull=True,
    ).exclude(
        currency_amount=0,
    )

    fx_data = {}
    for fp in fx_payments:
        direction_sign = Decimal('-1') if fp.is_output else Decimal('1')

        # 1. Reference'dan currency kodu oku
        fx_code = _extract_fx_code_from_reference(fp.reference)
        # 2. Exchange rate'den tespit et
        if not fx_code and fp.exchange_rate and fp.exchange_rate > 0:
            fx_code = _guess_fx_code_from_rate(fp.exchange_rate)
        # 3. Fallback
        if not fx_code:
            fx_code = 'Döviz'

        if fx_code not in fx_data:
            fx_data[fx_code] = Decimal('0')
        fx_data[fx_code] += direction_sign * (fp.currency_amount or Decimal('0'))

    return {k: str(v) for k, v in fx_data.items()} if fx_data else None


def _effective_amount_expr():
    """
    Komisyon (net_amount) varsa net_amount, yoksa amount dondurur.
    POS odemelerinde bankaya gercekte dusen net_amount'tur.
    """
    return Coalesce(F('payments__net_amount'), F('payments__amount'))


def _multicurrency_amount_expr(primary_cur='TRY'):
    """
    FAZ 20: Çoklu para birimi desteği (Merkez Döviz Kasası dahil).

    Kasanın currency alanına göre doğru tutarı seçer:
      - Birincil para birimi kasaları → net_amount → amount
      - FX kasaları (Merkez Döviz)    → currency_amount → amount (TL bakiyeyi gösterir)
      - Eski döviz kasaları (USD, EUR vb.) → currency_amount → amount

    get_bank_balance_qs() annotation'ları içinde kullanılır.
    FX kasası için bakiye TL olarak gösterilir (döviz bazlı gruplama detail'de yapılır).

    FAZ 20.x: Birincil para birimi artık parametre (önceki sabit 'TRY' geriye uyumlu default).
    """
    return Case(
        # Birincil para birimi kasaları: mevcut mantık (net_amount → amount)
        When(
            currency=primary_cur,
            then=Coalesce(
                F('payments__net_amount'),
                F('payments__amount'),
            ),
        ),
        # Döviz kasaları: currency_amount → amount (fallback)
        default=Coalesce(
            F('payments__currency_amount'),
            F('payments__amount'),
        ),
        output_field=DecimalField(max_digits=15, decimal_places=2),
    )


def get_bank_balance_qs(store):
    """
    BankAccount queryset'ine bakiye annotate'leri ekler.

    Her hesap icin:
      - total_in  : Gelen toplam (is_output=False)
      - total_out : Giden toplam (is_output=True)

    FAZ 15: is_cancelled=True olan Payment kayitlari hariç tutulur.
    FAZ 17: Döviz kasaları (USD, EUR vb.) için currency_amount kullanılır.
            TRY kasaları için mevcut net_amount → amount mantığı korunur.
    FAZ 18: is_approved=False olan Payment kayıtları bakiyeye dahil edilmez.

    Kullanim:
        qs = get_bank_balance_qs(store)
        for acc in qs:
            balance = (acc.total_in or 0) - (acc.total_out or 0)
    """
    # FAZ 17: _multicurrency_amount_expr() kasanın currency alanına göre
    # doğru tutarı seçer (birincil para → net_amount/amount, döviz → currency_amount/amount)
    # FAZ 20.x: Birincil para birimi StoreConfiguration'dan okunur.
    _primary_cur = _get_store_primary_currency(store, default='TRY')
    _amt = _multicurrency_amount_expr(primary_cur=_primary_cur)

    return BankAccount.objects.filter(
        store=store, is_deleted=False,
    ).annotate(
        total_in=Coalesce(
            Sum(
                Case(
                    When(
                        payments__is_output=False,
                        payments__is_cancelled=False,
                        payments__is_approved=True,  # FAZ 18
                        then=_amt,
                    ),
                    output_field=DecimalField(max_digits=15, decimal_places=2),
                ),
            ),
            _ZERO,
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
        total_out=Coalesce(
            Sum(
                Case(
                    When(
                        payments__is_output=True,
                        payments__is_cancelled=False,
                        payments__is_approved=True,  # FAZ 18
                        then=_amt,
                    ),
                    output_field=DecimalField(max_digits=15, decimal_places=2),
                ),
            ),
            _ZERO,
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
    )


def get_account_summary(account):
    """
    Tek bir banka hesabinin ozet bilgilerini dondurur.

    FAZ 15: is_cancelled=True olan Payment kayitlari hariç tutulur.
    FAZ 17: Döviz kasaları (USD, EUR vb.) için currency_amount kullanılır.
            TRY kasaları için mevcut net_amount → amount mantığı korunur.
    FAZ 18: is_approved=False olan Payment kayıtları bakiyeye dahil edilmez.

    Return dict:
        total_in        — Brut gelen toplam (kasanın para birimi cinsinden)
        total_out       — Brut giden toplam (kasanın para birimi cinsinden)
        total_in_net    — Net gelen toplam (komisyon dusulmus, sadece TRY kasaları)
        total_out_net   — Net giden toplam (sadece TRY kasaları)
        total_commission— Toplam komisyon tutari (TL cinsinden)
        balance         — Brut bakiye (total_in - total_out)
        balance_net     — Net bakiye  (total_in_net - total_out_net)
        is_foreign      — bool: Kasanın para birimi TRY değilse True
        pending_count   — int: Onay bekleyen ödeme sayısı (FAZ 18)
    """
    # FAZ 15+18: İptal edilmiş ve onaylanmamış ödemeleri kapsam dışında tut
    base_qs = Payment.objects.filter(bank_account=account, is_cancelled=False, is_approved=True)

    acct_currency = getattr(account, 'currency', 'TRY') or 'TRY'
    is_foreign = (acct_currency != 'TRY')

    if is_foreign:
        # FAZ 17: Döviz kasaları — currency_amount üzerinden toplam
        # currency_amount NULL ise fallback olarak amount kullanılır (eski kayıtlar)
        _amt_field = Coalesce(F('currency_amount'), F('amount'))

        agg = base_qs.aggregate(
            total_in=Coalesce(
                Sum(_amt_field, filter=Q(is_output=False)),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_out=Coalesce(
                Sum(_amt_field, filter=Q(is_output=True)),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            # Döviz kasalarında net_amount kavramı yok (POS komisyonu TL cinsindendir).
            # total_in_net ve total_out_net brüt ile aynı kalır.
            total_in_net=Coalesce(
                Sum(_amt_field, filter=Q(is_output=False)),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_out_net=Coalesce(
                Sum(_amt_field, filter=Q(is_output=True)),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            # Komisyon TL cinsindendir — döviz kasaları için yine de toplanır
            total_commission=Coalesce(
                Sum('commission_amount'),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
        )
    else:
        # TRY kasaları — mevcut mantık (net_amount → amount)
        agg = base_qs.aggregate(
            total_in=Coalesce(
                Sum('amount', filter=Q(is_output=False)),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_out=Coalesce(
                Sum('amount', filter=Q(is_output=True)),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_in_net=Coalesce(
                Sum(
                    Coalesce(F('net_amount'), F('amount')),
                    filter=Q(is_output=False),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_out_net=Coalesce(
                Sum(
                    Coalesce(F('net_amount'), F('amount')),
                    filter=Q(is_output=True),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_commission=Coalesce(
                Sum('commission_amount'),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
        )

    agg['balance'] = agg['total_in'] - agg['total_out']
    agg['balance_net'] = agg['total_in_net'] - agg['total_out_net']
    agg['is_foreign'] = is_foreign

    # FAZ 18: Onay bekleyen ödeme sayısı
    agg['pending_count'] = Payment.objects.filter(
        bank_account=account, is_cancelled=False, is_approved=False,
    ).count()

    # FAZ 20 / Faz 13.2: Merkez Döviz Kasası (currency='FX') için bakiye kırılımı.
    # SSOT: _get_fx_breakdown() — Payment.reference [KOD] etiketini önce okur,
    # yalnızca eski referanssız kayıtlar için kur aralığı fallback'ine düşer.
    agg['fx_breakdown'] = _get_fx_breakdown(account) if acct_currency == 'FX' else None

    return agg


# ──────────────────────────────────────────────────────
# 1. ANA LİSTE SAYFASI
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def bank_management_index(request):
    """Kasa Yonetimi ana sayfa — hesap listesi + bakiye DataTable."""
    return render(request, 'management/banks/index.html', {
        'title': 'Kasa Yonetimi',
    })


# ──────────────────────────────────────────────────────
# 1b. KONSOLİDE RAPOR — TÜM KASALARIN GENEL DURUMU
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def bank_consolidated_report(request):
    """
    Tum aktif kasalarin konsolide ozetini JSON olarak dondurur.
    TL kasaları ile Döviz kasaları KESİNLİKLE ayrı hesaplanır.
    """
    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Magaza bulunamadi.'})

    qs = BankAccount.objects.filter(
        store=store, is_deleted=False, is_active=True,
    )

    # FAZ 20.x: Birincil para birimi StoreConfiguration'dan okunur.
    _primary_cur = _get_store_primary_currency(store, default='TRY')

    # 1. SADECE BİRİNCİL PARA BİRİMİ KASALARINI ANA TOPLAMA DAHİL ET
    try_qs = qs.filter(currency=_primary_cur)

    # Tip bazli aggregate
    type_summary = {}
    grand_in = Decimal('0')
    grand_out = Decimal('0')
    grand_commission = Decimal('0')

    for acc_type, acc_label in BankAccount.AccountType.choices:
        type_qs = try_qs.filter(account_type=acc_type)

        # FAZ 15 & 18: İptal edilmiş ve onaylanmamış ödemeleri dışla
        agg = Payment.objects.filter(
            bank_account__in=type_qs,
            is_cancelled=False,
            is_approved=True,
        ).aggregate(
            total_in=Coalesce(
                Sum('amount', filter=Q(is_output=False)),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_out=Coalesce(
                Sum('amount', filter=Q(is_output=True)),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_commission=Coalesce(
                Sum('commission_amount'),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_in_net=Coalesce(
                Sum(
                    Coalesce(F('net_amount'), F('amount')),
                    filter=Q(is_output=False),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_out_net=Coalesce(
                Sum(
                    Coalesce(F('net_amount'), F('amount')),
                    filter=Q(is_output=True),
                ),
                Value(Decimal('0')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
        )

        balance_net = agg['total_in_net'] - agg['total_out_net']
        account_count = type_qs.count()

        type_summary[acc_type] = {
            'label': str(acc_label),
            'account_count': account_count,
            'total_in': str(agg['total_in']),
            'total_out': str(agg['total_out']),
            'total_commission': str(agg['total_commission']),
            'balance_net': str(balance_net),
        }

        grand_in += agg['total_in']
        grand_out += agg['total_out']
        grand_commission += agg['total_commission']

    grand_net = grand_in - grand_out - grand_commission

    # 2. DÖVİZ KASALARINI (Giriş, Çıkış ve Net Bakiyeler) AYRI TOPLA
    fx_summary = {}

    # 2A. Spesifik Döviz Kasaları (USD, EUR vb. — birincil ve FX olmayanlar)
    explicit_fx_accounts = qs.exclude(currency__in=[_primary_cur, 'FX'])
    for acc in explicit_fx_accounts:
        code = acc.currency
        if code not in fx_summary:
            fx_summary[code] = {'in': Decimal('0'), 'out': Decimal('0'), 'net': Decimal('0')}

        agg = Payment.objects.filter(
            bank_account=acc, is_cancelled=False, is_approved=True
        ).aggregate(
            t_in=Coalesce(Sum(Coalesce(F('currency_amount'), F('amount')), filter=Q(is_output=False)),
                          Value(Decimal('0')), output_field=DecimalField()),
            t_out=Coalesce(Sum(Coalesce(F('currency_amount'), F('amount')), filter=Q(is_output=True)),
                           Value(Decimal('0')), output_field=DecimalField())
        )
        fx_summary[code]['in'] += agg['t_in']
        fx_summary[code]['out'] += agg['t_out']
        fx_summary[code]['net'] += (agg['t_in'] - agg['t_out'])

    # 2B. Merkez Döviz Kasası (FX)
    fx_central_accounts = qs.filter(currency='FX')
    for acc in fx_central_accounts:
        payments = Payment.objects.filter(
            bank_account=acc, is_cancelled=False, is_approved=True
        ).exclude(currency_amount__isnull=True).exclude(currency_amount=0)

        for p in payments:
            code = _extract_fx_code_from_reference(p.reference) or (
                _guess_fx_code_from_rate(p.exchange_rate) if p.exchange_rate else 'Döviz')
            if code not in fx_summary:
                fx_summary[code] = {'in': Decimal('0'), 'out': Decimal('0'), 'net': Decimal('0')}

            amt = p.currency_amount or Decimal('0')
            if p.is_output:
                fx_summary[code]['out'] += amt
                fx_summary[code]['net'] -= amt
            else:
                fx_summary[code]['in'] += amt
                fx_summary[code]['net'] += amt

    return JsonResponse({
        'result': True,
        'data': {
            'by_type': type_summary,
            'grand': {
                'total_in': str(grand_in),
                'total_out': str(grand_out),
                'total_commission': str(grand_commission),
                'balance_net': str(grand_net),
                'account_count': try_qs.count(),  # TL hesap sayısı
            },
            'fx_summary': {k: {'in': str(v['in']), 'out': str(v['out']), 'net': str(v['net'])} for k, v in
                           fx_summary.items()} if fx_summary else None,
        }
    })


# ──────────────────────────────────────────────────────
# 2. DATATABLE — TÜM HESAPLARI LİSTELE (BAKİYE DAHİL)
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def bank_management_get_all(request):
    """
    DataTables server-side JSON endpoint.
    Her BankAccount icin bakiye annotate eder.
    """
    err = _require_store(request)
    if err:
        return JsonResponse({
            'draw': 0, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': [],
        })

    store = _get_store(request)

    draw = int(request.GET.get('draw', 0))
    length = int(request.GET.get('length', 25))
    start = int(request.GET.get('start', 0))
    search_value = request.GET.get('search[value]', '').strip()
    order_col_idx = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'asc')

    # Tip filtresi (opsiyonel)
    type_filter = request.GET.get('account_type', '').strip().upper()

    qs = get_bank_balance_qs(store)

    # Aktiflik filtresi
    active_only = request.GET.get('active_only', 'true').lower()
    if active_only != 'false':
        qs = qs.filter(is_active=True)

    # Tip filtresi
    valid_types = {c[0] for c in BankAccount.AccountType.choices}
    if type_filter and type_filter in valid_types:
        qs = qs.filter(account_type=type_filter)

    total_records = qs.count()

    # Arama
    if search_value:
        qs = qs.filter(
            Q(name__icontains=search_value) |
            Q(bank_name__icontains=search_value) |
            Q(iban__icontains=search_value)
        )

    filtered_count = qs.count()

    # Siralama
    column_map = {
        0: 'name',
        1: 'bank_name',
        2: 'account_type',
        3: 'iban',
        4: 'total_in',  # annotated field
        5: 'is_active',
    }
    order_col = column_map.get(order_col_idx, 'name')
    if order_dir == 'desc':
        order_col = f'-{order_col}'

    qs = qs.order_by(order_col)
    if length != -1:
        qs = qs[start:start + length]

    # Tip gorsel etiketleri
    type_labels = dict(BankAccount.AccountType.choices)

    data = []
    for acc in qs:
        total_in = acc.total_in or Decimal('0')
        total_out = acc.total_out or Decimal('0')
        balance = total_in - total_out

        data.append({
            'id': str(acc.id),
            'name': acc.name,
            'bank_name': acc.bank_name or '-',
            'account_type': acc.account_type,
            'account_type_display': str(type_labels.get(acc.account_type, acc.account_type)),
            'iban': acc.iban or '-',
            'currency': acc.currency,
            'total_in': str(total_in),
            'total_out': str(total_out),
            'balance': str(balance),
            'balance_is_negative': balance < 0,  # FAZ 16: Negatif bakiye UI bayrağı
            'is_active': acc.is_active,
            'reconciliation_tolerance': str(acc.reconciliation_tolerance),
            'fx_breakdown': _get_fx_breakdown(acc) if acc.currency == 'FX' else None,
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_count,
        'data': data,
    })


# ──────────────────────────────────────────────────────
# 3. DETAY SAYFASI
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def bank_management_detail(request, account_id):
    """Tek banka hesabinin detay sayfasi — ozet kartlari + islem gecmisi."""
    store = _get_store(request)
    if not store:
        return render(request, 'management/banks/detail.html', {
            'title': 'Hesap Bulunamadi',
            'error': 'Kullaniciya bagli magaza bulunamadi.',
        })

    account = get_object_or_404(
        BankAccount, id=account_id, store=store, is_deleted=False,
    )

    summary = get_account_summary(account)

    # FAZ 16: Negatif bakiye UI desteği.
    # Kuyumculuk operasyonlarında kasa bakiyesi geçici olarak negatife düşebilir
    # (ör. sabah virman gelmeden işlem yapılması). Backend'de engelleme yok,
    # yalnızca UI'da kırmızı renk + uyarı badge gösterilir.
    balance_net = summary.get('balance_net', Decimal('0'))
    summary['balance_is_negative'] = balance_net < 0

    # FAZ 20.4: FX kasa döviz kırılımını JSON string olarak template'e aktar
    fx_breakdown_json = ''
    if summary.get('fx_breakdown'):
        fx_breakdown_json = json.dumps(summary['fx_breakdown'])

    return render(request, 'management/banks/detail.html', {
        'title': f'{account.name} — Detay',
        'account': account,
        'summary': summary,
        'fx_breakdown_json': fx_breakdown_json,
    })


# ──────────────────────────────────────────────────────
# 4. DETAY DATATABLE — ÖDEME GEÇMİŞİ
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def bank_management_payments(request, account_id):
    """
    Belirli bir banka hesabina bagli Payment kayitlarini
    DataTables server-side JSON olarak dondurur.

    Filtreler:
        date_from, date_to   — tarih araligi (YYYY-MM-DD)
        payment_type         — CREDIT_CARD | TRANSFER | CASH | COMMISSION
        recon_status         — PENDING | MATCHED | PARTIAL | DISCREPANCY | MANUAL | NOT_REQUIRED
    """
    err = _require_store(request)
    if err:
        return JsonResponse({
            'draw': 0, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': [],
        })

    store = _get_store(request)

    # Hesabin bu magayaya ait oldugunu dogrula
    account = BankAccount.objects.filter(
        id=account_id, store=store, is_deleted=False,
    ).first()
    if not account:
        return JsonResponse({
            'draw': 0, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': [],
        })

    draw = int(request.GET.get('draw', 0))
    length = int(request.GET.get('length', 25))
    start = int(request.GET.get('start', 0))
    search_value = request.GET.get('search[value]', '').strip()
    order_col_idx = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'desc')

    # FAZ 15: İptal edilmiş ödemeleri kapsam dışında tut
    qs = Payment.objects.filter(bank_account=account, is_cancelled=False)
    total_records = qs.count()

    # --- Tarih filtresi ---
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if date_from:
        try:
            qs = qs.filter(date__date__gte=dt_datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(date__date__lte=dt_datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    # --- Odeme tipi filtresi ---
    payment_type = request.GET.get('payment_type', '').strip().upper()
    if payment_type:
        qs = qs.filter(payment_type=payment_type)

    # --- Mutabakat durumu filtresi ---
    recon_status = request.GET.get('recon_status', '').strip().upper()
    if recon_status:
        qs = qs.filter(reconciliation_status=recon_status)

    # --- Arama ---
    if search_value:
        qs = qs.filter(
            Q(process_no__icontains=search_value) |
            Q(reference__icontains=search_value)
        )

    filtered_count = qs.count()

    # --- Siralama ---
    column_map = {
        0: 'date',
        1: 'process_no',
        2: 'payment_type',
        3: 'amount',
        4: 'net_amount',
        5: 'is_output',
        6: 'reconciliation_status',
    }
    order_col = column_map.get(order_col_idx, 'date')
    if order_dir == 'desc':
        order_col = f'-{order_col}'

    qs = qs.select_related('process_group', 'performed_by').order_by(order_col)
    if length != -1:
        qs = qs[start:start + length]

    # Odeme tipi etiketleri
    type_labels = dict(Payment.PAYMENT_TYPE_CHOICES)
    recon_labels = dict(Payment.ReconciliationStatus.choices)

    data = []
    for p in qs:
        # Musteri bilgisini process_group uzerinden al
        customer_name = '-'
        if p.process_group_id:
            proc = p.process_group.process_items.select_related('customer').first()
            if proc and proc.customer:
                customer_name = f"{proc.customer.first_name} {proc.customer.last_name}".strip()

        # FAZ 17: Döviz kasaları için currency_amount ve exchange_rate bilgisi
        _is_foreign = (account.currency and account.currency != 'TRY')
        _display_amount = str(p.currency_amount) if (_is_foreign and p.currency_amount) else str(p.amount)
        _display_net = str(p.currency_amount) if (_is_foreign and p.currency_amount) else (
            str(p.net_amount) if p.net_amount else str(p.amount))

        data.append({
            'id': str(p.id),
            'date': timezone.localtime(p.date).strftime('%d.%m.%Y %H:%M') if p.date else '-',
            'date_raw': p.date.isoformat() if p.date else '',
            'process_no': p.process_no or '-',
            'payment_type': p.payment_type,
            'payment_type_display': str(type_labels.get(p.payment_type, p.payment_type)),
            'amount': _display_amount,
            'amount_eur': str(p.amount),
            'net_amount': _display_net,
            'commission_amount': str(p.commission_amount) if p.commission_amount else '0',
            'commission_rate': str(p.commission_rate_applied) if p.commission_rate_applied else '',
            'currency_amount': str(p.currency_amount) if p.currency_amount else None,
            'exchange_rate': str(p.exchange_rate) if p.exchange_rate else None,
            'is_output': p.is_output,
            'direction': 'OUT' if p.is_output else 'IN',
            'reconciliation_status': p.reconciliation_status,
            'reconciliation_display': str(recon_labels.get(p.reconciliation_status, p.reconciliation_status)),
            'customer_name': customer_name,
            'installment': p.installment,
            'reference': p.reference or '',
            'performed_by_name': (
                f"{p.performed_by.first_name} {p.performed_by.last_name}".strip()
                if p.performed_by else '-'
            ),
            'notes': p.notes or '',
            'is_transfer': bool(p.process_no and (p.process_no.startswith('V-') or p.process_no.startswith('VIR-'))),
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_count,
        'data': data,
    })


# ──────────────────────────────────────────────────────
# 5. EXPORT — PDF / EXCEL
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def bank_management_export(request, account_id):
    """
    Banka hesabi islem dokumunu PDF veya Excel olarak indirir.

    GET parametreleri:
        format    : 'pdf' (default) | 'excel'
        date_from : YYYY-MM-DD
        date_to   : YYYY-MM-DD
    """
    store = _get_store(request)
    if not store:
        return HttpResponse('Magaza bulunamadi.', status=400)

    account = get_object_or_404(
        BankAccount, id=account_id, store=store, is_deleted=False,
    )

    export_format = request.GET.get('format', 'pdf').lower()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    # FAZ 15: İptal edilmiş ödemeleri kapsam dışında tut
    qs = Payment.objects.filter(bank_account=account, is_cancelled=False).select_related('process_group')

    if date_from:
        try:
            qs = qs.filter(date__date__gte=dt_datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(date__date__lte=dt_datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    qs = qs.order_by('-date')

    # Ozet hesapla
    summary = qs.aggregate(
        total_in=Coalesce(
            Sum('amount', filter=Q(is_output=False)),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
        total_out=Coalesce(
            Sum('amount', filter=Q(is_output=True)),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
        total_in_net=Coalesce(
            Sum(
                Coalesce(F('net_amount'), F('amount')),
                filter=Q(is_output=False),
            ),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
        total_out_net=Coalesce(
            Sum(
                Coalesce(F('net_amount'), F('amount')),
                filter=Q(is_output=True),
            ),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
        total_commission=Coalesce(
            Sum('commission_amount'),
            Value(Decimal('0')),
            output_field=DecimalField(max_digits=15, decimal_places=2),
        ),
    )
    summary['balance'] = summary['total_in'] - summary['total_out']
    summary['balance_net'] = summary['total_in_net'] - summary['total_out_net']

    # Satir verileri
    type_labels = dict(Payment.PAYMENT_TYPE_CHOICES)
    rows = []
    for p in qs:
        customer_name = '-'
        if p.process_group_id:
            proc = p.process_group.process_items.select_related('customer').first()
            if proc and proc.customer:
                customer_name = f"{proc.customer.first_name} {proc.customer.last_name}".strip()

        # FAZ 20.4: FX kasası için gerçek döviz birimini tespit et
        _row_currency = account.currency
        _row_amount = p.amount
        _row_net = p.net_amount if p.net_amount else p.amount
        if account.currency == 'FX' and p.currency_amount:
            _row_amount = p.currency_amount
            _row_net = p.currency_amount
            # Reference'dan veya exchange_rate'den döviz kodu tespit
            _row_currency = _extract_fx_code_from_reference(p.reference) or (
                _guess_fx_code_from_rate(p.exchange_rate) if p.exchange_rate else 'Döviz'
            )

        rows.append({
            'date': timezone.localtime(p.date).strftime('%d.%m.%Y %H:%M') if p.date else '-',
            'process_no': p.process_no or '-',
            'payment_type_display': str(type_labels.get(p.payment_type, p.payment_type)),
            'amount': _row_amount,
            'net_amount': _row_net,
            'commission_amount': p.commission_amount or Decimal('0'),
            'direction': 'Cikis' if p.is_output else 'Giris',
            'customer_name': customer_name,
            'currency': _row_currency,  # FAZ 20.4: Gerçek birim
        })

    # FAZ 1 (TZ): timezone.localtime → Berlin saatine çevrilmiş aware datetime;
    # raporun başlık tarihi sistem clock'una değil Django TIME_ZONE'una bağlı.
    now_str = timezone.localtime(timezone.now()).strftime('%d.%m.%Y %H:%M')

    # FAZ 20.4: FX kasa döviz kırılımı
    fx_breakdown = _get_fx_breakdown(account) if account.currency == 'FX' else None

    context = {
        'account': account,
        'summary': summary,
        'rows': rows,
        'date_from': date_from or 'Baslangic',
        'date_to': date_to or 'Bugun',
        'print_datetime': now_str,
        'store': store,
        'fx_breakdown': fx_breakdown,
    }

    if export_format == 'excel':
        return _export_excel(context, rows, account)

    return _export_pdf(request, context)


def _export_pdf(request, context):
    """xhtml2pdf ile PDF olustur ve HttpResponse dondur."""
    from xhtml2pdf import pisa

    html_content = render_to_string(
        'management/banks/report_pdf.html', context, request=request,
    )

    pdf_buffer = io.BytesIO()
    pisa_status = pisa.CreatePDF(
        io.BytesIO(html_content.encode('utf-8')), dest=pdf_buffer,
    )

    if pisa_status.err:
        return HttpResponse('PDF olusturulurken hata olustu.', status=500)

    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    account_name = context['account'].name.replace(' ', '_')
    response['Content-Disposition'] = (
        f'attachment; filename=Banka_Rapor_{account_name}.pdf'
    )
    return response


def _export_excel(context, rows, account):
    """
    Basit CSV/Excel export.
    openpyxl yoksa CSV fallback kullanir.
    """
    try:
        import openpyxl
        return _export_xlsx(context, rows, account)
    except ImportError:
        return _export_csv(context, rows, account)


def _export_xlsx(context, rows, account):
    """openpyxl ile gercek xlsx dosyasi olusturur."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Islem Dokumu'

    # Baslik
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f'{account.name} — Islem Dokumu'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # Tarih araligi
    ws.merge_cells('A2:G2')
    ws['A2'].value = f"Donem: {context['date_from']} - {context['date_to']}"
    ws['A2'].alignment = Alignment(horizontal='center')

    # Ozet satiri
    ws['A4'].value = 'Toplam Giren:'
    ws['B4'].value = float(context['summary']['total_in'])
    ws['B4'].number_format = '#,##0.00'
    ws['C4'].value = 'Toplam Cikan:'
    ws['D4'].value = float(context['summary']['total_out'])
    ws['D4'].number_format = '#,##0.00'
    ws['E4'].value = 'Bakiye:'
    ws['F4'].value = float(context['summary']['balance'])
    ws['F4'].number_format = '#,##0.00'
    for cell in [ws['A4'], ws['C4'], ws['E4']]:
        cell.font = Font(bold=True)

    # Tablo basliklari
    headers = ['Tarih', 'Islem No', 'Odeme Tipi', 'Brut Tutar', 'Net Tutar', 'Yon', 'Musteri']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Veri satirlari
    for row_idx, row in enumerate(rows, 7):
        ws.cell(row=row_idx, column=1, value=row['date']).border = thin_border
        ws.cell(row=row_idx, column=2, value=row['process_no']).border = thin_border
        ws.cell(row=row_idx, column=3, value=row['payment_type_display']).border = thin_border

        amt_cell = ws.cell(row=row_idx, column=4, value=float(row['amount']))
        amt_cell.number_format = '#,##0.00'
        amt_cell.border = thin_border

        net_cell = ws.cell(row=row_idx, column=5, value=float(row['net_amount']))
        net_cell.number_format = '#,##0.00'
        net_cell.border = thin_border

        ws.cell(row=row_idx, column=6, value=row['direction']).border = thin_border
        ws.cell(row=row_idx, column=7, value=row['customer_name']).border = thin_border

    # Kolon genislikleri
    col_widths = [18, 14, 16, 16, 16, 10, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    account_name = account.name.replace(' ', '_')
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename=Banka_Rapor_{account_name}.xlsx'
    )
    return response


def _export_csv(context, rows, account):
    """openpyxl yoksa CSV fallback."""
    import csv

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename=Banka_Rapor_{account.name.replace(" ", "_")}.csv'
    )
    response.write('\ufeff')  # BOM for Excel UTF-8

    writer = csv.writer(response)
    writer.writerow(['Tarih', 'Islem No', 'Odeme Tipi', 'Brut Tutar', 'Net Tutar', 'Yon', 'Musteri'])

    for row in rows:
        writer.writerow([
            row['date'],
            row['process_no'],
            row['payment_type_display'],
            row['amount'],
            row['net_amount'],
            row['direction'],
            row['customer_name'],
        ])

    return response


# ──────────────────────────────────────────────────────
# 6. KASA ARASI TRANSFER (VİRMAN)
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def bank_transfer_view(request):
    if request.method != 'POST':
        return JsonResponse({'result': False, 'msg': 'Gecersiz istek.'})

    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Kullaniciya bagli magaza bulunamadi.'})

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'result': False, 'msg': 'Gecersiz JSON verisi.'})

    from_id = body.get('from_account_id', '').strip()
    to_id = body.get('to_account_id', '').strip()
    note = (body.get('note') or '').strip()

    if not from_id or not to_id:
        return JsonResponse({'result': False, 'msg': 'Gonderen ve alici kasa zorunludur.'})

    if from_id == to_id:
        return JsonResponse({'result': False, 'msg': 'Gonderen ve alici kasa ayni olamaz.'})

    try:
        amount = Decimal(str(body.get('amount', '0')))
    except (InvalidOperation, ValueError):
        return JsonResponse({'result': False, 'msg': 'Gecersiz tutar.'})

    if amount <= 0:
        return JsonResponse({'result': False, 'msg': 'Tutar sifirdan buyuk olmalidir.'})

    # Hesaplari dogrula
    from_acc = BankAccount.objects.filter(
        id=from_id, store=store, is_deleted=False, is_active=True,
    ).first()
    to_acc = BankAccount.objects.filter(
        id=to_id, store=store, is_deleted=False, is_active=True,
    ).first()

    if not from_acc:
        return JsonResponse({'result': False, 'msg': 'Gonderen kasa bulunamadi veya aktif degil.'})
    if not to_acc:
        return JsonResponse({'result': False, 'msg': 'Alici kasa bulunamadi veya aktif degil.'})

    # YENİ: Para birimi eşleşme kontrolü (Sadece aynı birimler arası transfer)
    from_currency = getattr(from_acc, 'currency', 'TRY') or 'TRY'
    to_currency = getattr(to_acc, 'currency', 'TRY') or 'TRY'

    if from_currency != to_currency:
        return JsonResponse({
            'result': False,
            'msg': f'Farklı para birimleri arasında transfere izin verilmiyor ({from_currency} ➔ {to_currency}).'
        })

    # process_no veritabanında max_length=15 karakter.
    # V-YYMMDDHHMMSS formatı = 14 karakter (güvenli).
    # FAZ 1 (TZ): timezone.localtime → Berlin saatine göre referans üretilir;
    # transfer_ref'in saat dilimi tutarlılığı raporlarla uyumlu kalır.
    short_timestamp = timezone.localtime(timezone.now()).strftime("%y%m%d%H%M%S")
    transfer_ref = f'V-{short_timestamp}'
    transfer_label = note or f'{from_acc.name} -> {to_acc.name}'

    # Döviz kasaları için ek veritabanı alanları
    _extra_fields = {}
    if from_currency != 'TRY':
        _extra_fields['currency_amount'] = amount
        _extra_fields['exchange_rate'] = Decimal('1')

    # FAZ 1 (TZ): Aware datetime — USE_TZ=True moduyla uyumlu, ORM'e doğrudan
    # yazılabilir, naive/aware karışıklığı oluşmaz.
    now = timezone.now()

    with transaction.atomic():
        # Cikis kaydi (gonderen kasadan para cikiyor)
        Payment.objects.create(
            process_no=transfer_ref,
            payment_type='CASH',
            amount=amount,
            is_output=True,
            bank_account=from_acc,
            reference=transfer_label,
            reconciliation_status='NOT_REQUIRED',
            date=now,
            performed_by=request.user,
            notes=note or '',
            **_extra_fields,
        )

        # Giris kaydi (alici kasaya para giriyor)
        Payment.objects.create(
            process_no=transfer_ref,
            payment_type='CASH',
            amount=amount,
            is_output=False,
            bank_account=to_acc,
            reference=transfer_label,
            reconciliation_status='NOT_REQUIRED',
            date=now,
            performed_by=request.user,
            notes=note or '',
            **_extra_fields,
        )

    return JsonResponse({
        'result': True,
        'msg': f'{amount:.2f} {from_currency} basariyla transfer edildi.',
        'reference': transfer_ref,
    })


# ──────────────────────────────────────────────────────
# 6b. TRANSFER DETAY (Modal icin)
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def bank_transfer_detail(request, payment_id):
    """
    Tek bir Payment ID'si uzerinden transfer (virman) detayini dondurur.
    Ayni process_no'ya sahip karsi kasayi bularak tam bilgi sunar.

    GET /bank-management/transfer-detail/<uuid:payment_id>/
    """
    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Magaza bulunamadi.'})

    payment = Payment.objects.select_related(
        'bank_account', 'performed_by',
    ).filter(id=payment_id).first()

    if not payment:
        return JsonResponse({'result': False, 'msg': 'Odeme kaydi bulunamadi.'})

    if not payment.process_no or not (payment.process_no.startswith('V-') or payment.process_no.startswith('VIR-')):
        return JsonResponse({'result': False, 'msg': 'Bu kayit bir transfer islemi degildir.'})

    # Karsi kasayi bul: ayni process_no, farkli yon
    counterpart = Payment.objects.select_related('bank_account').filter(
        process_no=payment.process_no,
        is_output=not payment.is_output,
        is_cancelled=False,
    ).first()

    # Gondereni ve aliciyi belirle
    if payment.is_output:
        from_name = payment.bank_account.name if payment.bank_account else '-'
        to_name = counterpart.bank_account.name if counterpart and counterpart.bank_account else '-'
    else:
        from_name = counterpart.bank_account.name if counterpart and counterpart.bank_account else '-'
        to_name = payment.bank_account.name if payment.bank_account else '-'

    return JsonResponse({
        'result': True,
        'data': {
            'reference': payment.process_no,
            'from_account': from_name,
            'to_account': to_name,
            'amount': str(payment.amount),
            'date': timezone.localtime(payment.date).strftime('%d.%m.%Y %H:%M') if payment.date else '-',
            'performed_by': (
                f"{payment.performed_by.first_name} {payment.performed_by.last_name}".strip()
                if payment.performed_by else '-'
            ),
            'notes': payment.notes or '-',
        },
    })


# ──────────────────────────────────────────────────────
# 7. GÜNLÜK KAPANIŞ (Z-RAPORU / FİZİKSEL SAYIM)
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def daily_close_view(request):
    """
    Gun sonu kapanisi olusturur.

    POST body:
        bank_account_id : UUID
        date            : YYYY-MM-DD (opsiyonel, default=bugun)
        physical_count  : Decimal (fiziksel sayim tutari)
        note            : str (opsiyonel)
    """
    if request.method != 'POST':
        return JsonResponse({'result': False, 'msg': 'Gecersiz istek.'})

    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Kullaniciya bagli magaza bulunamadi.'})

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'result': False, 'msg': 'Gecersiz JSON verisi.'})

    account_id = body.get('bank_account_id', '').strip()
    if not account_id:
        return JsonResponse({'result': False, 'msg': 'Kasa secimi zorunludur.'})

    account = BankAccount.objects.filter(
        id=account_id, store=store, is_deleted=False,
    ).first()
    if not account:
        return JsonResponse({'result': False, 'msg': 'Kasa bulunamadi.'})

    # Tarih
    date_str = body.get('date', '').strip()
    if date_str:
        try:
            close_date = dt_datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'result': False, 'msg': 'Gecersiz tarih formati (YYYY-MM-DD).'})
    else:
        # FAZ 1 (TZ): Berlin saatine göre tarih (UTC midnight kayması engellenir).
        close_date = timezone.localdate()

    # Fiziksel sayim
    try:
        physical_count = Decimal(str(body.get('physical_count', '0')))
    except (InvalidOperation, ValueError):
        return JsonResponse({'result': False, 'msg': 'Gecersiz fiziksel sayim tutari.'})

    note = (body.get('note') or '').strip()

    # Sistem bakiyesini hesapla
    summary = get_account_summary(account)
    system_balance = summary['balance_net']

    # Fark
    difference = physical_count - system_balance

    # Ayni gun + ayni kasa icin tekrar kapanisi engelle
    from apps.banking.models import DailyCashClose
    existing = DailyCashClose.objects.filter(
        store=store, bank_account=account, date=close_date,
    ).first()
    if existing:
        return JsonResponse({
            'result': False,
            'msg': f'{close_date.strftime("%d.%m.%Y")} tarihi icin bu kasada zaten kapanıs yapilmis.',
        })

    DailyCashClose.objects.create(
        store=store,
        bank_account=account,
        date=close_date,
        system_balance=system_balance,
        physical_count=physical_count,
        difference=difference,
        note=note,
        closed_by=request.user,
    )

    return JsonResponse({
        'result': True,
        'msg': f'Gun sonu kapanisi kaydedildi. Fark: {difference:.2f} TL',
        'data': {
            'date': close_date.strftime('%d.%m.%Y'),
            'system_balance': str(system_balance),
            'physical_count': str(physical_count),
            'difference': str(difference),
        }
    })


@login_required(login_url='login')
def daily_close_list(request, account_id):
    """
    Belirli bir kasanin gecmis kapanislarini JSON olarak dondurur.
    """
    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'data': []})

    from apps.banking.models import DailyCashClose
    closes = DailyCashClose.objects.filter(
        store=store, bank_account_id=account_id,
    ).order_by('-date')[:30]

    data = []
    for c in closes:
        data.append({
            'id': str(c.id),
            'date': c.date.strftime('%d.%m.%Y'),
            'system_balance': str(c.system_balance),
            'physical_count': str(c.physical_count),
            'difference': str(c.difference),
            'note': c.note or '',
            'closed_by': str(c.closed_by) if c.closed_by else '-',
            'created_at': timezone.localtime(c.created_at).strftime('%d.%m.%Y %H:%M') if c.created_at else '',
        })

    return JsonResponse({'result': True, 'data': data})


# ──────────────────────────────────────────────────────
# FAZ 19: BAKİYE DÜZELTME / AÇILIŞ FİŞİ
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def adjustment_payment(request):
    """
    FAZ 19: Kasa açılış bakiyesi veya bakiye düzeltme fişi.

    Bu işlem:
    - Kasanın bakiyesini doğrudan artırır (+) veya azaltır (-)
    - Payment tablosuna payment_type='ADJUSTMENT' olarak kaydedilir
    - Günlük satış/ciro raporlarına DAHİL OLMAZ
    - Karşılığında mal hareketi veya stok değişikliği OLMAZ

    POST parametreleri:
        account_id  — BankAccount UUID
        amount      — Tutar (pozitif: giriş, negatif: çıkış)
        notes       — Açıklama (ör. "Sisteme geçiş sermayesi")
        currency    — Döviz cinsi (FX kasaları için: USD, EUR vb. TRY kasaları için opsiyonel)
    """
    if request.method != 'POST':
        return JsonResponse({'result': False, 'msg': 'Geçersiz istek.'}, status=405)

    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Mağaza bulunamadı.'})

    account_id = request.POST.get('account_id', '').strip()
    amount_str = request.POST.get('amount', '').strip()
    notes = request.POST.get('notes', '').strip()
    adj_currency = request.POST.get('currency', '').strip().upper() or None

    if not account_id:
        return JsonResponse({'result': False, 'msg': 'Hesap bilgisi eksik.'})
    if not amount_str:
        return JsonResponse({'result': False, 'msg': 'Tutar girilmelidir.'})

    try:
        amount = Decimal(amount_str.replace(',', '.'))
    except Exception:
        return JsonResponse({'result': False, 'msg': 'Geçersiz tutar formatı.'})

    if amount == 0:
        return JsonResponse({'result': False, 'msg': 'Tutar 0 olamaz.'})

    account = BankAccount.objects.filter(
        id=account_id, store=store, is_deleted=False,
    ).first()
    if not account:
        return JsonResponse({'result': False, 'msg': 'Hesap bulunamadı.'})

    # Pozitif = GİRİŞ (kasaya para koyma / açılış sermayesi)
    # Negatif = ÇIKIŞ (kasadan para çekme / düzeltme)
    is_output = (amount < 0)
    abs_amount = abs(amount)

    # FAZ 20.2: Merkez Döviz Kasası (currency='FX') ise currency_amount ve exchange_rate ayarla
    _adj_extra = {}
    acct_currency = getattr(account, 'currency', 'TRY') or 'TRY'

    if acct_currency == 'FX' and adj_currency and adj_currency != 'TRY':
        _adj_extra['currency_amount'] = abs_amount
        # FAZ 13.3: Sentinel rate SSOT'tan alınır (services.FX_SENTINEL_MAP).
        _adj_extra['exchange_rate'] = FX_SENTINEL_MAP.get(adj_currency, Decimal('0.09'))
    elif acct_currency != 'TRY' and acct_currency != 'FX':
        _adj_extra['currency_amount'] = abs_amount
        _adj_extra['exchange_rate'] = Decimal('1')

    # Reference'a currency kodu prefix olarak yazılır: "[EUR] Açıklama"
    _ref_prefix = f'[{adj_currency}] ' if (acct_currency == 'FX' and adj_currency) else ''
    _ref_notes = notes[:90] if notes else 'Bakiye Düzeltme'

    Payment.objects.create(
        process_no=None,
        payment_type='ADJUSTMENT',
        amount=abs_amount,
        is_output=is_output,
        bank_account=account,
        reconciliation_status=Payment.ReconciliationStatus.NOT_REQUIRED,
        is_approved=True,
        reference=f'{_ref_prefix}{_ref_notes}',
        performed_by=request.user,
        **_adj_extra,
    )

    direction = 'ÇIKIŞ' if is_output else 'GİRİŞ'
    log.info(
        "FAZ19 ADJUSTMENT: account=%s amount=%s direction=%s notes=%s user=%s",
        account.name, abs_amount, direction, notes, request.user.username,
    )

    return JsonResponse({
        'result': True,
        'msg': f'{abs_amount:.2f} {account.currency} {direction} başarıyla kaydedildi.',
    })


# ════════════════════════════════════════════════════════════════════════════
# FAZ 31 / BUG-3 — MANUEL GİDER GİRİŞİ (2026-05-01)
# ════════════════════════════════════════════════════════════════════════════
#
# Müşteri şikayeti: "Kasada gider düşüremiyorum, gider düşecek ekran yok."
#
# Önceki durum:
#   - IncomeExpenseLedger ve CashboxLedger.EXPENSE modelleri MEVCUTTU
#   - Ancak bunlara YAZAN endpoint yoktu (sadece tahsilat/iskonto/kur farkı
#     gibi otomatik akışlar yazıyordu)
#   - Manuel "kira ödedim", "fatura ödedim", "personel maaşı" gibi gider
#     girişi için ekran/endpoint hiç yazılmamıştı
#
# Bu endpoint:
#   1. Payment (is_output=True, payment_type='EXPENSE') yazar →
#      kasa bakiyesi anında düşer (get_bank_balance_qs Payment'a dayanır)
#   2. CashboxLedger.EXPENSE yazar → audit trail
#   3. IncomeExpenseLedger.OTHER_EXPENSE yazar → P&L raporlamasına dahil
#
# Tüm yazımlar TEK atomik blokta. Bu üçünden biri patlasa hiçbiri yazılmaz.
# Mevcut adjustment_payment endpoint'ine DOKUNULMADI; o "açılış/sayım farkı"
# için kalır, bu yeni endpoint "gerçek gider" için kullanılır.
# ════════════════════════════════════════════════════════════════════════════

@login_required(login_url='login')
@transaction.atomic
def manual_expense(request):
    """Kasadan manuel gider düşümü.

    POST parametreleri:
        account_id          — BankAccount UUID (zorunlu)
        amount              — Gider tutarı (zorunlu, pozitif)
        currency            — Döviz cinsi (FX kasalarda zorunlu, diğerlerinde opsiyonel)
        description         — Açıklama (zorunlu, ör: "Kira", "Personel maaşı")
        expense_category_id — ExpenseCategory UUID (FAZ 61, OPSİYONEL)
                               Geçerli kategori varsa IncomeExpenseLedger.expense_category
                               alanına yazılır; yoksa NULL kalır (geriye uyum).

    Yan etkiler:
      • Payment(is_output=True, payment_type='EXPENSE') — kasa bakiyesi düşer
      • CashboxLedger.EXPENSE                              — audit/iz
      • IncomeExpenseLedger.OTHER_EXPENSE                  — gelir/gider defteri
                                                             (+ expense_category FK opsiyonel)
    """
    if request.method != 'POST':
        return JsonResponse({'result': False, 'msg': 'Geçersiz istek.'}, status=405)

    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Mağaza bulunamadı.'})

    account_id = request.POST.get('account_id', '').strip()
    amount_str = request.POST.get('amount', '').strip()
    description = request.POST.get('description', '').strip()
    exp_currency = request.POST.get('currency', '').strip().upper() or None
    category_id = request.POST.get('expense_category_id', '').strip() or None

    if not account_id:
        return JsonResponse({'result': False, 'msg': 'Hesap bilgisi eksik.'})
    if not amount_str:
        return JsonResponse({'result': False, 'msg': 'Tutar girilmelidir.'})
    if not description:
        return JsonResponse({'result': False, 'msg': 'Gider açıklaması zorunludur.'})

    try:
        amount = Decimal(amount_str.replace(',', '.'))
    except (InvalidOperation, ValueError):
        return JsonResponse({'result': False, 'msg': 'Geçersiz tutar formatı.'})

    if amount <= 0:
        return JsonResponse({'result': False, 'msg': 'Gider tutarı pozitif olmalıdır.'})

    account = BankAccount.objects.filter(
        id=account_id, store=store, is_deleted=False, is_active=True,
    ).first()
    if not account:
        return JsonResponse({'result': False, 'msg': 'Hesap bulunamadı veya aktif değil.'})

    # FAZ 61: Opsiyonel kategori doğrulaması (geriye uyumlu — eski çağrılar bypass eder).
    expense_category = None
    if category_id:
        from apps.banking.models import ExpenseCategory
        expense_category = ExpenseCategory.objects.filter(
            id=category_id, store=store, is_active=True,
        ).first()
        if not expense_category:
            return JsonResponse({
                'result': False,
                'msg': 'Seçilen gider kategorisi bulunamadı veya aktif değil.',
            })

    acct_currency = getattr(account, 'currency', 'TRY') or 'TRY'

    # ────────────────────────────────────────────────────────────────
    # 1) Payment kaydı (is_output=True → kasa bakiyesi anında düşer)
    # ────────────────────────────────────────────────────────────────
    _pay_extra = {}
    if acct_currency == 'FX' and exp_currency and exp_currency != 'TRY':
        _pay_extra['currency_amount'] = amount
        _pay_extra['exchange_rate'] = FX_SENTINEL_MAP.get(exp_currency, Decimal('0.09'))
    elif acct_currency != 'TRY' and acct_currency != 'FX':
        _pay_extra['currency_amount'] = amount
        _pay_extra['exchange_rate'] = Decimal('1')

    _ref_prefix = f'[{exp_currency}] ' if (acct_currency == 'FX' and exp_currency) else ''
    _ref_text = f'{_ref_prefix}GIDER: {description}'[:100]

    payment = Payment.objects.create(
        process_no=None,
        payment_type='EXPENSE',
        amount=amount,
        is_output=True,                                  # kasadan ÇIKIŞ
        bank_account=account,
        reconciliation_status=Payment.ReconciliationStatus.NOT_REQUIRED,
        is_approved=True,
        reference=_ref_text,
        performed_by=request.user,
        **_pay_extra,
    )

    # ────────────────────────────────────────────────────────────────
    # 2) CashboxLedger.EXPENSE (audit trail)
    # ────────────────────────────────────────────────────────────────
    # TL eşdeğeri:
    #   • TRY kasa → amount aynen
    #   • FX kasa → amount sentinel rate'i ham olduğu için TL eşdeğeri
    #     için anlık has kuru / döviz kuru çağırmak gerekir; FAZ 14 şu an
    #     bu kasa hareketi için TL eşdeğerini approximate olarak amount kabul ediyor.
    #     (Mevcut adjustment_payment de aynı yaklaşımı kullanıyor.)
    from apps.banking.models import CashboxLedger, IncomeExpenseLedger

    cb_currency_choice = (
        exp_currency if (acct_currency == 'FX' and exp_currency) else acct_currency
    )
    if cb_currency_choice not in ('TRY', 'USD', 'EUR', 'GBP', 'HS'):
        cb_currency_choice = 'TRY'

    try:
        prior_balance = account.get_balance(currency=cb_currency_choice)
    except Exception:
        prior_balance = Decimal('0')
    new_balance = (Decimal(str(prior_balance)) - amount).quantize(Decimal('0.01'))

    cashbox_entry = CashboxLedger.objects.create(
        cashbox=account,
        store=store,
        movement_type=CashboxLedger.MovementType.EXPENSE,
        amount=amount.quantize(Decimal('0.01')),
        currency=cb_currency_choice,
        amount_eur_equivalent=amount.quantize(Decimal('0.01')),
        exchange_rate=_pay_extra.get('exchange_rate'),
        balance_snapshot=new_balance,
        related_payment=payment,
        process_no=None,
        description=f'Manuel gider — {description}'[:255],
        created_by=request.user,
        ip_address=request.META.get('REMOTE_ADDR') or None,
        user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:512],
    )

    # ────────────────────────────────────────────────────────────────
    # 3) IncomeExpenseLedger.OTHER_EXPENSE (P&L kaydı)
    # ────────────────────────────────────────────────────────────────
    # Anlık has kuru — FX_LOSS_EXPENSE gibi altın bazlı zarar değil; pure TL gider.
    # amount_hs ve exchange_rate_eur sıfır kalır (modelin default'u 0).
    try:
        IncomeExpenseLedger.objects.create(
            store=store,
            entry_type=IncomeExpenseLedger.EntryType.OTHER_EXPENSE,
            amount_eur=amount.quantize(Decimal('0.01')),
            amount_hs=Decimal('0'),
            exchange_rate_eur=Decimal('0'),
            related_payment=payment,
            expense_category=expense_category,  # FAZ 61: opsiyonel kategori bağı
            description=description[:255],
            created_by=request.user,
            ip_address=request.META.get('REMOTE_ADDR') or None,
            user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:512],
        )
    except Exception as exc:
        # IncomeExpenseLedger yazımı başarısız olursa atomik blok geri sarılır.
        # Loglanır ve kullanıcıya hata döner.
        log.exception("manual_expense: IncomeExpenseLedger yazımı başarısız: %s", exc)
        raise

    log.info(
        "FAZ31 MANUAL_EXPENSE: account=%s amount=%s currency=%s description=%s "
        "category=%s user=%s",
        account.name, amount, cb_currency_choice, description,
        getattr(expense_category, 'name', '-'), request.user.username,
    )

    return JsonResponse({
        'result': True,
        'msg': f'{amount:.2f} {cb_currency_choice} gider olarak kaydedildi.',
    })


# ──────────────────────────────────────────────────────
# FAZ 18.2: HIZLI KASA OLUŞTURMA (Quick Create)
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def quick_create_account(request):
    """
    FAZ 18.2: Ödeme akışı sırasında eksik kasa/POS/banka hesabını
    anında oluşturmak için hafif JSON endpoint.

    POST parametreleri:
        name         — Hesap adı (zorunlu)
        account_type — CASH | POS | BANK (zorunlu)
        currency     — TRY | USD | EUR vb. (opsiyonel, default TRY)

    Returns:
        { result: true, id: "<uuid>", name: "...", account_type: "...", currency: "..." }
    """
    if request.method != 'POST':
        return JsonResponse({'result': False, 'msg': 'Geçersiz istek.'}, status=405)

    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Mağaza bulunamadı.'})

    name = (request.POST.get('name') or '').strip()
    account_type = (request.POST.get('account_type') or '').strip().upper()
    currency = (request.POST.get('currency') or 'TRY').strip().upper()

    if not name:
        return JsonResponse({'result': False, 'msg': 'Hesap adı zorunludur.'})

    valid_types = {c[0] for c in BankAccount.AccountType.choices}
    if account_type not in valid_types:
        return JsonResponse({'result': False, 'msg': f'Geçersiz hesap tipi: {account_type}'})

    acc = BankAccount.objects.create(
        store=store,
        name=name,
        account_type=account_type,
        currency=currency,
        is_active=True,
    )

    log.info(
        "FAZ18.2 QUICK_CREATE: store=%s type=%s currency=%s name=%s id=%s",
        store, account_type, currency, name, acc.id,
    )

    return JsonResponse({
        'result': True,
        'id': str(acc.id),
        'name': acc.name,
        'account_type': acc.account_type,
        'currency': acc.currency,
        'msg': f'"{name}" hesabı oluşturuldu.',
    })


# ──────────────────────────────────────────────────────
# FAZ 18: BEKLEYEN İŞLEMLER (Onaylı Kasa)
# ──────────────────────────────────────────────────────

@login_required(login_url='login')
def pending_payments_list(request):
    """
    FAZ 20: Onay bekleyen (is_approved=False) Payment kayıtlarını
    process_no bazlı GRUPLANMIŞ olarak DataTables JSON döndürür.

    Aynı process_no'ya sahip ödemeler tek satırda gösterilir.
    Satırda her ödemenin detayı (kasa, yön, tutar) listelenir.
    """
    store = _get_store(request)
    if not store:
        return JsonResponse({
            'draw': 0, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': [],
        })

    draw = int(request.GET.get('draw', 0))
    length = int(request.GET.get('length', 25))
    start = int(request.GET.get('start', 0))
    search_value = request.GET.get('search[value]', '').strip()

    qs = Payment.objects.filter(
        bank_account__store=store,
        is_cancelled=False,
        is_approved=False,
    ).select_related('bank_account')

    if search_value:
        qs = qs.filter(
            Q(process_no__icontains=search_value) |
            Q(bank_account__name__icontains=search_value)
        )

    # FAZ 20: process_no bazlı grupla
    from django.db.models import Count, Max, Min
    from collections import OrderedDict

    all_payments = qs.order_by('-date')

    # Process_no bazlı grupla
    grouped = OrderedDict()
    for p in all_payments:
        key = p.process_no or str(p.id)  # process_no yoksa (ADJUSTMENT vb.) tek satır
        if key not in grouped:
            grouped[key] = {
                'process_no': p.process_no or '-',
                'date': timezone.localtime(p.date).strftime('%d.%m.%Y %H:%M') if p.date else '-',
                'payments': [],
            }
        cur = p.bank_account.currency if p.bank_account else 'TRY'
        amt = float(p.currency_amount) if (cur != 'TRY' and p.currency_amount) else float(p.amount)
        grouped[key]['payments'].append({
            'bank_name': p.bank_account.name if p.bank_account else '-',
            'currency': cur,
            'amount': amt,
            'direction': 'ÇIKIŞ' if p.is_output else 'GİRİŞ',
            'is_output': p.is_output,
            'payment_type': p.payment_type,
        })

    total_records = len(grouped)
    filtered_count = total_records

    # Sayfalama
    group_list = list(grouped.values())
    if length != -1:
        group_list = group_list[start:start + length]

    type_labels = dict(Payment.PAYMENT_TYPE_CHOICES)
    data = []
    for g in group_list:
        # Özet: her ödeme satırını detay olarak birleştir
        detail_parts = []
        for pay in g['payments']:
            direction_icon = '↓' if not pay['is_output'] else '↑'
            detail_parts.append(
                f"{direction_icon} {pay['amount']:.2f} {pay['currency']} ({pay['bank_name']})"
            )
        detail_html = '<br>'.join(detail_parts)

        # Toplam TL tutarını göster (birincil)
        primary_type = g['payments'][0]['payment_type'] if g['payments'] else 'CASH'

        data.append({
            'process_no': g['process_no'],
            'date': g['date'],
            'payment_type_display': str(type_labels.get(primary_type, primary_type)),
            'detail_html': detail_html,
            'payment_count': len(g['payments']),
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_count,
        'data': data,
    })


@login_required(login_url='login')
@transaction.atomic
def approve_payment(request):
    """
    FAZ 20: Process_no bazlı TOPLU onay.
    Aynı process_no'ya sahip TÜM bekleyen ödemeleri aynı anda onaylar.
    Böylece döviz bozma gibi çift taraflı kayıtlar birlikte onaylanır.

    POST: { process_no: str }
    """
    if request.method != 'POST':
        return JsonResponse({'result': False, 'msg': 'Geçersiz istek.'}, status=405)

    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Mağaza bulunamadı.'})

    process_no = request.POST.get('process_no', '').strip()
    if not process_no:
        return JsonResponse({'result': False, 'msg': 'İşlem numarası gerekli.'})

    # FAZ 20: transaction.atomic dekoratör ile — select_for_update güvenle çalışır
    payments = Payment.objects.select_for_update().filter(
        process_no=process_no,
        bank_account__store=store,
        is_cancelled=False,
        is_approved=False,
    )

    count = payments.count()
    if count == 0:
        return JsonResponse({'result': False, 'msg': 'Bu işlem numarasında bekleyen ödeme bulunamadı.'})

    payments.update(is_approved=True)

    log.info(
        "FAZ20 TOPLU_ONAY: process_no=%s count=%d onaylayan=%s",
        process_no, count, request.user,
    )
    return JsonResponse({
        'result': True,
        'msg': f'{count} adet ödeme kaydı onaylandı (İşlem No: {process_no}).',
    })


@login_required(login_url='login')
@transaction.atomic
def reject_payment(request):
    """
    FAZ 20: Process_no bazlı TOPLU red.
    Aynı process_no'ya sahip TÜM bekleyen ödemeleri aynı anda iptal eder.

    POST: { process_no: str }
    """
    if request.method != 'POST':
        return JsonResponse({'result': False, 'msg': 'Geçersiz istek.'}, status=405)

    store = _get_store(request)
    if not store:
        return JsonResponse({'result': False, 'msg': 'Mağaza bulunamadı.'})

    process_no = request.POST.get('process_no', '').strip()
    if not process_no:
        return JsonResponse({'result': False, 'msg': 'İşlem numarası gerekli.'})

    # FAZ 20: transaction.atomic dekoratör ile — select_for_update güvenle çalışır
    payments = Payment.objects.select_for_update().filter(
        process_no=process_no,
        bank_account__store=store,
        is_cancelled=False,
        is_approved=False,
    )

    count = payments.count()
    if count == 0:
        return JsonResponse({'result': False, 'msg': 'Bu işlem numarasında bekleyen ödeme bulunamadı.'})

    payments.update(is_cancelled=True, cancelled_at=timezone.now())

    log.info(
        "FAZ20 TOPLU_RED: process_no=%s count=%d reddeden=%s",
        process_no, count, request.user,
    )
    return JsonResponse({
        'result': True,
        'msg': f'{count} adet ödeme kaydı reddedildi (İşlem No: {process_no}).',
    })
