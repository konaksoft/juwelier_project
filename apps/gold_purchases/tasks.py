# ============================================================================
# DOSYA: apps/gold_purchases/tasks.py
# AÇIKLAMA: Barkodlu Ürünler — ZIP Tam Yedekleme Celery Task'ı
#
# Kullanıcı "Tam Yedek Al" butonuna bastığında bu task arka planda çalışır.
# Üretilen ZIP arşivi:
#   yedek_YYYYMMDD_HHMM.zip
#   ├── urunler.xlsx         (tüm metin verileri + RFID + görsel dosya adı)
#   └── gorseller/           (ürün görselleri orijinal formatında)
#       ├── YZK0001.webp
#       └── ...
#
# Tamamlanan dosya GeneratedReports modeline kaydedilir ve frontend
# polling ile indirme linkini alır.
# ============================================================================

import logging
import os
import shutil
import tempfile
import zipfile

from celery import shared_task
from django.conf import settings
from django.core.files.base import ContentFile
from django.db.models import Q
from django.utils import timezone

log = logging.getLogger(__name__)


def _mark_failed(task_id: str, error: BaseException) -> None:
    """
    GeneratedReports kaydını FAILED olarak işaretler.
    Ana try-except bloğunun dışında çağrılabilmesi için bağımsız fonksiyon.
    """
    try:
        from apps.dashboard.models import GeneratedReports
        GeneratedReports.objects.filter(task_id=task_id).update(
            status='FAILED',
            error_message=str(error)[:500],
        )
    except Exception as inner_exc:
        log.error(f"FAILED durumu kaydedilemedi (task_id={task_id}): {inner_exc}")


@shared_task(bind=True, max_retries=0)
def generate_backup_zip_task(self, store_id, user_id):
    """
    Barkodlu ürünlerin tam yedeğini ZIP arşivi olarak oluşturur.

    İçerik:
        - urunler.xlsx: 12 sütunlu Excel (mevcut 10 + RFID Kodu + Görsel Dosya Adı)
        - gorseller/: Ürün görselleri {barkod}.{uzantı} formatında

    Sonuç GeneratedReports modeline kaydedilir.
    """
    # task_id, try bloğu dışında tanımlanıyor; except/finally her durumda erişebilsin.
    task_id = self.request.id or 'UNKNOWN'
    tmp_dir = None

    try:
        # ── Gerekli import'lar (fonksiyon seviyesinde — circular import riski önlenir) ──
        from apps.dashboard.models import GeneratedReports
        from apps.gold_purchases.models import GoldPurchases
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        log.info(f"Backup ZIP task başladı: task_id={task_id}, store_id={store_id}")

        # ── 1. GeneratedReports kaydını PENDING olarak oluştur/güncelle ──
        report, _ = GeneratedReports.objects.update_or_create(
            task_id=task_id,
            defaults={'report_type': 'BACKUP_ZIP', 'status': 'PENDING'},
        )

        # ── 2. Veriyi çek ──
        qs = (
            GoldPurchases.objects
            .filter(is_deleted=False, store_id=store_id)
            .select_related('product', 'supplier')
            .order_by('created_on')
        )
        total_count = qs.count()
        log.info(f"İşlenecek kayıt: {total_count}")

        # ── 3. Geçici klasör oluştur ──
        tmp_dir = tempfile.mkdtemp(prefix='kp_backup_')
        gorseller_dir = os.path.join(tmp_dir, 'gorseller')
        os.makedirs(gorseller_dir, exist_ok=True)

        # ── 4. openpyxl ile urunler.xlsx oluştur ──
        wb = Workbook()
        ws = wb.active
        ws.title = 'Rapor'

        headers = [
            'Barkod',
            'Tedarikçi',
            'Takı Tipi',
            'Gram',
            'Maliyet (Has)',
            'Satış (Has)',
            'Kar (%)',
            'Statü',
            'Tarih',
            'Durum',
            'RFID Kodu',
            'Görsel Dosya Adı',
        ]
        ws.append(headers)

        def f3(x):
            try:
                return ('%.3f' % float(x)) if x is not None else '0.000'
            except Exception:
                return '0.000'

        def f2(x):
            try:
                return ('%.2f' % float(x)) if x is not None else ''
            except Exception:
                return ''

        images_copied = 0

        for r in qs.iterator(chunk_size=200):
            prod = r.product
            if not prod:
                continue

            is_sold = (r.is_status is False) or (getattr(prod, 'is_completed', False) is True)
            stat_text = 'Satıldı' if is_sold else 'Tezgahta'
            active_text = 'Aktif' if prod.is_active else 'Pasif'

            created_on = r.created_on
            created_text = ''
            if created_on:
                try:
                    created_text = created_on.astimezone(
                        timezone.get_current_timezone()
                    ).strftime('%d/%m/%Y %H:%M')
                except Exception:
                    created_text = str(created_on)

            # Görsel dosya adı ve kopyalama
            gorsel_dosya_adi = ''
            if prod.image and prod.image.name and prod.image.name != 'default/default.png':
                src_path = os.path.join(settings.MEDIA_ROOT, prod.image.name)
                if os.path.isfile(src_path):
                    _, ext = os.path.splitext(prod.image.name)
                    barcode_safe = (prod.barcode or str(prod.id)).replace('/', '_')
                    dest_name = f"{barcode_safe}{ext}"
                    dest_path = os.path.join(gorseller_dir, dest_name)
                    try:
                        shutil.copy2(src_path, dest_path)
                        gorsel_dosya_adi = dest_name
                        images_copied += 1
                    except Exception as copy_err:
                        log.warning(f"Görsel kopyalama hatası ({prod.barcode}): {copy_err}")

            supplier_name = ''
            if r.supplier:
                supplier_name = r.supplier.company_name or ''

            ws.append([
                prod.barcode or '',
                supplier_name,
                prod.jewelry_type or '',
                f3(prod.gram),
                f3(prod.buy_price_hs),
                f3(prod.sale_price_hs),
                f2(prod.profit),
                stat_text,
                created_text,
                active_text,
                prod.rfid_code or '',
                gorsel_dosya_adi,
            ])

        # Sütun genişlikleri
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        xlsx_path = os.path.join(tmp_dir, 'urunler.xlsx')
        wb.save(xlsx_path)

        # ── 5. ZIP arşivi oluştur ──
        now_str = timezone.now().strftime('%Y%m%d_%H%M')
        zip_filename = f'barkodlu_urun_yedek_{now_str}.zip'
        zip_path = os.path.join(tmp_dir, zip_filename)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(xlsx_path, 'urunler.xlsx')
            for fname in os.listdir(gorseller_dir):
                fpath = os.path.join(gorseller_dir, fname)
                if os.path.isfile(fpath):
                    zf.write(fpath, f'gorseller/{fname}')

        # ── 6. GeneratedReports'a kaydet — ATOMIK ──
        # save=True: dosya diske yazılır VE DB kaydı aynı anda güncellenir.
        # save=False + ayrı save() iki işlem arasında exception riski taşır.
        with open(zip_path, 'rb') as f:
            zip_content = ContentFile(f.read())

        report.status = 'SUCCESS'
        report.file.save(zip_filename, zip_content, save=True)  # save=True → tek DB yazması

        log.info(
            f"Backup ZIP tamamlandı: {zip_filename}, "
            f"toplam={total_count}, görsel={images_copied}"
        )

    # ── Exception: hem normal hataları hem de Celery sinyallerini yakala ──
    except BaseException as exc:
        # BaseException: SoftTimeLimitExceeded, SystemExit, KeyboardInterrupt dahil
        log.exception(f"Backup ZIP task FAILED (task_id={task_id}): {exc}")
        _mark_failed(task_id, exc)
        # SoftTimeLimitExceeded dışındaki BaseException'ları yeniden fırlat
        # (Celery'nin kendi yönetimine bırak)
        if not isinstance(exc, Exception):
            raise

    finally:
        # ── Geçici klasörü HER DURUMDA temizle (disk dolmasını önle) ──
        if tmp_dir and os.path.isdir(tmp_dir):
            try:
                shutil.rmtree(tmp_dir)
                log.debug(f"Geçici klasör temizlendi: {tmp_dir}")
            except Exception as cleanup_err:
                log.warning(f"Geçici klasör temizleme hatası: {cleanup_err}")
