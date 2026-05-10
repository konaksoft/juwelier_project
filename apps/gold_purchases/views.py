import random
from datetime import datetime  # Sadece bu import kalmalı
from decimal import InvalidOperation, Decimal
from pathlib import Path

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum, Value, DecimalField, F, Func, Subquery, OuterRef, Count, ExpressionWrapper, Case, When
from django.db.models.functions import Coalesce, Cast, Lower, Replace
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods

# Modeller ve diğer importlar
from apps.activity_logs.views import write_log
from apps.gold_purchases.models import GoldPurchases, ProductCategory, BarcodeTemplate
from apps.helpers.image_resize import process_image
from apps.products.importers import import_excel_as_products_and_purchases
from apps.products.models import Products
from apps.roles.decorators import role_required
from apps.definitions.categories.models import Categories
from apps.settings.models import (
    StoreLabelSettings,
    default_small_config, default_large_config,
    default_diamond_small_config, default_diamond_large_config,
    default_watch_small_config, default_watch_large_config,
)
from apps.suppliers.models import Suppliers, SupplierLedger
from apps.suppliers.services import get_ledger_currency
from apps.process.models import Process
from apps.process.views import generate_process_no

# --- FAZ 9.7: StockService ve ConversionService entegrasyonu ---
from apps.stock_management.services.stock_service import StockService, InsufficientStockError
from apps.stock_management.models import StockLedger, StockSnapshot
from apps.stock_management.services.price_service import PriceService
from apps.stock_management.services.conversion_service import ConversionService

# --- PIVOT FAZ E (2026-04-23): Çoklu Maden ürün uzantı tabloları ---
from apps.products.models import MaterialType, DiamondDetail, DiamondStone, WatchDetail


# --- YARDIMCI FONKSİYONLAR ---

def parse_decimal_locale(val, default="0"):
    if val is None:
        return Decimal(default)
    s = str(val).strip().replace(' ', '')
    if not s:
        return Decimal(default)
    if ',' in s and '.' in s:
        last_comma = s.rfind(',')
        last_dot = s.rfind('.')
        if last_comma > last_dot:
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    else:
        if ',' in s:
            s = s.replace('.', '')
            s = s.replace(',', '.')
    if s in ('.', '-.', '', '-'):
        return Decimal(default)
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal(default)


def generate_rfid_hex():
    """ZPL için 24 karakterlik Hex kodu üretir (Tarih + Random)."""
    now = datetime.now()
    time_part = now.strftime('%d%m%Y%H%M%S')  # 14 karakter
    random_part = str(random.randint(1, 9999999999)).zfill(10)  # 10 karakter
    return f"{time_part}{random_part}"


def generate_barcode(product_name, store, prefix=None):
    """
    Gap-Filling barkod üretici.
    Silinen kayıtların bıraktığı boşlukları doldurarak aralıksız sıralı barkod üretir.

    Algoritma:
    1. prefix belirlenir (ProductCategory.barcode_prefix veya product_name'den türetilir)
    2. Aynı prefix + store için aktif (is_deleted=False) barkodlar çekilir
    3. Sayısal kısımlar ayıklanarak bir set oluşturulur
    4. 1'den başlayarak sette olmayan ilk sayı bulunur (gap-filling)
    5. Bulunan sayı 5 haneli sıfır dolgusuyla formatlanır (ör: YZK-00002)
    """
    if prefix:
        prefix = prefix.strip().upper()
    else:
        s = (product_name or '').strip().upper()
        if not s:
            prefix = "PR"
        else:
            first_char = ""
            last_char = ""
            for ch in s:
                if ch.isalnum():
                    first_char = ch
                    break
            for ch in reversed(s):
                if ch.isalnum():
                    last_char = ch
                    break
            prefix = (first_char + last_char).strip() or "PR"

    # Tire yok, prefix doğrudan sayıyla birleşir (YZK00002 formatı)
    width = 4
    full_prefix = prefix

    # Aktif (silinmemiş) kayıtlardan aynı prefix ile başlayan barkodları çek
    existing = Products.objects.filter(
        is_deleted=False, store=store, barcode__istartswith=full_prefix
    ).exclude(barcode__isnull=True).exclude(barcode__exact="").values_list('barcode', flat=True)

    # Kullanılan sayıları bir set'e topla
    used_numbers = set()
    pref_len = len(full_prefix)
    for b in existing:
        bb = (str(b) if b is not None else "").strip().upper()
        if not bb.startswith(full_prefix):
            continue
        suf = bb[pref_len:]
        # Sadece tam sayısal ve doğru genişlikteki suffix'leri kabul et
        if suf.isdigit():
            used_numbers.add(int(suf))

    # Gap-Filling: 1'den başlayarak kullanılmayan ilk sayıyı bul
    n = 1
    while n in used_numbers:
        n += 1

    # Son kontrol: veritabanında gerçekten bu barkod yoksa kullan
    candidate = f"{full_prefix}{str(n).zfill(width)}".upper()
    while Products.objects.filter(is_deleted=False, store=store, barcode__iexact=candidate).exists():
        n += 1
        candidate = f"{full_prefix}{str(n).zfill(width)}".upper()

    return candidate


def _tr_lower(s: str) -> str:
    if not s: return ""
    return s.replace("İ", "i").replace("I", "ı").lower()


def _parse_tr_date(d: str):
    try:
        return datetime.strptime(d.strip(), '%d/%m/%Y').date()
    except Exception:
        return None


# ============================================================================
# Pırlanta etiket alanı çözümleyici (2026-04-28)
# ----------------------------------------------------------------------------
# Etiket render'ı (hem ZPL `get_print_data` hem HTML preview `print_barcode_normal`)
# Pırlanta ürünleri için DiamondDetail özet alanlarını + DiamondStone kayıtlarını
# birleştirerek anlamlı veri üretir.
#
# Mimari sorun: yeni form (R-01..R-12 refactor sonrası) kullanıcının çoğunlukla
# `DiamondStone[]` tablosuna veri girmesini öngörüyor; `DiamondDetail` özet
# alanları (carat_weight, color_grade, ...) çoğu zaman boş kalıyor.
# Çözüm: özet boşsa ilk taşın değerleri + toplam karat fallback'i.
#
# Para birimi: GOLD'un tek " S" (Satış) suffix'i Pırlanta'da yetersiz — döviz
# bazlı satıldığı için para birimi de basılır. TRY → "₺/TL", USD → "$",
# EUR → "EUR", GBP → "£" formatında suffix.
# ============================================================================

def _resolve_diamond_label_data(p):
    """
    Pırlanta ürünü için etikete basılacak alanları üretir.

    Args:
        p: Products instance (material_type=DIAMOND beklenir; değilse boş dönüş).

    Returns:
        dict (str → str). Tüm alanlar string; boş alanlar '' döner. Anahtarlar:
            total_carat_str  : "0.85 ct" (boşsa '')
            color_grade      : "F" / "Yeşil" / ''
            clarity_grade    : "VS1" / ''
            cut_grade        : "EXCELLENT" / ''
            certificate_lab  : "GIA" / '' (NONE filtrelenir)
            certificate_no   : "1234567890" / ''
            sale_price_str   : "1.200,00" (locale formatlı)
            sale_currency    : "USD" / "TRY" / ...
            price_with_ccy   : "1200,00 USD" / "1500,00 ₺" (suffix ile)
    """
    out = {
        'total_carat_str': '',
        'color_grade':     '',
        'clarity_grade':   '',
        'cut_grade':       '',
        'certificate_lab': '',
        'certificate_no':  '',
        'sale_price_str':  '0,00',
        'sale_currency':   '',
        'price_with_ccy':  '0,00',
        # FAZ DIA-LBL (2026-04-28): Montür altın bilgileri (opsiyonel — etikette
        # show/hide config ile kontrol edilir, default visible=False).
        'mount_karat':     '',   # "18K" / "" (NONE filtrelenir)
        'mount_gram_str':  '',   # "4,10 gr" / ""
    }

    dd = getattr(p, 'diamond_detail', None)

    # ── Fiyat: önce DiamondDetail.sale_price (döviz), boşsa Products.sale_price_eur ──
    raw_price = 0.0
    raw_currency = ''
    if dd and dd.sale_price is not None and float(dd.sale_price) > 0:
        raw_price = float(dd.sale_price)
        raw_currency = (dd.sale_currency or 'USD').upper()
    elif getattr(p, 'sale_price_eur', None) and float(p.sale_price_eur or 0) > 0:
        raw_price = float(p.sale_price_eur)
        raw_currency = 'TRY'

    if raw_price > 0:
        out['sale_price_str'] = f"{raw_price:.2f}".replace('.', ',')
    out['sale_currency'] = raw_currency

    _ccy_suffix_map = {
        'TRY': '₺',
        'USD': '$',
        'EUR': 'EUR',
        'GBP': '£',
    }
    suffix = _ccy_suffix_map.get(raw_currency, raw_currency or 'S')
    out['price_with_ccy'] = f"{out['sale_price_str']} {suffix}".strip()

    if not dd:
        return out

    # ── Taş kayıtları (sıralı: position artan) ──
    try:
        stones = list(dd.stones.all().order_by('position'))
    except Exception:
        stones = []

    first_stone = stones[0] if stones else None

    # ── Toplam karat: DiamondDetail.carat_weight varsa onu, yoksa stones toplamı ──
    total_carat = 0.0
    if dd.carat_weight and float(dd.carat_weight) > 0:
        total_carat = float(dd.carat_weight)
    elif stones:
        total_carat = sum(float(s.carat_weight or 0) for s in stones)
    if total_carat > 0:
        out['total_carat_str'] = f"{total_carat:.2f} ct"

    # ── 4C alanları: özet → ilk taş fallback zinciri ──
    def _pick(summary_val, stone_attr):
        v = (summary_val or '').strip() if isinstance(summary_val, str) else (summary_val or '')
        if v:
            return str(v)
        if first_stone:
            sv = getattr(first_stone, stone_attr, None)
            if sv:
                sv = str(sv).strip()
                if sv:
                    return sv
        return ''

    out['color_grade']   = _pick(dd.color_grade,   'color_grade')
    out['clarity_grade'] = _pick(dd.clarity_grade, 'clarity_grade')
    out['cut_grade']     = _pick(dd.cut_grade,     'cut_grade')

    cert_lab = _pick(dd.certificate_lab, 'certificate_lab')
    if cert_lab.upper() == 'NONE':
        cert_lab = ''
    out['certificate_lab'] = cert_lab

    out['certificate_no'] = _pick(dd.certificate_no, 'certificate_no')

    # ── Montür altın alanları: NONE / 0.000 olanlar boş döner ──
    mk_raw = (dd.mount_karat or '').strip() if isinstance(dd.mount_karat, str) else (dd.mount_karat or '')
    if mk_raw and str(mk_raw).upper() != 'NONE':
        out['mount_karat'] = str(mk_raw)

    try:
        mg_val = float(dd.mount_gram or 0)
    except (TypeError, ValueError):
        mg_val = 0.0
    if mg_val > 0:
        out['mount_gram_str'] = f"{mg_val:.2f} gr".replace('.', ',')

    return out


@login_required
def get_print_data(request):
    ids_param = request.GET.get('ids', '')
    if not ids_param:
        return JsonResponse({'result': False, 'error_msg': 'Seçili ürün yok.'})

    product_ids = ids_param.split(',')
    # Pırlanta etiketinde 4C fallback'i için DiamondStone kayıtları gereklidir;
    # N+1 önlemek için prefetch_related ile eager-load.
    records = GoldPurchases.objects.filter(id__in=product_ids).select_related(
        'product', 'product__diamond_detail', 'product__watch_detail', 'store', 'supplier'
    ).prefetch_related('product__diamond_detail__stones')

    # ── Mağaza etiket ayarlarını oku ──
    bottom_left_type = 'MILYEM'
    layout_mode = 'STANDARD'
    active_size = 'small'
    bc_x, bc_y, bc_h, bc_vis = 200, 127, 35, True
    # RFID Modu varsayılanı True — mevcut RFID müşterilerini koruma altına alır
    rfid_mode = True
    # Barkod çizgi kalınlığı (ZPL ^BY modül genişliği) için ayar; JSON config'te saklanır
    barcode_lines_width = 70

    # Tüm material konfigürasyonlarını önceden yükle; ürün loop'unda material_type'a göre seçilir
    configs_by_material = {
        'GOLD':    {'small': default_small_config(),         'large': default_large_config()},
        'DIAMOND': {'small': default_diamond_small_config(), 'large': default_diamond_large_config()},
        'WATCH':   {'small': default_watch_small_config(),   'large': default_watch_large_config()},
    }

    try:
        settings = StoreLabelSettings.objects.filter(store=request.user.store).first()
        if settings:
            active_size = settings.active_size or 'small'
            # Her material için DB'deki override'ı defaults üzerine merge et
            def _merge(default_cfg, saved):
                merged = dict(default_cfg)
                if saved:
                    merged.update(saved)
                return merged
            configs_by_material['GOLD']['small']    = _merge(default_small_config(),         settings.small_design)
            configs_by_material['GOLD']['large']    = _merge(default_large_config(),         settings.large_design)
            configs_by_material['DIAMOND']['small'] = _merge(default_diamond_small_config(), getattr(settings, 'diamond_small_design', None))
            configs_by_material['DIAMOND']['large'] = _merge(default_diamond_large_config(), getattr(settings, 'diamond_large_design', None))
            configs_by_material['WATCH']['small']   = _merge(default_watch_small_config(),   getattr(settings, 'watch_small_design', None))
            configs_by_material['WATCH']['large']   = _merge(default_watch_large_config(),   getattr(settings, 'watch_large_design', None))

            bottom_left_type = settings.label_bottom_left_type or 'MILYEM'
            layout_mode = settings.label_layout_mode or 'STANDARD'
            bc_x = settings.barcode_lines_x if settings.barcode_lines_x is not None else 200
            bc_y = settings.barcode_lines_y if settings.barcode_lines_y is not None else 127
            bc_h = settings.barcode_lines_height if settings.barcode_lines_height is not None else 35
            bc_vis = settings.barcode_lines_visible if settings.barcode_lines_visible is not None else True
            # rfid_mode alanı migration sonrası erişilebilir; erken erişimde AttributeError'dan koru
            rfid_mode = getattr(settings, 'rfid_mode', True)
            if rfid_mode is None:
                rfid_mode = True
            try:
                # Barcode line genişliği için aktif boyut + GOLD config kullanılır (ortak ayar)
                blw_val = int(configs_by_material['GOLD'][active_size].get('barcode_lines_width', 70))
                if 10 <= blw_val <= 200:
                    barcode_lines_width = blw_val
            except (ValueError, TypeError):
                pass
    except Exception:
        pass

    # GOLD config'i geriye dönük uyum için "config" değişkenine de tut (fallback)
    config = configs_by_material['GOLD'][active_size]

    # Kullanıcı Genişlik ayarını (10–200) ZPL ^BY modül genişliğine (1–5) map'le.
    # ZPL ^BY module_width 1–10 aralığını kabul eder; üst sınır 5 ile etiketin
    # taşması engellenir. Oran (ratio) sabit 3 kalır.
    def _map_barcode_module_width(blw):
        if blw < 40: return 1
        if blw < 70: return 2
        if blw < 100: return 3
        if blw < 150: return 4
        return 5

    bc_module_w = _map_barcode_module_width(barcode_lines_width)

    label_data = []

    try:
        _store = request.user.store
        store_name = _store.barcode_title or getattr(_store, 'title', None) or getattr(_store, 'name', None) or "KUYUM PLUS"
    except Exception:
        store_name = "KUYUM PLUS"

    def clean_text(text):
        if not text: return ""
        return text.replace('İ', 'I').replace('ı', 'i').replace('Ş', 'S').replace('ş', 's').replace('Ğ', 'G').replace(
            'ğ', 'g').upper()

    store_name = clean_text(store_name)

    # Düzen sabitleri (dot cinsinden)
    SIDE_BY_SIDE_OFFSET = 400
    LABEL_HEIGHT = 250  # Yansımalı mod için etiket yüksekliği (dot)

    for item in records:
        p = item.product
        if not p: continue

        if not p.rfid_code:
            p.rfid_code = generate_rfid_hex()
            p.save()

        # Ürünün material_type'ına göre config seç (GOLD geriye dönük varsayılan)
        product_material = getattr(p, 'material_type', None) or MaterialType.GOLD
        if product_material not in configs_by_material:
            product_material = MaterialType.GOLD
        config = configs_by_material[product_material][active_size]

        barcode = p.barcode or ""
        rfid = p.rfid_code or ""

        gram_val = f"{p.gram:.2f}".replace('.', ',') if p.gram else "0,00"
        cost_val = f"{p.buy_price_hs:.2f}".replace('.', ',') if p.buy_price_hs else "0,00"
        sale_price = f"{p.sale_price_hs:.2f}".replace('.', ',') if p.sale_price_hs else "0,00"
        jewelry_type = clean_text(p.jewelry_type or "URUN")
        supplier_name = clean_text(item.supplier.company_name if item.supplier else '')
        ring_size_val = (p.ring_size or '').strip()

        # ── Sol alt köşe verisi: Milyem veya Maliyet ──
        if bottom_left_type == 'MALIYET':
            bottom_left_val = cost_val + ' M'
        else:
            # Milyem: tam sayı olarak bas (585 M, 750 M vb.)
            try:
                milyem_int = int(float(p.product_mileage))
                bottom_left_val = f"{milyem_int} M"
            except (ValueError, TypeError):
                bottom_left_val = "0 M"

        gold_rate = f"{float(p.gold_rate):.0f}" if p.gold_rate else ""
        karat_text = ""
        if gold_rate == "585":
            karat_text = "14K"
        elif gold_rate == "916":
            karat_text = "22K"
        elif gold_rate == "750":
            karat_text = "18K"
        elif gold_rate == "416":
            karat_text = "10K"
        elif gold_rate == "333":
            karat_text = "8K"
        elif gold_rate == "995":
            karat_text = "24K"
        else:
            karat_text = f"{gold_rate}"

        def z(key, text_value, x_offset=0, inverted=False, y_override=None):
            field_conf = config.get(key, {})
            if not field_conf:
                field_conf = default_small_config().get(key, {})

            if not field_conf.get('visible', True):
                return ""

            x = field_conf.get('x', 0) + x_offset
            y = y_override if y_override is not None else field_conf.get('y', 0)
            f = field_conf.get('font', 20)

            rotation = 'I' if inverted else 'N'
            return f"^FT{x},{y}^A0{rotation},{f},{f}^FH\\^FD{text_value}^FS"

        # ── Barkod çizgileri ZPL komutu ──
        def build_barcode_cmd(x_offset=0, inverted=False, mirror_y=False):
            if not bc_vis:
                return ""
            bx = bc_x + x_offset
            by = bc_y
            if mirror_y:
                by = LABEL_HEIGHT - bc_y + 100
            rotation = 'I' if inverted else 'N'
            # ^BY: modül genişliği (bc_module_w), oran (3) — yükseklik ^BC'de verilir
            # Modül genişliği kullanıcının "Barkod Çizgileri Genişlik" ayarından map'lenir
            # ^BC{rotation},{bc_h},N,N — 2. parametre = barkod yüksekliği (dot)
            return f"^FW{rotation}^BY{bc_module_w},3^FT{bx},{by}^BC{rotation},{bc_h},N,N^FD{barcode}^FS^FWN"

        # ── Material'a özel metin alanı listesi ──
        # GOLD: karat/gram/milyem; DIAMOND: 4C + sertifika; WATCH: marka/model/referans
        if product_material == MaterialType.DIAMOND:
            # 2026-04-28: DiamondDetail özet alanları boşsa DiamondStone'dan
            # auto-derive eden helper kullanılır. Fiyat dd.sale_price'tan
            # (döviz) okunur — eskiden yanlışlıkla GOLD alanı sale_price_hs
            # kullanılıyordu, bu yüzden Pırlanta etiketinde "0,00" çıkıyordu.
            dl = _resolve_diamond_label_data(p)
            carat_val   = clean_text(dl['total_carat_str'])
            color_val   = clean_text(dl['color_grade'])
            clarity_val = clean_text(dl['clarity_grade'])
            cut_val     = clean_text(dl['cut_grade'])
            cert_lab    = clean_text(dl['certificate_lab'])
            cert_no     = clean_text(dl['certificate_no'])
            mount_karat_val = clean_text(dl['mount_karat'])
            mount_gram_val  = clean_text(dl['mount_gram_str'])
            # FAZ DIA-LBL (2026-04-28): show_currency toggle — config'te kapalıysa
            # döviz simgesi (₺/$/€/£) basılmaz; yalnız sayı görünür.
            _price_cfg = config.get('price', {}) or {}
            _show_ccy  = bool(_price_cfg.get('show_currency', True))
            diamond_price_label = dl['price_with_ccy'] if _show_ccy else dl['sale_price_str']
            text_field_list = [
                ('store_name',      store_name),
                ('barcode_no',      barcode),
                ('carat_weight',    carat_val),
                ('color_grade',     color_val),
                ('clarity_grade',   clarity_val),
                ('cut_grade',       cut_val),
                ('mount_karat',     mount_karat_val),
                ('mount_gram',      mount_gram_val),
                ('price',           diamond_price_label),
                ('certificate_lab', cert_lab),
                ('certificate_no',  cert_no),
                ('supplier',        supplier_name),
            ]
        elif product_material == MaterialType.WATCH:
            wd = getattr(p, 'watch_detail', None)
            brand_val    = clean_text(wd.brand or "") if wd else ""
            model_val    = clean_text(wd.model_name or "") if wd else ""
            ref_val      = clean_text(wd.reference_no or "") if wd else ""
            movement_val = clean_text(wd.movement_type or "") if wd else ""
            condition_v  = clean_text(wd.condition or "") if wd else ""
            box_val      = clean_text(wd.box_papers or "") if wd else ""
            text_field_list = [
                ('store_name',    store_name),
                ('barcode_no',    barcode),
                ('brand',         brand_val),
                ('model_name',    model_val),
                ('reference_no',  ref_val),
                ('movement_type', movement_val),
                ('condition',     condition_v),
                ('price',         sale_price + ' S'),
                ('supplier',      supplier_name),
                ('box_papers',    box_val),
            ]
        else:
            text_field_list = [
                ('store_name',   store_name),
                ('barcode_no',   barcode),
                ('quality_text', karat_text),
                ('price',        sale_price + ' S'),
                ('mileage',      bottom_left_val),
                ('gram',         gram_val + ' gr'),
                ('jewelry_type', jewelry_type),
                ('supplier',     supplier_name),
                ('ring_size',    ring_size_val),
            ]

        # ── Tüm metin alanlarını oluşturan yardımcı ──
        def build_text_fields(x_offset=0, inverted=False, mirror_y=False):
            fields = ""
            for key, val in text_field_list:
                if mirror_y:
                    fc = config.get(key, {}) or default_small_config().get(key, {})
                    orig_y = fc.get('y', 0)
                    new_y = LABEL_HEIGHT - orig_y + 100
                    fields += z(key, val, x_offset=x_offset, inverted=inverted, y_override=new_y)
                else:
                    fields += z(key, val, x_offset=x_offset, inverted=inverted)
            return fields

        # ── BİRİNCİ YARI (Her zaman basılır) ──
        # ^RFW sadece rfid_mode=True iken yazılır. RFID'siz (standart) Zebra
        # yazıcılarda ^RFW void/kilitlenme'ye neden olabilir.
        rfid_cmd = f"^RFW,H,2,12,1^FD{rfid}^FS" if rfid_mode else ""
        first_half = (
            build_barcode_cmd()
            + rfid_cmd
            + "^CI28"
            + build_text_fields()
        )

        # ── İKİNCİ YARI (Düzene göre) ──
        second_half = ""

        if layout_mode == 'REFLECTED':
            # Yansımalı (Kelebek): 180° ters yazı + Y aynalama + ters barkod
            second_half = (
                build_barcode_cmd(inverted=True, mirror_y=True)
                + build_text_fields(inverted=True, mirror_y=True)
            )

        elif layout_mode == 'SIDE_BY_SIDE':
            # Yan Yana (Çiftli): X offsetli ikinci kopya + barkod çizgileri
            second_half = (
                build_barcode_cmd(x_offset=SIDE_BY_SIDE_OFFSET)
                + build_text_fields(x_offset=SIDE_BY_SIDE_OFFSET)
            )

        zpl_code = f"^XA{first_half}{second_half}^XZ"

        label_data.append({
            'barcode': barcode,
            'name': jewelry_type,
            'gram': gram_val,
            'zpl': zpl_code
        })

    return JsonResponse({'result': True, 'labels': label_data})


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_GOLD_PURCHASE_ADD')
@require_http_methods(["POST"])
def gold_purchase_add(request):
    record_id = request.POST.get('gold_purchase_id')
    store = request.user.store
    # Kategori kontrolü (Hata almamak için try-except veya get_or_create kullanılabilir)
    try:
        category = Categories.objects.get(name='Barkodlu Ürünler', is_deleted=False)
    except Categories.DoesNotExist:
        # Eğer kategori yoksa ilki veya varsayılan
        category = Categories.objects.filter(is_deleted=False).first()

    if record_id:
        record = get_object_or_404(Products, id=record_id)
        supplier_val = request.POST.get('supplier_id') or None
        GoldPurchases.objects.filter(
            product=record, store=store, is_deleted=False
        ).update(supplier_id=supplier_val)
    else:
        record = Products()
        record.store = store

        # --- GÖREV 2: ProductCategory barcode_prefix kullan ---
        jt = request.POST.get('jewelry_type', '')
        cat_prefix = None
        try:
            pc = ProductCategory.objects.get(store=store, name=jt, is_deleted=False)
            cat_prefix = pc.barcode_prefix
        except ProductCategory.DoesNotExist:
            pass

        record.barcode = generate_barcode(jt, store=store, prefix=cat_prefix)
        # RFID Otomatik Oluştur
        if not record.rfid_code:
            record.rfid_code = generate_rfid_hex()

        new_record = GoldPurchases()
        new_record.created_by = request.user
        new_record.store = store
        new_record.supplier_id = request.POST.get('supplier_id') or None

    try:
        with transaction.atomic():
            record.category = category
            record.jewelry_type = request.POST.get('jewelry_type') or ''
            record.name = record.jewelry_type
            record.created_on = timezone.now()

            # Sayısal değerleri güvenli parse et
            record.retail_lower_limit = parse_decimal_locale(request.POST.get('retail_lower_limit'))
            record.retail_top_limit = parse_decimal_locale(request.POST.get('retail_top_limit'))
            record.wholesale_lower_limit = parse_decimal_locale(request.POST.get('wholesale_lower_limit'))
            record.wholesale_top_limit = parse_decimal_locale(request.POST.get('wholesale_top_limit'))
            record.profit = parse_decimal_locale(request.POST.get('profit'))
            record.product_mileage = parse_decimal_locale(request.POST.get('product_mileage'))
            record.labor_mileage = parse_decimal_locale(request.POST.get('labor_mileage'))
            record.piece_labor = parse_decimal_locale(request.POST.get('piece_labor'))
            record.gold_rate = parse_decimal_locale(request.POST.get('gold_rate'))
            record.buy_price_hs = parse_decimal_locale(request.POST.get('buy_price_hs'))
            record.sale_price_hs = parse_decimal_locale(request.POST.get('sale_price_hs'))
            record.gram = parse_decimal_locale(request.POST.get('gram'), default="0.000")
            record.ring_size = (request.POST.get('ring_size') or '').strip()
            record.is_gram_bullion = False

            image = request.FILES.get('image')
            if image:
                filename, processed = process_image(image)
                record.image.save(filename, processed, save=False)

            # Barkod kontrolü
            record.barcode = (record.barcode or '').strip().upper()
            if record.store_id and record.barcode:
                exists = Products.objects.filter(
                    is_deleted=False, store=record.store, barcode__iexact=record.barcode
                ).exclude(pk=record.pk).exists()
                if exists:
                    record.barcode = generate_barcode(record.jewelry_type, store=record.store)

            # RFID kontrol (Edit durumunda boşsa doldur)
            if not record.rfid_code:
                record.rfid_code = generate_rfid_hex()

            record.save()

            if not record_id:
                new_record.product = record
                new_record.is_status = True
                new_record.save()

                # ── FAZ 9.7: 3'lü Senaryo Entegrasyonu ──────────────────────
                # Senaryo A — Stoktan Dönüşüm: is_from_stock → ConversionService
                # Senaryo B — Hazır Alım:       supplier var  → PURCHASE
                # Senaryo C — Açılış Stoğu:     supplier yok  → INITIAL
                # ─────────────────────────────────────────────────────────────
                is_from_stock = request.POST.get('is_from_stock')

                if is_from_stock:
                    # ── Senaryo A: Stoktan Dönüşüm (Hurda veya Bilezik) ──
                    source_type = request.POST.get('source_type', 'Hurda')
                    milyem = record.product_mileage or record.gold_rate

                    if source_type == 'Bilezik':
                        # Bilezik havuzundan dönüşüm
                        bilezik_cat = Categories.objects.filter(
                            name='Bilezik', is_deleted=False
                        ).first()
                        source_product = Products.objects.filter(
                            store=store,
                            category=bilezik_cat,
                            product_mileage=milyem,
                            is_deleted=False,
                        ).first() if bilezik_cat else None
                    else:
                        # Hurda havuzundan dönüşüm (varsayılan)
                        source_product = Products.objects.filter(
                            store=store,
                            is_scrap=True,
                            product_mileage=milyem,
                            is_deleted=False,
                        ).first()

                    if not source_product:
                        raise ValueError(
                            f"Bu ayar ({milyem}) için mağazanızda "
                            f"{source_type} ürünü bulunamadı. "
                            f"Lütfen önce ilgili {source_type.lower()} kaydını oluşturun."
                        )

                    melt_loss_gram = parse_decimal_locale(
                        request.POST.get('melt_loss_gram', '0')
                    )
                    used_source_gram = record.gram + melt_loss_gram

                    ConversionService.convert_scrap_to_product(
                        source_scrap=source_product,
                        target_product=record,
                        store=store,
                        used_scrap_gram=used_source_gram,
                        target_quantity_pieces=1,
                        target_quantity_gram=record.gram,
                        melt_loss_gram=melt_loss_gram,
                        user=request.user,
                        notes=(
                            f"Stoktan üretim ({source_type}): "
                            f"{source_product.name} → {record.name}"
                        ),
                    )

                elif record.gram and record.gram > 0:
                    # Stoktan üretim değil — normal stok girişi
                    # Has Altın TL kuru (WAC hesabı için)
                    hs_rate_eur = Decimal('0.0000')
                    try:
                        hs_data = PriceService.get_price('GOLD_24K')
                        hs_rate_eur = hs_data.get('buy_tl', Decimal('0'))
                        if hs_rate_eur <= 0:
                            hs_prod = Products.objects.filter(
                                name__icontains='Has Altın'
                            ).only('buy_price_eur').first()
                            if hs_prod:
                                hs_rate_eur = Decimal(str(hs_prod.buy_price_eur or 0))
                    except Exception:
                        pass

                    # ── BIRIM MALIYET (WAC icin gram basina HS) ──────────────
                    # FAZ 34 KOK NEDEN FIX (2026-05-01):
                    # gold_purchases/index.html:2249'daki JS sunu hesapliyor:
                    #     totalHas = ((product_mileage + labor_mileage)/1000) * gram + piece_labor
                    #     buyPriceHs.value = totalHas
                    # Yani form'daki "Maliyet (Has)" alani, urunun TOPLAM Has
                    # karsiligini tutar (orn. 15gr x 0.685 = 10.275). Eskiden
                    # bu deger record.buy_price_hs den dogrudan unit_cost_hs e
                    # geciyordu; StockService.record_entry WAC formulu ise
                    # gram bazli (gram x unit_cost_hs) calisiyor:
                    #     new_wac = (15 x 10.275) / 15 = 10.275 HS/gr
                    # Sonuc: Dashboard "stock_gram x WAC" = 15 x 10.275 = 154 HS
                    # gibi imkansiz bir HAS toplami uretiyordu.
                    #
                    # Cozum: Toplam Has degerini gram'a bolerek gercek BIRIM
                    # maliyeti elde et. 1 gramlik urunde toplam == birim
                    # oldugu icin bug 1 gramda kendini gizliyordu.
                    _total_buy_hs = Decimal(str(record.buy_price_hs or 0))
                    _gram = Decimal(str(record.gram or 0))
                    if _gram > 0 and _total_buy_hs > 0:
                        unit_cost_hs = (_total_buy_hs / _gram).quantize(Decimal('0.0001'))
                    else:
                        unit_cost_hs = Decimal('0')

                    unit_cost_eur = Decimal('0.00')
                    if hs_rate_eur > 0 and unit_cost_hs > 0:
                        unit_cost_eur = (unit_cost_hs * hs_rate_eur).quantize(
                            Decimal('0.01')
                        )

                    supplier_id = request.POST.get('supplier_id')
                    if supplier_id:
                        # ── Senaryo B: Hazır Alım (Tedarikçiden) ──
                        reason = StockLedger.Reason.PURCHASE
                        ref_type = 'gold_purchase'
                        notes_text = f"Tedarikçi alımı: {record.name}"
                    else:
                        # ── Senaryo C: Açılış Stoğu ──
                        reason = StockLedger.Reason.INITIAL
                        ref_type = 'initial_stock'
                        notes_text = f"Açılış stoğu: {record.name}"

                    StockService.record_entry(
                        product=record,
                        store=store,
                        quantity_gram=record.gram,
                        quantity_pieces=1,
                        reason=reason,
                        ref_type=ref_type,
                        ref_id=f"gold_purchase_{new_record.id}",
                        unit_cost_hs=unit_cost_hs,
                        unit_cost_eur=unit_cost_eur,
                        hs_rate_eur=hs_rate_eur,
                        user=request.user,
                        notes=notes_text,
                    )

                    # ── Tedarikçi Cari + İşlem Entegrasyonu ──
                    process_ledger = request.POST.get('process_supplier_ledger')
                    if process_ledger and supplier_id:
                        try:
                            supplier = Suppliers.objects.get(id=supplier_id)
                            # Has maliyeti — yuvarlama YOK, tam Decimal hassasiyeti korunur
                            ledger_amount_hs = record.buy_price_hs or Decimal('0')
                            if ledger_amount_hs > 0:
                                gp_process_no = generate_process_no()

                                # 1. SupplierLedger (Cari Borç)
                                # PIVOT FAZ E (2026-04-23): currency='HS' hardcode kaldırıldı.
                                # get_ledger_currency() material_type bazlı doğru birimi seçer:
                                #   GOLD -> HS, SILVER -> HG, WATCH/DIAMOND -> fiat (TRY default).
                                SupplierLedger.objects.create(
                                    supplier=supplier,
                                    product=record,
                                    transaction_type=SupplierLedger.ENTRY,
                                    quantity_piece=1,
                                    quantity_gram=record.gram,
                                    amount_value=ledger_amount_hs,
                                    currency=get_ledger_currency(record, fiat_currency='TRY'),
                                    process_no=gp_process_no,
                                    description=f"Hazır Alım: {record.jewelry_type or record.name} - {record.barcode}",
                                    is_active=True,
                                )

                                # 2. Process Kaydı (İşlemler ekranında görünsün)
                                Process.objects.create(
                                    store=store,
                                    process_no=gp_process_no,
                                    process_type='RETAIL',
                                    transaction_type='PURCHASE',
                                    product=record,
                                    supplier=supplier,
                                    employee=request.user,
                                    piece=1,
                                    gram=record.gram,
                                    price_hs=ledger_amount_hs,
                                    unit_price=unit_cost_eur,
                                    amount=unit_cost_eur,
                                    is_status='COMPLETED',
                                    is_deleted=False,
                                )
                        except Suppliers.DoesNotExist:
                            pass

        response_data = {'result': True}
        # Kaydet ve Yazdır için ürün bilgilerini döndür
        if not record_id:
            response_data['gold_purchase_id'] = str(new_record.id)
        else:
            gp = GoldPurchases.objects.filter(product=record, store=store, is_deleted=False).first()
            response_data['gold_purchase_id'] = str(gp.id) if gp else ''
        response_data['product_id'] = str(record.id)
        response_data['barcode'] = record.barcode or ''
        response_data['rfid_code'] = record.rfid_code or ''
        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({'error': True, 'error_msg': str(e)})


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_GOLD_PURCHASES_INDEX')
def gold_purchases_index(request):
    store = request.user.store
    suppliers = Suppliers.objects.filter(is_deleted=False, store=store)

    DEC18_6 = DecimalField(max_digits=18, decimal_places=6)
    # FAZ 44 — 1.05 EŞİK KURALI:
    # Products.buy_price_hs iki çağdan veri tutuyor:
    #   - Legacy gold_purchases formundan giriş: TOPLAM HS (örn. 10.575)
    #   - Perakende alış akışı (retail_views.py:1355) üzerine yazımı: BİRİM HS (örn. 0.705)
    # Saf altın fraksiyonu ≤ 1.000 olduğundan 1.05 üstü değer kesinlikle legacy
    # toplamdır → doğrudan toplanır. 1.05 ve altı birim demektir → gram ile çarpılır.
    buy_price_cast = Cast(F('product__buy_price_hs'), DEC18_6)
    gram_cast = Cast(F('product__gram'), DEC18_6)
    has_expr = Case(
        When(product__buy_price_hs__gt=Decimal('1.05'), then=buy_price_cast),
        default=ExpressionWrapper(buy_price_cast * gram_cast, output_field=DEC18_6),
        output_field=DEC18_6,
    )
    zero_dec = Value(Decimal('0'), output_field=DEC18_6)

    base_qs = GoldPurchases.objects.filter(is_deleted=False, store=store)

    # FAZ 9.8: Satıldı = is_status=False VEYA product.is_completed=True
    sold_q = Q(is_status=False) | Q(product__is_completed=True)

    agg = base_qs.aggregate(
        has_active=Coalesce(Sum(has_expr, filter=~sold_q, output_field=DEC18_6), zero_dec),
        has_sold=Coalesce(Sum(has_expr, filter=sold_q, output_field=DEC18_6), zero_dec),
        active_count=Count('id', filter=~sold_q),
        sold_count=Count('id', filter=sold_q),
        total_count=Count('id'),
    )

    # --- GÖREV 2: Dinamik takı tipleri ---
    product_categories = ProductCategory.objects.filter(
        store=store, is_deleted=False, is_active=True
    ).order_by('name')

    # --- GÖREV 3: Barkod şablonları ---
    # FAZ E UI FIX (2026-04-23): Şablonlar material_type bazında ayrıştırıldı.
    # - barcode_templates (LEGACY): Sadece Altın şablonları — mevcut altın tab
    #   template pill rendering'i bu key'e bağımlı, tek satır kırılmasın diye korundu.
    # - gold_templates / diamond_templates / watch_templates: Yeni tab-spesifik key'ler.
    all_templates = BarcodeTemplate.objects.filter(
        store=store, is_deleted=False, is_active=True
    ).order_by('template_name')

    context = {
        'suppliers': suppliers,
        'title': 'Tedarikçiden Altın Alımı',
        'gp_total_has_active': agg['has_active'],
        'gp_total_has_sold': agg['has_sold'],
        'gp_active_count': agg['active_count'],
        'gp_sold_count': agg['sold_count'],
        'gp_total_count': agg['total_count'],
        'product_categories': product_categories,
        # Geriye dönük uyum: barcode_templates = Sadece GOLD (Altın tab pill'leri değişmesin)
        'barcode_templates': all_templates.filter(material_type='GOLD'),
        # Material-type bazlı kırılım (yeni)
        'gold_templates': all_templates.filter(material_type='GOLD'),
        'diamond_templates': all_templates.filter(material_type='DIAMOND'),
        'watch_templates': all_templates.filter(material_type='WATCH'),
    }
    return render(request, 'management/gold_purchases/index.html', context)


@login_required(login_url='login')
def get_all(request):
    try:
        draw = int(request.GET.get('draw', 1))
        length = int(request.GET.get('length', 10))
        start = int(request.GET.get('start', 0))
        search_value = request.GET.get('search[value]', '').strip()
        order_column_index = request.GET.get('order[0][column]', '0')
        order_column = request.GET.get(f'columns[{order_column_index}][data]', 'created_on')
        order_direction = request.GET.get('order[0][dir]', 'desc')

        store = request.user.store

        gold_rate_param = request.GET.get('gold_rate', '')
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        status_filter = request.GET.get('status_filter', 'all')
        label_status = request.GET.get('label_status', 'not_printed')
        # FAZ E UI FIX v2 (2026-04-23): Materyal tipi filtresi.
        # Boş/None → filtre uygulanmaz (tüm tipler; eski davranış = regression yok).
        # Geçersiz değer de yoksayılır (whitelist kontrolü).
        material_type_param = (request.GET.get('material_type') or '').upper().strip()

        qs = GoldPurchases.objects.filter(is_deleted=False, store=store)

        if order_direction == 'desc':
            order_column = f'-{order_column}'

        qs = GoldPurchases.objects.filter(is_deleted=False, store=store)
        # --- SEKME FİLTRESİ ---
        if label_status == 'printed':
            qs = qs.filter(is_labeled=True)
        elif label_status == 'not_printed':
            qs = qs.filter(is_labeled=False)

        # FAZ 9.8: Satıldı = is_status=False VEYA product.is_completed=True
        sold_q = Q(is_status=False) | Q(product__is_completed=True)
        if status_filter == 'on_shelf':
            qs = qs.exclude(sold_q)
        elif status_filter == 'sold':
            qs = qs.filter(sold_q)

        # FAZ E UI FIX v2 (2026-04-23): Materyal tipi filtresi uygulaması.
        # Whitelist ile korunur; geçersiz değer sessizce yoksayılır.
        if material_type_param in ('GOLD', 'SILVER', 'DIAMOND', 'WATCH'):
            qs = qs.filter(product__material_type=material_type_param)

        df = _parse_tr_date(date_from) if date_from else None
        dt_ = _parse_tr_date(date_to) if date_to else None
        if df: qs = qs.filter(created_on__date__gte=df)
        if dt_: qs = qs.filter(created_on__date__lte=dt_)
        if gold_rate_param: qs = qs.filter(product__gold_rate=gold_rate_param)

        queryset = qs.values(
            'product__id', 'supplier_id', 'id',
            'product__name', 'product__image', 'product__barcode', 'product__gram',
            'product__jewelry_type', 'created_on',
            'product__buy_price_hs', 'supplier__company_name',
            'product__sale_price_hs', 'product__is_active',
            'product__profit', 'product__gold_rate',
            'product__product_mileage', 'product__labor_mileage', 'product__piece_labor',
            'is_status',
            'product__is_completed',
            # FAZ DIA-DT (2026-04-28): DIAMOND-specific fields for material-aware DataTable
            'product__material_type',
            'product__sale_price_eur',
            'product__diamond_detail__sale_price',
            'product__diamond_detail__sale_currency',
            'product__diamond_detail__carat_weight',
            'product__diamond_detail__color_grade',
            'product__diamond_detail__clarity_grade',
            'product__diamond_detail__cut_grade',
            'product__diamond_detail__certificate_lab',
            'product__diamond_detail__certificate_no',
        ).annotate(
            diamond_stones_total_carat=Sum('product__diamond_detail__stones__carat_weight'),
            diamond_stones_count=Count('product__diamond_detail__stones'),
        )

        total_records = queryset.count()

        if search_value:
            q_basic = (Q(product__name__icontains=search_value) | Q(product__barcode__icontains=search_value))
            queryset = queryset.annotate(
                name_tr_norm=Lower(
                    Replace(Replace(F('product__jewelry_type'), Value('İ'), Value('i')), Value('I'), Value('ı')))
            ).filter(q_basic | Q(name_tr_norm__contains=_tr_lower(search_value)))

        filtered_records = queryset.count()

        if length != -1:
            queryset = queryset.order_by(order_column)[start:start + length]
        else:
            queryset = queryset.order_by(order_column)

        return JsonResponse({
            "draw": draw,
            "recordsFiltered": filtered_records,
            "recordsTotal": total_records,
            "data": list(queryset)
        })
    except Exception as e:
        return JsonResponse({"error": True, "error_msg": str(e)}, status=500)


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_DELETE')
def delete(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids[]')
        try:
            with transaction.atomic():
                records = GoldPurchases.objects.filter(id__in=ids).select_related('product')
                store = request.user.store

                for gp in records:
                    product = gp.product
                    if not product:
                        gp.is_deleted = True
                        gp.save(update_fields=['is_deleted'])
                        continue

                    # 1. Bağlı Process kayıtlarını iptal et + SupplierLedger'ı kapat
                    linked_processes = Process.objects.filter(
                        product=product, store=store, is_deleted=False
                    )
                    for proc in linked_processes:
                        # Stok ters hareketi (alış iptali → stoktan düş)
                        if proc.transaction_type == 'PURCHASE' and not proc.waiting_stock:
                            piece = int(proc.piece or 0)
                            gram = Decimal(str(proc.gram or 0))
                            if piece > 0 or gram > 0:
                                _hs_rate_eur = Decimal('0')
                                try:
                                    _hs_data = PriceService.get_price('GOLD_24K')
                                    _hs_rate_eur = Decimal(str(_hs_data.get('buy_tl', Decimal('0'))))
                                except Exception:
                                    pass
                                snap = StockSnapshot.objects.filter(product=product, store=store).first()
                                u_hs = snap.weighted_avg_cost_hs if snap else Decimal('0')
                                u_tl = snap.weighted_avg_cost_eur if snap else Decimal('0')
                                try:
                                    StockService.record_exit(
                                        product=product, store=store,
                                        quantity_gram=gram, quantity_pieces=piece,
                                        reason=StockLedger.Reason.RETURN_OUT,
                                        ref_type='cancel_gold_purchase',
                                        ref_id=f"cancel_{proc.process_no}",
                                        unit_cost_hs=u_hs, unit_cost_eur=u_tl,
                                        hs_rate_eur=_hs_rate_eur,
                                        user=request.user,
                                        notes=f"Barkodlu ürün silme: {product.barcode}",
                                    )
                                except InsufficientStockError:
                                    pass

                        # SupplierLedger → pasife çek
                        if proc.process_no:
                            SupplierLedger.objects.filter(
                                process_no=proc.process_no, is_active=True
                            ).update(is_active=False)

                        # Process → iptal
                        proc.is_status = 'CANCELED'
                        proc.is_deleted = True
                        proc.save(update_fields=['is_status', 'is_deleted'])

                    # 2. Process kaydı olmayan ancak GP- formatıyla oluşturulmuş
                    #    eski SupplierLedger kayıtlarını da kapat
                    SupplierLedger.objects.filter(
                        product=product, is_active=True
                    ).update(is_active=False)

                    # 3. Eğer bağlı Process yoksa ama stokta varsa düş
                    if not linked_processes.exists():
                        if product.gram and product.gram > 0:
                            snap = StockSnapshot.objects.filter(product=product, store=store).first()
                            if snap and snap.stock_gram > 0:
                                _hs_rate_eur = Decimal('0')
                                try:
                                    _hs_data = PriceService.get_price('GOLD_24K')
                                    _hs_rate_eur = Decimal(str(_hs_data.get('buy_tl', Decimal('0'))))
                                except Exception:
                                    pass
                                try:
                                    StockService.record_exit(
                                        product=product, store=store,
                                        quantity_gram=snap.stock_gram,
                                        quantity_pieces=snap.stock_pieces,
                                        reason=StockLedger.Reason.RETURN_OUT,
                                        ref_type='cancel_gold_purchase',
                                        ref_id=f"gp_delete_{gp.id}",
                                        unit_cost_hs=snap.weighted_avg_cost_hs,
                                        unit_cost_eur=snap.weighted_avg_cost_eur,
                                        hs_rate_eur=_hs_rate_eur,
                                        user=request.user,
                                        notes=f"Barkodlu ürün silme (stok düşümü): {product.barcode}",
                                    )
                                except InsufficientStockError:
                                    pass

                    # 4. GoldPurchases + Products soft-delete
                    gp.is_deleted = True
                    gp.save(update_fields=['is_deleted'])

                    product.is_deleted = True
                    product.barcode = ''
                    product.save(update_fields=['is_deleted', 'barcode'])

            # Dashboard "Mağaza Varlıkları" cache'ini temizle (snap değişmese
            # bile — ör. ürün hiç stoğu olmayan bir kayıttı — cache'in bayat
            # kalmaması için burada da açıkça invalidation yapıyoruz).
            try:
                from django.core.cache import cache as _cache
                _cache.delete(f"dashboard_assets_summary:{store.id}")
            except Exception:
                pass

            return JsonResponse({'result': True})
        except Exception as e:
            return JsonResponse({'result': False, 'error': True, 'error_msg': str(e)})
    return JsonResponse({'result': False, 'error': True, 'error_msg': 'Geçersiz istek.'})


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_CHANGE_STATUS')
def change_status(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids[]')
        try:
            with transaction.atomic():
                products = Products.objects.filter(id__in=ids)
                for p in products:
                    p.is_active = not p.is_active
                    p.save(update_fields=['is_active'])
                return JsonResponse({'result': True, 'updated': products.count()})
        except Exception as e:
            return JsonResponse({'result': False, 'error': True, 'error_msg': str(e)})
    return JsonResponse({'result': False, 'error': True, 'error_msg': 'Geçersiz istek.'})


@login_required(login_url='login')
def gold_purchases_stats(request):
    # Bu fonksiyon index.html içinde JS ile çağrılıyor
    store = request.user.store
    DEC18_6 = DecimalField(max_digits=18, decimal_places=6)
    # FAZ 44 — 1.05 EŞİK KURALI (gold_purchases_index ile aynı SSOT).
    buy_price_cast = Cast(F('product__buy_price_hs'), DEC18_6)
    gram_cast = Cast(F('product__gram'), DEC18_6)
    has_expr = Case(
        When(product__buy_price_hs__gt=Decimal('1.05'), then=buy_price_cast),
        default=ExpressionWrapper(buy_price_cast * gram_cast, output_field=DEC18_6),
        output_field=DEC18_6,
    )
    zero_dec = Value(Decimal('0'), output_field=DEC18_6)
    base_qs = GoldPurchases.objects.filter(is_deleted=False, store=store)

    # FAZ 9.8: Satıldı = is_status=False VEYA product.is_completed=True
    sold_q = Q(is_status=False) | Q(product__is_completed=True)

    agg = base_qs.aggregate(
        has_active=Coalesce(Sum(has_expr, filter=~sold_q, output_field=DEC18_6), zero_dec),
        has_sold=Coalesce(Sum(has_expr, filter=sold_q, output_field=DEC18_6), zero_dec),
        active_count=Count('id', filter=~sold_q),
        sold_count=Count('id', filter=sold_q),
        total_count=Count('id'),
    )

    fmt = lambda x: f"{float(x):.3f}"
    return JsonResponse({
        "result": True,
        "total_active": fmt(agg['has_active']),
        "total_sold": fmt(agg['has_sold']),
        "active_count": agg['active_count'],
        "sold_count": agg['sold_count'],
        "total_count": agg['total_count'],
    })


class RoundDecimal(Func):
    function = 'ROUND'
    template = "%(function)s(%(expressions)s, 2)"
    output_field = DecimalField(decimal_places=2)


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_GOLD_PURCHASES_INDEX')
def gold_purchases_import_from_disk(request):
    excel_path = Path(settings.BASE_DIR) / "3.xlsx"

    if not excel_path.exists():
        return JsonResponse({"result": False, "error": True,
                             "error_msg": f"Dosya yok: {excel_path}"}, status=404)
    try:
        result = import_excel_as_products_and_purchases(
            file_obj_or_path=str(excel_path),
            store_id=request.user.store_id,
            created_by=request.user,
            default_price_currency="HS",
        )
        print("IMPORT RESULT:", result)
        return JsonResponse({"result": True, **result})
    except Exception as exc:
        print(exc)
        return JsonResponse({"result": False, "error": True, "error_msg": str(exc)}, status=500)


@login_required(login_url='login')
@require_http_methods(["POST"])
def gold_purchases_import_excel(request):
    """
    Kullanıcının yüklediği Excel dosyasından barkodlu ürün toplu import işlemi.
    Mevcut import_excel_as_products_and_purchases() fonksiyonunu kullanır.
    """
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({
            'result': False, 'error_msg': 'Lütfen bir Excel dosyası seçiniz.'
        }, status=400)

    allowed_ext = ('.xlsx', '.xls')
    if not excel_file.name.lower().endswith(allowed_ext):
        return JsonResponse({
            'result': False,
            'error_msg': 'Sadece .xlsx ve .xls dosyaları kabul edilir.'
        }, status=400)

    try:
        result = import_excel_as_products_and_purchases(
            file_obj_or_path=excel_file,
            store_id=request.user.store_id,
            created_by=request.user,
            default_price_currency="HS",
        )
        return JsonResponse({'result': True, **result})
    except Exception as exc:
        return JsonResponse({
            'result': False, 'error_msg': str(exc)
        }, status=500)


def print_barcode(request):
    product_ids = request.GET.get('ids', '').split(',')

    # Seçilen ürünleri getir
    products = GoldPurchases.objects.filter(id__in=product_ids).select_related('product')

    # Mağaza etiket ayarından RFID Modunu oku (varsayılan True — mevcut müşteriler korunur)
    rfid_mode = True
    try:
        store = getattr(request.user, 'store', None)
        if store:
            _ls = StoreLabelSettings.objects.filter(store=store).first()
            if _ls is not None:
                rfid_mode = getattr(_ls, 'rfid_mode', True)
                if rfid_mode is None:
                    rfid_mode = True
    except Exception:
        pass

    zpl_data = []

    for item in products:
        product = item.product

        # Eğer ürünün RFID kodu yoksa oluştur ve kaydet
        if not product.rfid_code:
            product.rfid_code = generate_rfid_hex()
            product.save()

        # Değerleri ZPL için hazırla
        rfid_code = product.rfid_code
        barcode = product.barcode
        barcode_title = product.store.barcode_title
        ad = product.name[:15]  # ZPL'de taşmasın diye kısaltılabilir
        gram = f"{product.gram:.2f}".replace('.', ',')

        # ^RFW sadece RFID'li yazıcılarda gönderilmeli
        rfw_line = f"^RFW,H,2,12,1^FD{rfid_code}^FS" if rfid_mode else ""

        # ZPL Şablonu
        zpl = f"""
        ^XA
        ^BY1,3,30^FT165,127^BCN,,N,N^FH\\^FD{barcode}^FS
        {rfw_line}
        ^FT180,90^A0N,20,20^FH\\^CI28^FDS{barcode_title}^FS
        ^FT180,145^A0N,20,20^FH\\^CI28^FD{barcode}^FS
        ^FT180,175^A0N,20,20^FH\\^CI28^FD{gram}gr^FS
        ^FT180,220^A0N,20,20^FH\\^CI28^FD{product.jewelry_type}^FS
        ^XZ
        """
        zpl_data.append(zpl)

    # Bu ZPL verisini yazıcıya gönderen servise veya template'e iletin
    return render(request, 'management/gold_purchases/print-barcode.html', {'zpl_list': zpl_data})


@login_required(login_url='login')
def print_barcode_normal(request):
    try:
        ids_param = request.GET.get('ids', '')
        if not ids_param:
            return render(request, 'management/gold_purchases/print-barcode.html', {
                'products': [], 'layout_mode': 'STANDARD',
                'config': default_small_config(),
                'barcode_lines': {'x': 200, 'y': 127, 'height': 35, 'visible': True},
            })

        product_ids = [pid.strip() for pid in ids_param.split(',') if pid.strip()]

        records = GoldPurchases.objects.filter(
            id__in=product_ids
        ).select_related(
            'product', 'product__store', 'product__diamond_detail', 'product__watch_detail', 'supplier'
        ).prefetch_related('product__diamond_detail__stones')

        # Mağaza etiket ayarlarını oku — tüm dinamik tasarım verisi dahil
        bottom_left_type = 'MILYEM'
        layout_mode = 'STANDARD'
        _active_sz = 'small'
        # Tüm material'lar için defaults + DB merge
        configs_by_material = {
            'GOLD':    {'small': default_small_config(),         'large': default_large_config()},
            'DIAMOND': {'small': default_diamond_small_config(), 'large': default_diamond_large_config()},
            'WATCH':   {'small': default_watch_small_config(),   'large': default_watch_large_config()},
        }
        barcode_lines = {'x': 200, 'y': 127, 'height': 35, 'visible': True}
        label_settings = None
        try:
            store = getattr(request.user, 'store', None)
            if store:
                label_settings = StoreLabelSettings.objects.filter(store=store).first()
                if label_settings:
                    bottom_left_type = label_settings.label_bottom_left_type or 'MILYEM'
                    layout_mode = label_settings.label_layout_mode or 'STANDARD'
                    _active_sz = label_settings.active_size or 'small'

                    def _merge(default_cfg, saved):
                        merged = dict(default_cfg)
                        if saved:
                            merged.update(saved)
                        return merged

                    configs_by_material['GOLD']['small']    = _merge(default_small_config(),         label_settings.small_design)
                    configs_by_material['GOLD']['large']    = _merge(default_large_config(),         label_settings.large_design)
                    configs_by_material['DIAMOND']['small'] = _merge(default_diamond_small_config(), getattr(label_settings, 'diamond_small_design', None))
                    configs_by_material['DIAMOND']['large'] = _merge(default_diamond_large_config(), getattr(label_settings, 'diamond_large_design', None))
                    configs_by_material['WATCH']['small']   = _merge(default_watch_small_config(),   getattr(label_settings, 'watch_small_design', None))
                    configs_by_material['WATCH']['large']   = _merge(default_watch_large_config(),   getattr(label_settings, 'watch_large_design', None))

                    barcode_lines = {
                        'x': label_settings.barcode_lines_x,
                        'y': label_settings.barcode_lines_y,
                        'height': label_settings.barcode_lines_height,
                        'visible': label_settings.barcode_lines_visible,
                    }
        except Exception:
            pass

        # ── Print ölçeği: x/y offset + font hesapla ──
        _preview_w = 550 if _active_sz == 'large' else 400
        _scale = 75.0 / _preview_w
        _defaults_by_material = {
            'GOLD':    default_large_config()         if _active_sz == 'large' else default_small_config(),
            'DIAMOND': default_diamond_large_config() if _active_sz == 'large' else default_diamond_small_config(),
            'WATCH':   default_watch_large_config()   if _active_sz == 'large' else default_watch_small_config(),
        }

        # Her material config'ine print_dx/print_dy/print_font enjekte et
        for _mat, _pair in configs_by_material.items():
            _cfg_for_size = _pair[_active_sz]
            _defaults_ref = _defaults_by_material[_mat]
            for _key, _val in _cfg_for_size.items():
                if isinstance(_val, dict) and 'x' in _val:
                    _dv = _defaults_ref.get(_key, {})
                    _val['print_dx'] = round((_val.get('x', 0) - _dv.get('x', 0)) * _scale)
                    _val['print_dy'] = round((_val.get('y', 0) - _dv.get('y', 0)) * _scale)
                    _val['print_font'] = max(6, round(_val.get('font', 20) / 2))

        # GOLD config'i "config" değişkenine ata — eski template uyumu
        config = configs_by_material['GOLD'][_active_sz]

        # Barkod çizgileri offset
        barcode_lines['print_dx'] = round((barcode_lines.get('x', 200) - 200) * _scale)
        barcode_lines['print_dy'] = round((barcode_lines.get('y', 127) - 127) * _scale)

        products = []
        for item in records:
            p = item.product
            if not p:
                continue

            # Firma adı fallback zinciri
            barcode_title = 'KUYUM PLUS'
            try:
                if p.store:
                    barcode_title = p.store.barcode_title or getattr(p.store, 'title', None) or 'KUYUM PLUS'
            except Exception:
                pass

            # Sayısal değerleri güvenli formatlama
            sale_price_hs = f"{float(p.sale_price_hs or 0):.2f}"
            buy_price_hs = f"{float(p.buy_price_hs or 0):.2f}"
            gram = f"{float(p.gram or 0):.2f}"
            gold_rate = f"{float(p.gold_rate or 0):.0f}" if p.gold_rate else ""

            # Sol alt köşe verisi: Milyem veya Maliyet
            if bottom_left_type == 'MALIYET':
                bottom_left_val = f"{buy_price_hs} Has"
            else:
                try:
                    milyem_int = int(float(p.product_mileage or 0))
                    bottom_left_val = f"{milyem_int} M"
                except (ValueError, TypeError):
                    bottom_left_val = "0 M"

            # Ürünün material_type'ına göre kendi config'ini ve material'a özel alanları çöz
            product_material = getattr(p, 'material_type', None) or 'GOLD'
            if product_material not in configs_by_material:
                product_material = 'GOLD'
            product_config = configs_by_material[product_material][_active_sz]

            # Diamond detayları — özet boşsa DiamondStone'dan auto-derive (2026-04-28)
            dd = getattr(p, 'diamond_detail', None)
            wd = getattr(p, 'watch_detail', None)
            dl = _resolve_diamond_label_data(p) if product_material == 'DIAMOND' else None

            # DIAMOND için fiyat alanı: sale_price_hs (GOLD) yerine helper'dan üretilen
            # döviz bazlı fiyat string'i (suffix dahil). Template `product__sale_price_hs`
            # değişkeniyle döviz fiyatını render eder; geriye uyum için aynı anahtar kullanıldı.
            # FAZ DIA-LBL (2026-04-28): show_currency=False → döviz simgesi gösterilmez.
            if dl is not None:
                _price_cfg = (product_config or {}).get('price', {}) or {}
                _show_ccy  = bool(_price_cfg.get('show_currency', True))
                sale_price_display = dl['price_with_ccy'] if _show_ccy else dl['sale_price_str']
            else:
                sale_price_display = sale_price_hs

            products.append({
                'material_type': product_material,
                'config': product_config,
                'barcode_title': barcode_title,
                'product__barcode': p.barcode or '',
                'product__sale_price_hs': sale_price_display,
                'product__buy_price_hs': buy_price_hs,
                'product__gram': gram,
                'product__gold_rate': gold_rate,
                'bottom_left_val': bottom_left_val,
                'supplier__company_name': item.supplier.company_name if item.supplier else '',
                'jewelry_type': p.jewelry_type or '',
                'ring_size': p.ring_size or '',
                # Diamond alanları (helper kullanır; özet boşsa DiamondStone fallback)
                'carat_weight':    (dl['total_carat_str']  if dl else ''),
                'color_grade':     (dl['color_grade']      if dl else ''),
                'clarity_grade':   (dl['clarity_grade']    if dl else ''),
                'cut_grade':       (dl['cut_grade']        if dl else ''),
                'certificate_lab': (dl['certificate_lab']  if dl else ''),
                'certificate_no':  (dl['certificate_no']   if dl else ''),
                # Montür altın bilgileri (opsiyonel; default visible=False)
                'mount_karat':     (dl['mount_karat']      if dl else ''),
                'mount_gram':      (dl['mount_gram_str']   if dl else ''),
                # Watch alanları
                'brand':         (wd.brand or '') if wd else '',
                'model_name':    (wd.model_name or '') if wd else '',
                'reference_no':  (wd.reference_no or '') if wd else '',
                'movement_type': (wd.movement_type or '') if wd else '',
                'condition':     (wd.condition or '') if wd else '',
                'box_papers':    (wd.box_papers or '') if wd else '',
            })

        return render(request, 'management/gold_purchases/print-barcode.html', {
            'products': products,
            'layout_mode': layout_mode,
            'config': config,
            'barcode_lines': barcode_lines,
        })

    except Exception as e:
        import traceback
        return HttpResponse(
            f"<pre>HATA: {e}\n\n{traceback.format_exc()}</pre>",
            status=500
        )


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_GOLD_PURCHASES_INDEX')
def gold_purchases_export(request):
    try:
        store = request.user.store

        fmt = (request.GET.get('format') or 'csv').strip().lower()
        if fmt not in ('csv', 'excel', 'pdf'):
            fmt = 'csv'

        gold_rate_param = (request.GET.get('gold_rate') or '').strip()
        date_from = (request.GET.get('date_from') or '').strip()
        date_to = (request.GET.get('date_to') or '').strip()
        status_filter = (request.GET.get('status_filter') or 'all').strip().lower()
        search_value = (request.GET.get('search') or request.GET.get('search[value]') or '').strip()

        order_column = (request.GET.get('order_column') or 'created_on').strip()
        order_dir = (request.GET.get('order_dir') or 'desc').strip().lower()

        allowed_order = {
            'product__barcode': 'product__barcode',
            'supplier__company_name': 'supplier__company_name',
            'product__jewelry_type': 'product__jewelry_type',
            'product__gram': 'product__gram',
            'product__buy_price_hs': 'product__buy_price_hs',
            'product__sale_price_hs': 'product__sale_price_hs',
            'product__profit': 'product__profit',
            'is_status': 'is_status',
            'created_on': 'created_on',
            'product__is_active': 'product__is_active',
        }

        order_field = allowed_order.get(order_column, 'created_on')
        if order_dir == 'desc':
            order_field = '-' + order_field

        qs = GoldPurchases.objects.filter(is_deleted=False, store=store)

        # FAZ 9.8: Satıldı = is_status=False VEYA product.is_completed=True
        sold_q = Q(is_status=False) | Q(product__is_completed=True)
        if status_filter == 'on_shelf':
            qs = qs.exclude(sold_q)
        elif status_filter == 'sold':
            qs = qs.filter(sold_q)

        df = _parse_tr_date(date_from) if date_from else None
        dt_ = _parse_tr_date(date_to) if date_to else None
        if df:
            qs = qs.filter(created_on__date__gte=df)
        if dt_:
            qs = qs.filter(created_on__date__lte=dt_)

        if gold_rate_param:
            qs = qs.filter(product__gold_rate=gold_rate_param)

        if search_value:
            q_basic = (Q(product__name__icontains=search_value) | Q(product__barcode__icontains=search_value))
            qs = qs.annotate(
                name_tr_norm=Lower(
                    Replace(
                        Replace(F('product__jewelry_type'), Value('İ'), Value('i')),
                        Value('I'), Value('ı')
                    )
                )
            ).filter(q_basic | Q(name_tr_norm__contains=_tr_lower(search_value)))

        rows = qs.values(
            'product__barcode',
            'supplier__company_name',
            'product__jewelry_type',
            'product__gram',
            'product__buy_price_hs',
            'product__sale_price_hs',
            'product__profit',
            'is_status',
            'product__is_completed',
            'created_on',
            'product__is_active',
        ).order_by(order_field)

        now_str = timezone.now().strftime('%Y%m%d-%H%M')

        headers = [
            'Barkod',
            'Tedarikçi',
            'Takı Tipi',
            'Gram',
            'Maliyet (Has)',
            'Satış (Has)',
            'Kar (%)',
            'Statü',
            'Tarih',
            'Durum',
        ]

        data_list = []
        for r in rows:
            # FAZ 9.8: Satıldı = is_status=False VEYA product.is_completed=True
            is_sold = (r.get('is_status') is False) or (r.get('product__is_completed') is True)
            stat_text = 'Satıldı' if is_sold else 'Tezgahta'
            is_active = r.get('product__is_active') is True
            active_text = 'Aktif' if is_active else 'Pasif'
            created_on = r.get('created_on')
            created_text = created_on.astimezone(timezone.get_current_timezone()).strftime(
                '%d/%m/%Y %H:%M') if created_on else ''

            def f3(x):
                try:
                    return ('%.3f' % float(x)) if x is not None else '0.000'
                except Exception:
                    return '0.000'

            def f2(x):
                try:
                    return ('%.2f' % float(x)) if x is not None else ''
                except Exception:
                    return ''

            data_list.append([
                (r.get('product__barcode') or ''),
                (r.get('supplier__company_name') or ''),
                (r.get('product__jewelry_type') or ''),
                f3(r.get('product__gram')),
                f3(r.get('product__buy_price_hs')),
                f3(r.get('product__sale_price_hs')),
                f2(r.get('product__profit')),
                stat_text,
                created_text,
                active_text,
            ])

        if fmt == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = 'Rapor'

                ws.append(headers)
                for row in data_list:
                    ws.append(row)

                for i in range(1, len(headers) + 1):
                    col = get_column_letter(i)
                    ws.column_dimensions[col].width = 18

                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="barkodlu-urun-raporu-' + now_str + '.xlsx"'
                wb.save(response)
                return response
            except Exception:
                fmt = 'csv'

        if fmt == 'pdf':
            try:
                from django.template.loader import render_to_string
                from weasyprint import HTML

                html = render_to_string('management/gold_purchases/export_pdf.html', {
                    'headers': headers,
                    'rows': data_list,
                    'title': 'Barkodlu Ürün Raporu',
                    'generated_at': timezone.now(),
                })

                pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
                response = HttpResponse(pdf_bytes, content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="barkodlu-urun-raporu-' + now_str + '.pdf"'
                return response
            except Exception:
                html_rows = ''.join(
                    ['<tr>' + ''.join(['<td>' + (str(c) if c is not None else '') + '</td>' for c in row]) + '</tr>' for
                     row in data_list])
                html_head = ''.join(['<th>' + h + '</th>' for h in headers])
                html = '<html><head><meta charset="utf-8"><title>Barkodlu Ürün Raporu</title></head><body>' \
                       '<h3>Barkodlu Ürün Raporu</h3>' \
                       '<table border="1" cellspacing="0" cellpadding="6"><thead><tr>' + html_head + '</tr></thead><tbody>' + html_rows + '</tbody></table>' \
                                                                                                                                          '</body></html>'
                response = HttpResponse(html, content_type='text/html; charset=utf-8')
                response['Content-Disposition'] = 'attachment; filename="barkodlu-urun-raporu-' + now_str + '.html"'
                return response

        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="barkodlu-urun-raporu-' + now_str + '.csv"'
        response.write('\ufeff')

        writer = csv.writer(response, delimiter=';')
        writer.writerow(headers)
        for row in data_list:
            writer.writerow(row)
        return response

    except Exception as e:
        return JsonResponse({'result': False, 'error': True, 'error_msg': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def mark_as_printed(request):
    ids = request.POST.getlist('ids[]')
    if not ids:
        return JsonResponse({'result': False, 'message': 'ID listesi boş.'})

    try:
        with transaction.atomic():
            GoldPurchases.objects.filter(id__in=ids, store=request.user.store).update(is_labeled=True)
        return JsonResponse({'result': True})
    except Exception as e:
        return JsonResponse({'result': False, 'message': str(e)})


# ═══════════════════════════════════════════════════════════════════════════════
# GÖREV 1: Ürün Detay Endpoint (Düzenle butonu için)
# ═══════════════════════════════════════════════════════════════════════════════

@login_required(login_url='login')
def get_details(request):
    """Düzenle butonu tıklandığında ürünün tüm bilgilerini döner."""
    record_id = request.GET.get('id')
    if not record_id:
        return JsonResponse({'result': False, 'error_msg': 'ID gerekli.'})

    gp = get_object_or_404(
        GoldPurchases,
        id=record_id,
        store=request.user.store,
        is_deleted=False
    )
    p = gp.product
    if not p:
        return JsonResponse({'result': False, 'error_msg': 'Ürün bulunamadı.'})

    image_url = ''
    if p.image:
        try:
            image_url = p.image.url
        except Exception:
            image_url = ''

    return JsonResponse({
        'result': True,
        'product_id': str(p.id),
        'jewelry_type': p.jewelry_type or '',
        'gram': str(p.gram or ''),
        'gold_rate': str(int(float(p.gold_rate))) if p.gold_rate else '',
        'product_mileage': str(p.product_mileage or ''),
        'labor_mileage': str(p.labor_mileage or ''),
        'piece_labor': str(p.piece_labor or ''),
        'buy_price_hs': str(p.buy_price_hs or ''),
        'sale_price_hs': str(p.sale_price_hs or ''),
        'profit': str(p.profit or ''),
        'supplier_id': str(gp.supplier_id) if gp.supplier_id else '',
        'ring_size': p.ring_size or '',
        'image': image_url,
        'barcode': p.barcode or '',
        'rfid_code': p.rfid_code or '',
    })


# ═══════════════════════════════════════════════════════════════════════════════
# GÖREV 2: ProductCategory CRUD (Takı Tipleri Yönetimi)
# ═══════════════════════════════════════════════════════════════════════════════

@login_required(login_url='login')
def category_list(request):
    """Mağazaya ait aktif kategorileri JSON olarak döner."""
    store = request.user.store
    cats = ProductCategory.objects.filter(
        store=store, is_deleted=False, is_active=True
    ).order_by('name').values('id', 'name', 'barcode_prefix')
    return JsonResponse({'result': True, 'categories': list(cats)})


@login_required(login_url='login')
@require_http_methods(["POST"])
def category_add(request):
    """Yeni takı tipi kategorisi ekler."""
    store = request.user.store
    name = (request.POST.get('name') or '').strip()
    barcode_prefix = (request.POST.get('barcode_prefix') or '').strip().upper()

    if not name:
        return JsonResponse({'result': False, 'error_msg': 'Kategori adı zorunludur.'})
    if not barcode_prefix:
        return JsonResponse({'result': False, 'error_msg': 'Barkod kısayolu (prefix) zorunludur.'})
    if len(barcode_prefix) > 10:
        return JsonResponse({'result': False, 'error_msg': 'Prefix en fazla 10 karakter olabilir.'})

    # Aynı prefix bu mağazada var mı?
    if ProductCategory.objects.filter(store=store, barcode_prefix=barcode_prefix, is_deleted=False).exists():
        return JsonResponse({'result': False, 'error_msg': f'"{barcode_prefix}" kısayolu zaten kullanılıyor.'})

    # Aynı isim bu mağazada var mı?
    if ProductCategory.objects.filter(store=store, name=name, is_deleted=False).exists():
        return JsonResponse({'result': False, 'error_msg': f'"{name}" adında bir kategori zaten var.'})

    cat = ProductCategory.objects.create(
        store=store,
        name=name,
        barcode_prefix=barcode_prefix,
    )
    return JsonResponse({
        'result': True,
        'category': {
            'id': str(cat.id),
            'name': cat.name,
            'barcode_prefix': cat.barcode_prefix,
        }
    })


@login_required(login_url='login')
@require_http_methods(["POST"])
def category_delete(request):
    """Kategoriyi soft-delete yapar."""
    store = request.user.store
    cat_id = request.POST.get('id')
    if not cat_id:
        return JsonResponse({'result': False, 'error_msg': 'ID gerekli.'})

    try:
        cat = ProductCategory.objects.get(id=cat_id, store=store, is_deleted=False)
        cat.is_deleted = True
        cat.save(update_fields=['is_deleted'])
        return JsonResponse({'result': True})
    except ProductCategory.DoesNotExist:
        return JsonResponse({'result': False, 'error_msg': 'Kategori bulunamadı.'})


# ═══════════════════════════════════════════════════════════════════════════════
# GÖREV 3: BarcodeTemplate CRUD (Veri Giriş Şablonları)
# ═══════════════════════════════════════════════════════════════════════════════

@login_required(login_url='login')
def template_list(request):
    """Mağazaya ait aktif şablonları JSON olarak döner.

    FAZ E UI FIX (2026-04-23):
      - ?material_type=GOLD/DIAMOND/WATCH query parametresi ile filtreleme.
      - material_type boş/eksikse TÜM şablonlar döner (geriye dönük uyum).
      - extra_data alanı da dönüşe eklendi (Pırlanta/Saat'e özel alanlar).
    """
    store = request.user.store
    material_type = (request.GET.get('material_type') or '').upper().strip()

    qs = BarcodeTemplate.objects.filter(
        store=store, is_deleted=False, is_active=True
    )
    if material_type in ('GOLD', 'SILVER', 'DIAMOND', 'WATCH'):
        qs = qs.filter(material_type=material_type)

    templates = qs.order_by('template_name').values(
        'id', 'template_name', 'material_type', 'jewelry_type', 'gold_rate',
        'product_mileage', 'labor_mileage', 'piece_labor', 'profit',
        'ring_size', 'process_supplier_ledger', 'supplier_id', 'extra_data',
    )
    result = []
    for t in templates:
        t['product_mileage'] = str(t['product_mileage'])
        t['labor_mileage'] = str(t['labor_mileage'])
        t['piece_labor'] = str(t['piece_labor'])
        t['profit'] = str(t['profit'])
        t['id'] = str(t['id'])
        t['supplier_id'] = str(t['supplier_id']) if t['supplier_id'] else ''
        # extra_data zaten dict; JSONField None dönerse boş dict'e çevir
        if t.get('extra_data') is None:
            t['extra_data'] = {}
        result.append(t)
    return JsonResponse({'result': True, 'templates': result})


@login_required(login_url='login')
@require_http_methods(["POST"])
def template_save(request):
    """Yeni barkod şablonu kaydeder.

    FAZ E UI FIX (2026-04-23):
      - material_type POST parametresi okunur (default 'GOLD').
      - Aynı template_name + aynı material_type varsa güncellenir
        (önceki davranış sadece isim bazlıydı → pırlanta şablonu altın
        şablonunun üstüne yazabiliyordu; bu bug engellendi).
      - extra_data POST'tan JSON string olarak gelir; Pırlanta/Saat'e
        özel alanlar (mount_karat, sale_currency, ...) buraya yazılır.
    """
    import json  # Lokal import — extra_data parse için

    store = request.user.store
    template_name = (request.POST.get('template_name') or '').strip()

    if not template_name:
        return JsonResponse({'result': False, 'error_msg': 'Şablon adı zorunludur.'})

    material_type = (request.POST.get('material_type') or 'GOLD').upper().strip()
    if material_type not in ('GOLD', 'SILVER', 'DIAMOND', 'WATCH'):
        material_type = 'GOLD'

    # extra_data: Pırlanta/Saat'e özel JSON payload
    extra_raw = request.POST.get('extra_data') or ''
    try:
        extra_data = json.loads(extra_raw) if extra_raw else {}
        if not isinstance(extra_data, dict):
            extra_data = {}
    except (ValueError, TypeError):
        extra_data = {}

    # Aynı isim + aynı material_type varsa güncelle (material_type izolasyonu)
    existing = BarcodeTemplate.objects.filter(
        store=store,
        template_name=template_name,
        material_type=material_type,
        is_deleted=False,
    ).first()

    supplier_id = request.POST.get('supplier_id') or None

    ring_size = (request.POST.get('ring_size') or '').strip()
    process_supplier_ledger = request.POST.get('process_supplier_ledger') in ('true', 'on', '1', 'True')

    if existing:
        existing.material_type = material_type
        existing.jewelry_type = request.POST.get('jewelry_type', '')
        existing.gold_rate = request.POST.get('gold_rate', '')
        existing.product_mileage = parse_decimal_locale(request.POST.get('product_mileage'))
        existing.labor_mileage = parse_decimal_locale(request.POST.get('labor_mileage'))
        existing.piece_labor = parse_decimal_locale(request.POST.get('piece_labor'))
        existing.profit = parse_decimal_locale(request.POST.get('profit'))
        existing.ring_size = ring_size
        existing.process_supplier_ledger = process_supplier_ledger
        existing.supplier_id = supplier_id
        existing.extra_data = extra_data
        existing.save()
        tmpl = existing
    else:
        tmpl = BarcodeTemplate.objects.create(
            store=store,
            template_name=template_name,
            material_type=material_type,
            jewelry_type=request.POST.get('jewelry_type', ''),
            gold_rate=request.POST.get('gold_rate', ''),
            product_mileage=parse_decimal_locale(request.POST.get('product_mileage')),
            labor_mileage=parse_decimal_locale(request.POST.get('labor_mileage')),
            piece_labor=parse_decimal_locale(request.POST.get('piece_labor')),
            profit=parse_decimal_locale(request.POST.get('profit')),
            ring_size=ring_size,
            process_supplier_ledger=process_supplier_ledger,
            supplier_id=supplier_id,
            extra_data=extra_data,
        )

    return JsonResponse({
        'result': True,
        'template': {
            'id': str(tmpl.id),
            'template_name': tmpl.template_name,
            'material_type': tmpl.material_type,
            'jewelry_type': tmpl.jewelry_type,
            'gold_rate': tmpl.gold_rate,
            'product_mileage': str(tmpl.product_mileage),
            'labor_mileage': str(tmpl.labor_mileage),
            'piece_labor': str(tmpl.piece_labor),
            'profit': str(tmpl.profit),
            'ring_size': tmpl.ring_size or '',
            'process_supplier_ledger': tmpl.process_supplier_ledger,
            'supplier_id': str(tmpl.supplier_id) if tmpl.supplier_id else '',
            'extra_data': tmpl.extra_data or {},
        }
    })


@login_required(login_url='login')
@require_http_methods(["POST"])
def template_delete(request):
    """Şablonu soft-delete yapar."""
    store = request.user.store
    tmpl_id = request.POST.get('id')
    if not tmpl_id:
        return JsonResponse({'result': False, 'error_msg': 'ID gerekli.'})

    try:
        tmpl = BarcodeTemplate.objects.get(id=tmpl_id, store=store, is_deleted=False)
        tmpl.is_deleted = True
        tmpl.save(update_fields=['is_deleted'])
        return JsonResponse({'result': True})
    except BarcodeTemplate.DoesNotExist:
        return JsonResponse({'result': False, 'error_msg': 'Şablon bulunamadı.'})


# ═══════════════════════════════════════════════════════════════════════════════
# GÖREV 4: Kategori (Takı Tipi) Bazlı Rapor
# ═══════════════════════════════════════════════════════════════════════════════

@login_required(login_url='login')
def gold_purchases_category_report(request):
    """
    ProductCategory (Takı Tipi) bazında gruplanmış özet rapor döner.
    Her kategori için: toplam adet, tezgahtaki gram/has, satılan gram/has.
    """
    store = request.user.store

    DEC18_6 = DecimalField(max_digits=18, decimal_places=6)
    zero_dec = Value(Decimal('0'), output_field=DEC18_6)

    base_qs = GoldPurchases.objects.filter(is_deleted=False, store=store)

    # FAZ 9.8: Satıldı tanımı
    sold_q = Q(is_status=False) | Q(product__is_completed=True)

    gram_expr = Cast(F('product__gram'), DEC18_6)
    has_expr = Cast(F('product__buy_price_hs'), DEC18_6)

    rows = (
        base_qs
        .values(category=F('product__jewelry_type'))
        .annotate(
            total_count=Count('id'),
            tezgahta_gram=Coalesce(
                Sum(gram_expr, filter=~sold_q, output_field=DEC18_6), zero_dec
            ),
            tezgahta_has=Coalesce(
                Sum(has_expr, filter=~sold_q, output_field=DEC18_6), zero_dec
            ),
            satilan_gram=Coalesce(
                Sum(gram_expr, filter=sold_q, output_field=DEC18_6), zero_dec
            ),
            satilan_has=Coalesce(
                Sum(has_expr, filter=sold_q, output_field=DEC18_6), zero_dec
            ),
        )
        .order_by('category')
    )

    fmt = lambda x: f"{float(x):.3f}"
    data = []
    for r in rows:
        data.append({
            'category': r['category'] or 'Tanımsız',
            'total_count': r['total_count'],
            'tezgahta_gram': fmt(r['tezgahta_gram']),
            'tezgahta_has': fmt(r['tezgahta_has']),
            'satilan_gram': fmt(r['satilan_gram']),
            'satilan_has': fmt(r['satilan_has']),
        })

    return JsonResponse({'result': True, 'data': data})


# ═══════════════════════════════════════════════════════════════════════════════
# GÖREV 6: ZIP Tam Yedekleme — Export & Import
# ═══════════════════════════════════════════════════════════════════════════════

@login_required(login_url='login')
@require_http_methods(["POST"])
def backup_export(request):
    """
    Tam yedek ZIP arşivi oluşturma işlemini Celery'ye gönderir.
    GeneratedReports modeline PENDING kayıt oluşturur ve task_id döner.
    Frontend bu task_id ile /dashboard/check-report-status/<task_id> polling yapar.
    """
    from apps.dashboard.models import GeneratedReports
    from apps.gold_purchases.tasks import generate_backup_zip_task

    store = request.user.store
    result = generate_backup_zip_task.delay(
        store_id=str(store.id),
        user_id=str(request.user.id),
    )

    task_id = result.id
    GeneratedReports.objects.update_or_create(
        task_id=task_id,
        defaults={'report_type': 'BACKUP_ZIP', 'status': 'PENDING'},
    )

    return JsonResponse({
        'result': True,
        'task_id': task_id,
        'message': 'Yedekleme başlatıldı. Hazır olduğunda bildirilecek.',
    })


@login_required(login_url='login')
@require_http_methods(["POST"])
def backup_import(request):
    """
    ZIP arşivinden tam geri yükleme.
    ZIP içeriği:
        urunler.xlsx   — ürün verileri (RFID + görsel dosya adı dahil)
        gorseller/     — ürün görselleri {barkod}.{ext}
    """
    import os
    import shutil
    import tempfile
    import zipfile

    from django.core.files.base import ContentFile

    zip_file = request.FILES.get('zip_file')
    if not zip_file:
        return JsonResponse({
            'result': False, 'error_msg': 'Lütfen bir ZIP dosyası seçiniz.'
        }, status=400)

    if not zip_file.name.lower().endswith('.zip'):
        return JsonResponse({
            'result': False, 'error_msg': 'Sadece .zip dosyaları kabul edilir.'
        }, status=400)

    tmp_dir = None
    try:
        # 1. ZIP'i geçici klasöre aç
        tmp_dir = tempfile.mkdtemp(prefix='kp_import_')
        zip_path = os.path.join(tmp_dir, 'upload.zip')
        with open(zip_path, 'wb') as f:
            for chunk in zip_file.chunks():
                f.write(chunk)

        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(tmp_dir)

        # 2. urunler.xlsx dosyasını bul
        xlsx_path = os.path.join(tmp_dir, 'urunler.xlsx')
        if not os.path.isfile(xlsx_path):
            return JsonResponse({
                'result': False,
                'error_msg': 'ZIP içinde "urunler.xlsx" dosyası bulunamadı.'
            }, status=400)

        # 3. gorseller/ klasöründeki dosyaları indexle (barkod → tam yol)
        gorseller_dir = os.path.join(tmp_dir, 'gorseller')
        image_map = {}
        if os.path.isdir(gorseller_dir):
            for fname in os.listdir(gorseller_dir):
                fpath = os.path.join(gorseller_dir, fname)
                if os.path.isfile(fpath):
                    # Dosya adından barkod kısmını çıkar (uzantı hariç)
                    barcode_key = os.path.splitext(fname)[0]
                    image_map[barcode_key] = fpath

        # 4. Import fonksiyonunu çağır (RFID + image_map desteğiyle)
        result = import_excel_as_products_and_purchases(
            file_obj_or_path=xlsx_path,
            store_id=request.user.store_id,
            created_by=request.user,
            default_price_currency="HS",
            image_map=image_map,
        )

        return JsonResponse({'result': True, **result})

    except Exception as exc:
        return JsonResponse({
            'result': False, 'error_msg': str(exc)
        }, status=500)

    finally:
        # Geçici klasörü temizle
        if tmp_dir and os.path.isdir(tmp_dir):
            try:
                shutil.rmtree(tmp_dir)
            except Exception:
                pass


# ═══════════════════════════════════════════════════════════════════════════════
# FAZ 10: Kategori + Ayar Bazlı Detaylı Rapor
# ═══════════════════════════════════════════════════════════════════════════════

def _mileage_to_ayar(mileage):
    """product_mileage değerini Türkçe ayar etiketine çevirir. Yakın değerler de doğru eşlenir."""
    try:
        mi = int(float(mileage or 0))
    except (ValueError, TypeError):
        return str(mileage or '-')
    # Aralık bazlı eşleme (995 ve 999 ikisi de 24 Ayar gibi yakın değerler)
    if mi >= 990: return '24 Ayar'
    if mi >= 950: return '23 Ayar'
    if mi >= 900: return '22 Ayar'
    if mi >= 850: return '21 Ayar'
    if mi >= 720: return '18 Ayar'
    if mi >= 580: return '14 Ayar'
    if mi >= 410: return '10 Ayar'
    if mi >= 370: return '9 Ayar'
    if mi >= 320: return '8 Ayar'
    return f'{mi} M'


def _fmt_tr(val, decimals=2):
    """
    Türkçe lokal sayı formatı: binlik ayıracı '.', ondalık ayıracı ','
    Örn: 10 → '10,00' | 1000 → '1.000,00' | 19.5 → '19,50'
    """
    try:
        v = float(val or 0)
    except (ValueError, TypeError):
        v = 0.0
    # Önce US locale ile formatla ('1,000.00'), sonra karakterleri değiştir
    s = f"{v:,.{decimals}f}"
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


def _build_detailed_report_rows(store):
    """
    Kategori + Ayar bazında; Tezgahta/Satılan için Adet, Gram, Has (saf altın gramı),
    Maliyet (buy_price_hs) değerlerini döner. Tek ORM sorgusu (FAZ 9.8 uyumlu).
    """
    DEC18_6 = DecimalField(max_digits=18, decimal_places=6)
    zero_dec = Value(Decimal('0'), output_field=DEC18_6)
    base_qs = GoldPurchases.objects.filter(is_deleted=False, store=store)

    # FAZ 9.8
    sold_q = Q(is_status=False) | Q(product__is_completed=True)

    gram_expr = Cast(F('product__gram'), DEC18_6)
    # Has gram = fiziksel gram × milyem / 1000  (saf altın ağırlık eşdeğeri)
    has_gram_expr = ExpressionWrapper(
        Cast(F('product__gram'), DEC18_6)
        * Cast(F('product__product_mileage'), DEC18_6)
        / Value(Decimal('1000'), output_field=DEC18_6),
        output_field=DEC18_6,
    )
    # FAZ 44 — 1.05 EŞİK KURALI:
    # Maliyet (Has cinsinden) = buy_price_hs (legacy total) VEYA buy_price_hs × gram (new unit)
    # 1.05 üstü değer legacy toplamdır → doğrudan kullan.
    # 1.05 ve altı birim demektir → gram ile çarp.
    buy_price_cast = Cast(F('product__buy_price_hs'), DEC18_6)
    maliyet_expr = Case(
        When(product__buy_price_hs__gt=Decimal('1.05'), then=buy_price_cast),
        default=ExpressionWrapper(buy_price_cast * gram_expr, output_field=DEC18_6),
        output_field=DEC18_6,
    )

    rows = (
        base_qs
        .values(
            category=F('product__jewelry_type'),
            mileage=F('product__product_mileage'),
        )
        .annotate(
            tezgahta_count=Count('id', filter=~sold_q),
            tezgahta_gram=Coalesce(Sum(gram_expr, filter=~sold_q, output_field=DEC18_6), zero_dec),
            tezgahta_has=Coalesce(Sum(has_gram_expr, filter=~sold_q, output_field=DEC18_6), zero_dec),
            tezgahta_maliyet=Coalesce(Sum(maliyet_expr, filter=~sold_q, output_field=DEC18_6), zero_dec),
            satilan_count=Count('id', filter=sold_q),
            satilan_gram=Coalesce(Sum(gram_expr, filter=sold_q, output_field=DEC18_6), zero_dec),
            satilan_has=Coalesce(Sum(has_gram_expr, filter=sold_q, output_field=DEC18_6), zero_dec),
            satilan_maliyet=Coalesce(Sum(maliyet_expr, filter=sold_q, output_field=DEC18_6), zero_dec),
        )
        .order_by('category', 'mileage')
    )

    data = []
    for r in rows:
        data.append({
            'category': r['category'] or 'Tanımsız',
            'ayar': _mileage_to_ayar(r['mileage']),
            'tezgahta_count': r['tezgahta_count'] or 0,
            'tezgahta_gram': _fmt_tr(r['tezgahta_gram']),
            'tezgahta_has': _fmt_tr(r['tezgahta_has']),
            'tezgahta_maliyet': _fmt_tr(r['tezgahta_maliyet']),
            'satilan_count': r['satilan_count'] or 0,
            'satilan_gram': _fmt_tr(r['satilan_gram']),
            'satilan_has': _fmt_tr(r['satilan_has']),
            'satilan_maliyet': _fmt_tr(r['satilan_maliyet']),
            # Ham değerler (JS filtreleme/toplam için)
            '_raw_tezgahta_gram': float(r['tezgahta_gram']),
            '_raw_tezgahta_has': float(r['tezgahta_has']),
            '_raw_tezgahta_maliyet': float(r['tezgahta_maliyet']),
            '_raw_satilan_gram': float(r['satilan_gram']),
            '_raw_satilan_has': float(r['satilan_has']),
            '_raw_satilan_maliyet': float(r['satilan_maliyet']),
        })
    return data


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_GOLD_PURCHASES_INDEX')
def get_barcoded_products_report(request):
    """
    Kategori + Ayar bazında detaylı rapor AJAX endpoint'i.
    Adet, Gram, Has (saf altın), Maliyet (Has) değerlerini döner.
    """
    store = request.user.store
    data = _build_detailed_report_rows(store)
    return JsonResponse({'result': True, 'data': data})


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_GOLD_PURCHASES_INDEX')
def export_detailed_report_pdf(request):
    """
    Kategori + Ayar bazlı raporu xhtml2pdf ile PDF olarak indirir (landscape A4).
    Adet, Gram, Has (saf altın), Maliyet (Has) kolonlarıyla detaylı analiz.
    Filtre parametreleri: status (all/tezgahta/satilan), ayar, category_q
    """
    from io import BytesIO
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return HttpResponse('xhtml2pdf kutuphanesi yuklu degil.', status=500)
    from django.template.loader import render_to_string

    store = request.user.store
    all_data = _build_detailed_report_rows(store)

    # Filtre parametreleri
    status_filter = (request.GET.get('status') or 'all').strip().lower()
    ayar_filter = (request.GET.get('ayar') or '').strip()
    category_q = (request.GET.get('category_q') or '').strip()
    category_q_low = _tr_lower(category_q) if category_q else ''

    data = []
    for r in all_data:
        if ayar_filter and r['ayar'] != ayar_filter:
            continue
        if category_q_low and category_q_low not in _tr_lower(r['category']):
            continue
        if status_filter == 'tezgahta' and r['tezgahta_count'] == 0:
            continue
        if status_filter == 'satilan' and r['satilan_count'] == 0:
            continue
        data.append(r)

    # Genel toplamlar (ham değerler üzerinden hesapla, sonra TR formatına çevir)
    total_tezgahta_count = sum(d['tezgahta_count'] for d in data)
    total_satilan_count = sum(d['satilan_count'] for d in data)
    total_tezgahta_gram = sum(d['_raw_tezgahta_gram'] for d in data)
    total_tezgahta_has = sum(d['_raw_tezgahta_has'] for d in data)
    total_tezgahta_maliyet = sum(d['_raw_tezgahta_maliyet'] for d in data)
    total_satilan_gram = sum(d['_raw_satilan_gram'] for d in data)
    total_satilan_has = sum(d['_raw_satilan_has'] for d in data)
    total_satilan_maliyet = sum(d['_raw_satilan_maliyet'] for d in data)

    store_name = (
        getattr(store, 'barcode_title', None)
        or getattr(store, 'title', None)
        or getattr(store, 'name', None)
        or 'Kuyum Plus'
    )
    report_date = timezone.now().strftime('%d/%m/%Y %H:%M')

    # Aktif filtre özeti (raporda görünsün diye)
    filter_labels = []
    if status_filter == 'tezgahta':
        filter_labels.append('Durum: Sadece Tezgahtakiler')
    elif status_filter == 'satilan':
        filter_labels.append('Durum: Sadece Satılanlar')
    if ayar_filter:
        filter_labels.append(f'Ayar: {ayar_filter}')
    if category_q:
        filter_labels.append(f'Takı Tipi: {category_q}')

    html_string = render_to_string(
        'management/gold_purchases/detailed_report_pdf.html',
        {
            'store_name': store_name,
            'report_date': report_date,
            'data': data,
            'filter_labels': filter_labels,
            'total_tezgahta_count': total_tezgahta_count,
            'total_satilan_count': total_satilan_count,
            'total_tezgahta_gram': _fmt_tr(total_tezgahta_gram),
            'total_tezgahta_has': _fmt_tr(total_tezgahta_has),
            'total_tezgahta_maliyet': _fmt_tr(total_tezgahta_maliyet),
            'total_satilan_gram': _fmt_tr(total_satilan_gram),
            'total_satilan_has': _fmt_tr(total_satilan_has),
            'total_satilan_maliyet': _fmt_tr(total_satilan_maliyet),
            'grand_total_count': total_tezgahta_count + total_satilan_count,
        },
    )

    pdf_buffer = BytesIO()
    pdf_status = pisa.CreatePDF(BytesIO(html_string.encode('utf-8')), dest=pdf_buffer)
    if pdf_status.err:
        return HttpResponse('PDF olusturulurken hata olustu.', status=500)

    now_str = timezone.now().strftime('%Y%m%d-%H%M')
    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="detayli-rapor-{now_str}.pdf"'
    )
    return response


# ============================================================================
# PIVOT FAZ E (2026-04-23): ÇOKLU MADEN BATCH ENTRY — WATCH ve DIAMOND
# ============================================================================
# Tasarım notları:
#   - SADECE material_type ∈ {WATCH, DIAMOND} kabul eder. GOLD/SILVER için
#     mevcut `gold_purchase_add` veya scraps endpoint'i kullanılır.
#   - Her AJAX çağrısı = 1 ürün. Kuyumcu sayfayı yenilemeden 50 pırlanta
#     girebilir; her giriş bağımsız transaction.atomic().
#   - SupplierLedger currency: get_ledger_currency() + fiat_currency.
#   - Stok adet bazlı: quantity_gram=0, quantity_pieces=max(1, stock_pieces).
#   - Satış fiyatı: WatchDetail/DiamondDetail.sale_currency + sale_price.
#     Satış ekranı bunu günlük kurla TL'ye çevirerek kullanır.
#
# Güvenlik:
#   - Fixed price_currency (HS) hardcode YOK → get_ledger_currency router.
#   - quantity_gram > 0 YOK → StockService._validate_material_type_quantities
#     zaten reddeder, ama burada da garanti altında (Decimal('0') gönderilir).
#   - Products.clean() material_type=WATCH/DIAMOND için metal alanlarını
#     zorla 0'a çeker (DOM spoofing koruması).
# ============================================================================


def _parse_diamond_stones_payload(request):
    """
    Form'dan gelen çoklu taş payload'ını parse eder.

    Beklenen POST format (tekrarlı name'ler):
        stone_role[]          = CENTER | SIDE | ACCENT
        stone_position[]      = 1, 2, 3 ...
        stone_carat[]         = 0.26
        stone_shape[]         = ROUND | PRINCESS ...
        stone_color[]         = D..N
        stone_clarity[]       = FL..I3
        stone_cut[]           = EXCELLENT..POOR
        stone_cert_lab[]      = GIA | IGI | NONE ...
        stone_cert_no[]       = (boş veya sertifika no)

    Dönüş: list[dict] — DiamondStone.objects.create() için hazır dict'ler.
    Eğer payload yoksa boş liste döner; caller summary 4C alanlarını kullanır.
    """
    roles       = request.POST.getlist('stone_role[]')
    types       = request.POST.getlist('stone_type[]')
    positions   = request.POST.getlist('stone_position[]')
    carats      = request.POST.getlist('stone_carat[]')
    shapes      = request.POST.getlist('stone_shape[]')
    colors      = request.POST.getlist('stone_color[]')
    clarities   = request.POST.getlist('stone_clarity[]')
    cuts        = request.POST.getlist('stone_cut[]')
    cert_labs   = request.POST.getlist('stone_cert_lab[]')
    cert_nos    = request.POST.getlist('stone_cert_no[]')

    if not carats:
        return []

    _VALID_TYPES = {'DIAMOND', 'EMERALD', 'RUBY', 'SAPPHIRE', 'PEARL', 'ALEXANDRITE', 'OTHER'}

    out = []
    accepted_idx = 0
    for i, raw_carat in enumerate(carats):
        carat = parse_decimal_locale(raw_carat)
        if carat <= 0:
            continue
        try:
            pos = int(positions[i]) if i < len(positions) and positions[i] else (accepted_idx + 1)
        except (ValueError, TypeError):
            pos = accepted_idx + 1

        stone_type_raw = ((types[i] if i < len(types) else '') or 'DIAMOND').upper().strip()
        stone_type = stone_type_raw if stone_type_raw in _VALID_TYPES else 'DIAMOND'

        explicit_role = ((roles[i] if i < len(roles) else '') or '').upper().strip()
        if explicit_role in {'CENTER', 'SIDE', 'ACCENT'}:
            role = explicit_role
        else:
            role = 'CENTER' if (stone_type == 'DIAMOND' and accepted_idx == 0) else 'SIDE'

        if stone_type == 'DIAMOND':
            color_val = (colors[i] if i < len(colors) else '') or None
        else:
            raw_color = (colors[i] if i < len(colors) else '') or ''
            color_val = raw_color.strip()[:24] or None

        out.append({
            'stone_type':     stone_type,
            'role':           role,
            'position':       pos,
            'carat_weight':   carat,
            'shape':          (shapes[i] if i < len(shapes) else '') or None,
            'color_grade':    color_val,
            'clarity_grade':  (clarities[i] if i < len(clarities) else '') or None,
            'cut_grade':      (cuts[i] if i < len(cuts) else '') or None,
            'certificate_lab': (cert_labs[i] if i < len(cert_labs) else '') or None,
            'certificate_no':  ((cert_nos[i].strip() if i < len(cert_nos) and cert_nos[i] else None) or None),
        })
        accepted_idx += 1
    return out


@login_required(login_url='login')
@role_required('GOLD_PURCHASES_GOLD_PURCHASE_ADD')
@require_http_methods(["POST"])
def multi_material_product_add(request):
    """
    PIVOT: Saat ve Pırlanta için batch AJAX ürün ekleme endpoint'i.

    Akış:
        1. material_type doğrulaması (sadece WATCH/DIAMOND kabul).
        2. Products kaydı oluşturulur (material_type doğru set edilir).
        3. WATCH → WatchDetail, DIAMOND → DiamondDetail + DiamondStone(s).
        4. GoldPurchases tracker kaydı (barkod yazdırma için).
        5. StockService.record_entry (quantity_gram=0, pieces=1+).
        6. supplier_id varsa SupplierLedger + Process (get_ledger_currency
           ile material-aware currency seçimi).
        7. JSON dönüş (barkod, ID, material_type).

    Tüm işlemler tek transaction.atomic() içinde; herhangi bir adım
    başarısız olursa her şey geri alınır.
    """
    store = request.user.store

    # ---- 1. material_type kontrolü ----
    mat_type = (request.POST.get('material_type') or '').upper()
    if mat_type not in (MaterialType.WATCH, MaterialType.DIAMOND):
        return JsonResponse({
            'error': True,
            'error_msg': (
                f"Bu endpoint yalnızca WATCH veya DIAMOND için kullanılabilir. "
                f"Gelen: '{mat_type}'. Altın için /gold-purchases/add kullanın."
            ),
        }, status=400)

    # ---- 1b. POST-QA GUARD: Pırlanta için minimum taş kontrolü ----
    # Frontend validation bypass edildiğinde veya ağ gecikmesiyle boş stones
    # payload'ı gelirse DiamondDetail taşsız olarak kaydedilir ve
    # total_carat_weight property'si 0/None döner. Bu veri butunlugu
    # ihlalini transaction açılmadan önce burada engelliyoruz.
    # _parse_diamond_stones_payload 0 karatli satirlari zaten atladigi icin
    # bos liste donerse gercek taş yok demektir.
    if mat_type == MaterialType.DIAMOND:
        _stones_preview = _parse_diamond_stones_payload(request)
        if not _stones_preview:
            return JsonResponse({
                'error': True,
                'error_msg': (
                    'Pırlanta ürünü en az 1 taş kaydı gerektirir. '
                    'Lütfen taş tablosuna geçerli karat değerine (> 0) sahip '
                    'en az bir taş ekleyin.'
                ),
            }, status=400)

    # ---- 2. Kategori (mat_type'a göre dinamik) ----
    # WATCH  → 'Saat'      kategorisi
    # DIAMOND → 'Pırlanta' kategorisi
    # Hiçbiri bulunmazsa 'Barkodlu Ürünler' fallback (eski davranış).
    # Sol menüde ürünler doğru sekmede gösterilebilsin diye gerekli.
    _category_name_map = {
        MaterialType.WATCH:   'Saat',
        MaterialType.DIAMOND: 'Pırlanta',
    }
    _preferred_category_name = _category_name_map.get(mat_type)
    category = None
    if _preferred_category_name:
        category = Categories.objects.filter(
            name=_preferred_category_name, is_deleted=False
        ).first()
    if category is None:
        category = Categories.objects.filter(
            name='Barkodlu Ürünler', is_deleted=False
        ).first()
    if category is None:
        category = Categories.objects.filter(is_deleted=False).first()

    try:
        with transaction.atomic():
            # ---- 3. Products kaydı ----
            record = Products()
            record.store = store
            record.category = category
            record.created_by_id = request.user.id
            record.created_on = timezone.now()

            # KRİTİK: material_type POST'tan OKUNUR ve KAYDEDİLİR.
            # (Önceki products/views.py::product_add bug'ı burada YOK.)
            record.material_type = mat_type

            record.name = (
                request.POST.get('name')
                or request.POST.get('jewelry_type')
                or ('Saat' if mat_type == 'WATCH' else 'Pırlanta')
            )
            record.jewelry_type = request.POST.get('jewelry_type') or ''
            record.brand = request.POST.get('brand') or ''
            record.description = request.POST.get('description') or ''

            # Satış/Alış TL karşılıkları (opsiyonel; döviz × kur önceden hesaplanmışsa)
            record.sale_price_eur = parse_decimal_locale(
                request.POST.get('sale_price_eur'), default="0.00"
            )
            record.buy_price_eur = parse_decimal_locale(
                request.POST.get('buy_price_eur'), default="0.00"
            )

            # WATCH/DIAMOND için price_currency döviz olarak saklanır
            # (legacy alan, CurrencyChoices içinden seçilir)
            sale_currency_raw = (request.POST.get('sale_currency') or 'USD').upper()
            if sale_currency_raw in ('USD', 'EUR', 'GBP', 'TRY', 'CAD', 'QAR'):
                record.price_currency = sale_currency_raw

            # piece_labor ve fixed_labor_amount: WATCH/DIAMOND için ANLAMLIDIR
            # (Products.clean() bu alanlara dokunmaz)
            record.piece_labor = parse_decimal_locale(
                request.POST.get('piece_labor'), default="0.00"
            )
            record.fixed_labor_amount = parse_decimal_locale(
                request.POST.get('fixed_labor_amount'), default="0.00"
            )

            record.profit = parse_decimal_locale(
                request.POST.get('profit'), default="0.000"
            )

            jt = (record.jewelry_type or '').strip()
            cat_prefix = None
            if jt:
                try:
                    pc = ProductCategory.objects.get(
                        store=store, name=jt, is_deleted=False
                    )
                    cat_prefix = pc.barcode_prefix
                except ProductCategory.DoesNotExist:
                    pass
            if cat_prefix is None and mat_type == MaterialType.DIAMOND:
                cat_prefix = 'PIR'
            record.barcode = generate_barcode(
                jt or mat_type.lower(), store=store, prefix=cat_prefix
            )
            record.rfid_code = generate_rfid_hex()
            record.is_gram_bullion = False  # Adet bazlı ürün

            # Resim
            image = request.FILES.get('image')
            if image:
                filename, processed = process_image(image)
                record.image.save(filename, processed, save=False)

            # Products.clean() otomatik olarak çağrılır (save override);
            # WATCH/DIAMOND için gram/hs/mileage alanlarını 0'a çeker.
            record.save()

            # ---- 4. Uzantı tablosu (WatchDetail veya DiamondDetail + DiamondStone) ----
            sale_price_foreign = parse_decimal_locale(
                request.POST.get('sale_price'), default="0.00"
            )

            if mat_type == MaterialType.WATCH:
                # Geçerli sale_currency değerleri WatchDetail.SaleCurrency'den
                valid_watch_cur = ('USD', 'EUR', 'GBP', 'CHF', 'TRY')
                watch_sale_cur = (
                    sale_currency_raw if sale_currency_raw in valid_watch_cur else 'USD'
                )

                # year_of_mfg: güvenli int parse
                year_raw = (request.POST.get('watch_year_of_mfg') or '').strip()
                year_val = int(year_raw) if year_raw.isdigit() else None

                # case_diameter: güvenli decimal parse
                cd_val = parse_decimal_locale(
                    request.POST.get('watch_case_diameter'), default="0"
                )
                case_diameter = cd_val if cd_val and cd_val > 0 else None

                # warranty_date: None'a düşürülebilir
                warranty_raw = (request.POST.get('watch_warranty_date') or '').strip()

                # ensure_detail_extension sinyali Products.save() sonrası boş
                # WatchDetail yaratmış olabilir. create() yerine update_or_create
                # kullanarak o boş kaydın üstüne yazıyoruz. Double-submit zaten
                # JS validate-first lock ile önleniyor; ayrı guard gerekmez.
                WatchDetail.objects.update_or_create(
                    product=record,
                    defaults=dict(
                        brand=(request.POST.get('watch_brand') or '').strip() or None,
                        model_name=(request.POST.get('watch_model_name') or '').strip() or None,
                        reference_no=(request.POST.get('watch_reference_no') or '').strip() or None,
                        serial_no=(request.POST.get('watch_serial_no') or '').strip() or None,
                        movement_type=(request.POST.get('watch_movement_type') or '').strip() or None,
                        case_material=(request.POST.get('watch_case_material') or '').strip() or None,
                        case_diameter=case_diameter,
                        year_of_mfg=year_val,
                        warranty_date=(warranty_raw or None),
                        box_papers=bool(request.POST.get('watch_box_papers')),
                        condition=(request.POST.get('watch_condition') or 'NEW'),
                        sale_currency=watch_sale_cur,
                        sale_price=sale_price_foreign,
                    ),
                )

            else:  # DIAMOND
                # Geçerli sale_currency değerleri DiamondDetail.SaleCurrency'den
                valid_diamond_cur = ('USD', 'EUR', 'GBP', 'TRY')
                diamond_sale_cur = (
                    sale_currency_raw if sale_currency_raw in valid_diamond_cur else 'USD'
                )

                # Montür gramı (is_mounted türetmek için okunur)
                mount_gram_val = parse_decimal_locale(
                    request.POST.get('mount_gram'), default="0.000"
                )

                # Piece-level summary 4C alanları (tekil taşlı ürün için veya özet)
                summary_carat = parse_decimal_locale(
                    request.POST.get('diamond_carat_weight'), default="0"
                )

                # ensure_detail_extension sinyali Products.save() sonrası boş
                # DiamondDetail yaratmış olabilir. create() yerine update_or_create
                # kullanarak o boş kaydın üstüne yazıyoruz. Double-submit zaten
                # JS validate-first lock ile önleniyor; ayrı guard gerekmez.
                diamond_detail, _ = DiamondDetail.objects.update_or_create(
                    product=record,
                    defaults=dict(
                        mount_metal=(request.POST.get('mount_metal') or 'GOLD_YELLOW'),
                        mount_karat=(request.POST.get('mount_karat') or '18K'),
                        mount_gram=mount_gram_val,
                        sale_currency=diamond_sale_cur,
                        sale_price=sale_price_foreign,
                        carat_weight=(summary_carat if summary_carat > 0 else None),
                        shape=(request.POST.get('diamond_shape') or '').strip() or None,
                        color_grade=(request.POST.get('diamond_color_grade') or '').strip() or None,
                        clarity_grade=(request.POST.get('diamond_clarity_grade') or '').strip() or None,
                        cut_grade=(request.POST.get('diamond_cut_grade') or '').strip() or None,
                        certificate_lab=(request.POST.get('diamond_certificate_lab') or 'NONE'),
                        certificate_no=(request.POST.get('diamond_certificate_no') or '').strip() or None,
                        # P-07 (2026-04-27): Üretici referans kodu (örn: RAI5.61-27).
                        # Opsiyonel — accordion alanı; boş gönderilirse None saklanır.
                        supplier_ref=(request.POST.get('diamond_supplier_ref') or '').strip() or None,
                        fluorescence=(request.POST.get('diamond_fluorescence') or '').strip() or None,
                        is_mounted=bool(mount_gram_val > 0),
                    ),
                )

                # Çoklu taş (D1, D2, ...) payload'ı
                stones_payload = _parse_diamond_stones_payload(request)
                for stone_data in stones_payload:
                    DiamondStone.objects.create(
                        diamond_detail=diamond_detail,
                        **stone_data,
                    )

            # ---- 5. GoldPurchases tracker kaydı (barkod/rapor entegrasyonu) ----
            gp_record = GoldPurchases.objects.create(
                product=record,
                store=store,
                supplier_id=(request.POST.get('supplier_id') or None),
                created_by=request.user,
                is_status=True,
            )

            # ---- 6. Stok girişi — ADET BAZLI (quantity_gram=0) ----
            try:
                stock_pieces = int(request.POST.get('stock_pieces', 1) or 1)
            except (ValueError, TypeError):
                stock_pieces = 1
            stock_pieces = max(1, stock_pieces)  # Garanti: en az 1 adet

            supplier_id = request.POST.get('supplier_id')
            if supplier_id:
                reason = StockLedger.Reason.PURCHASE
                ref_type = 'gold_purchase'
                notes_text = f"Tedarikçi alımı ({mat_type}): {record.name}"
            else:
                reason = StockLedger.Reason.INITIAL
                ref_type = 'initial_stock'
                notes_text = f"Açılış stoğu ({mat_type}): {record.name}"

            # StockService._validate_material_type_quantities() WATCH/DIAMOND
            # için quantity_gram==0 ve quantity_pieces>=1 kuralını zorlar.
            StockService.record_entry(
                product=record,
                store=store,
                quantity_gram=Decimal('0.0000'),
                quantity_pieces=stock_pieces,
                reason=reason,
                ref_type=ref_type,
                ref_id=f"gold_purchase_{gp_record.id}",
                unit_cost_hs=Decimal('0.0000'),
                unit_cost_eur=(record.buy_price_eur or Decimal('0')),
                hs_rate_eur=Decimal('0.0000'),
                user=request.user,
                notes=notes_text,
            )

            # ---- 7. SupplierLedger + Process (opsiyonel; checkbox ile tetiklenir) ----
            process_ledger = request.POST.get('process_supplier_ledger')
            if process_ledger and supplier_id:
                try:
                    supplier = Suppliers.objects.get(id=supplier_id)
                    buy_tl = record.buy_price_eur or Decimal('0')

                    if buy_tl > 0:
                        gp_process_no = generate_process_no()

                        # get_ledger_currency: WATCH/DIAMOND -> fiat (TRY/USD/...)
                        # Metal bazlı (GOLD/SILVER) buradan geçmez — zaten farklı endpoint.
                        # fiat_currency router'a sale_currency olarak verilir.
                        fiat_param = (
                            sale_currency_raw
                            if sale_currency_raw in ('TRY', 'USD', 'EUR', 'GBP', 'CAD', 'QAR')
                            else 'TRY'
                        )
                        try:
                            ledger_ccy = get_ledger_currency(
                                record, fiat_currency=fiat_param
                            )
                        except ValueError:
                            # Geçersiz fiat -> güvenli default TRY
                            ledger_ccy = 'TRY'

                        # Cari borç amount_value: ledger_ccy cinsinden.
                        # Döviz ise sale_price_foreign; TRY ise buy_tl.
                        ledger_amount = (
                            sale_price_foreign
                            if ledger_ccy in ('USD', 'EUR', 'GBP')
                            else buy_tl
                        )

                        SupplierLedger.objects.create(
                            supplier=supplier,
                            product=record,
                            transaction_type=SupplierLedger.ENTRY,
                            quantity_piece=stock_pieces,
                            quantity_gram=Decimal('0'),
                            amount_value=ledger_amount,
                            currency=ledger_ccy,
                            process_no=gp_process_no,
                            description=(
                                f"Alım ({mat_type}): "
                                f"{record.jewelry_type or record.name} - {record.barcode}"
                            ),
                            is_active=True,
                        )

                        # Process kaydı (İşlemler ekranında görünür)
                        Process.objects.create(
                            store=store,
                            process_no=gp_process_no,
                            process_type='RETAIL',
                            transaction_type='PURCHASE',
                            product=record,
                            supplier=supplier,
                            employee=request.user,
                            piece=stock_pieces,
                            gram=Decimal('0'),
                            price_hs=Decimal('0'),
                            unit_price=buy_tl,
                            amount=buy_tl,
                            is_status='COMPLETED',
                            is_deleted=False,
                        )
                except Suppliers.DoesNotExist:
                    pass

            return JsonResponse({
                'result': True,
                'gold_purchase_id': str(gp_record.id),
                'product_id': str(record.id),
                'barcode': record.barcode or '',
                'rfid_code': record.rfid_code or '',
                'material_type': mat_type,
            })

    except ValueError as ve:
        # StockService validation hataları ve parse hataları buraya düşer
        return JsonResponse({'error': True, 'error_msg': str(ve)}, status=400)
    except Exception as e:
        return JsonResponse({'error': True, 'error_msg': str(e)}, status=500)
