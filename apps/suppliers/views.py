from django.db.models import Q
from decimal import Decimal
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.db.models import Sum, DecimalField, Value, Avg
from django.db.models.aggregates import Max
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from xhtml2pdf import pisa

# Modellerini import etmeyi unutma
from apps.activity_logs.views import write_log
from apps.gold_purchases.models import GoldPurchases
from apps.process.models import Process
from apps.roles.decorators import role_required
from apps.suppliers.models import Suppliers, SupplierLedger
from apps.suppliers.services import SupplierDeleteService, SupplierDeleteBlocked


# --- GÖRÜNÜM VE SAYFA ---

@login_required(login_url='login')
@role_required('SUPPLIERS_SUPPLIERS_DETAIL')
def suppliers_detail(request, record_id):
    s = get_object_or_404(Suppliers, id=record_id)

    # 1. Finansal Bakiye (Ledger'dan)
    bal = s.balance_summary().get('HS', {'receivable': 0, 'payable': 0, 'net': 0})

    # 2. Temel Sorgular
    # Process tablosu (Toptan, Ödeme, İade vb.)
    process_qs = Process.objects.filter(supplier=s, is_deleted=False).exclude(is_status='CANCELED')
    # GoldPurchases tablosu (Barkodlu, tekil ürün alımları)
    barcode_qs = GoldPurchases.objects.filter(supplier=s, is_deleted=False, is_active=True)

    # --- KPI: TOPLAM HACİM (Has Altın Gramı) ---

    # DÜZELTME BURADA YAPILDI: 0.0 yerine Decimal('0') kullanıldı.

    # A) Process tablosundan gelen gramlar (Toptan Alış + Stok Girişi)
    proc_vol = process_qs.filter(transaction_type__in=['PURCHASE', 'STOCK_IN']).aggregate(
        t_gram=Coalesce(Sum('gram'), Decimal('0')),
        t_hs=Coalesce(Sum('price_hs'), Decimal('0'))
    )

    # B) Barkodlu ürünlerden gelen gramlar (Products tablosundaki gram üzerinden)
    bar_vol = barcode_qs.aggregate(
        t_gram=Coalesce(Sum('product__gram'), Decimal('0'))
    )

    # Toplam Gram Hacmi (Decimal -> Float dönüşümü yaparak topla)
    total_volume_gram = float(proc_vol['t_gram']) + float(bar_vol['t_gram'])

    # --- KPI: İADE ORANI ---
    total_tx = process_qs.count()
    return_tx = process_qs.filter(transaction_type='RETURN').count()

    return_rate = 0
    if total_tx > 0:
        return_rate = (return_tx / total_tx) * 100

    # --- KPI: TOPLAM İŞLEM SAYISI ---
    # Process satırları + Barkodlu ürün adedi
    total_interaction = total_tx + barcode_qs.count()

    # --- KPI: SON HAREKET ---
    last_proc = process_qs.order_by('-date').first()
    last_bar = barcode_qs.order_by('-created_on').first()

    last_date = None
    # Tarih karşılaştırması için process date (datetime) ile goldpurchases created_on (datetime) kıyaslanır
    if last_proc and last_bar:
        # Her ikisi de varsa en yeniyi al
        if last_proc.date and last_bar.created_on:
            last_date = last_proc.date if last_proc.date > last_bar.created_on else last_bar.created_on
        elif last_proc.date:
            last_date = last_proc.date
        else:
            last_date = last_bar.created_on
    elif last_proc:
        last_date = last_proc.date
    elif last_bar:
        last_date = last_bar.created_on

    # --- ÜRÜN ANALİZİ (TAB 2) ---
    # En çok işlem gören ürün gruplarını Process tablosundan çekiyoruz
    top_products = (process_qs
    .filter(transaction_type__in=['PURCHASE', 'STOCK_IN'], product__isnull=False)
    .values('product__name', 'product__category__name')  # Ürün adına göre grupla
    .annotate(
        total_qty=Sum('piece'),
        total_gr=Sum('gram')
    )
    .order_by('-total_gr')[:5]  # Gramaj olarak en büyük 5 kalem
    )

    # ════════════════════════════════════════════════════════════════════
    # FAZ 2 / KARAR-01 — Karar Şeridi (Decision Strip) Context (2026-04-29)
    # ════════════════════════════════════════════════════════════════════
    # Sayfanın en üst öncelikli karar destek alanı için server-side hesaplama.
    # Server-side tercih edildi çünkü ilk paint'te flicker olmaması ve
    # esnafın sayfayı açar açmaz net pozisyonu görmesi kritik.
    #
    # Bileşenler:
    #   1. Dominant Net (HS) — büyük başlık olarak gösterilir
    #   2. Currency Pills — HS dışındaki açık döviz pozisyonları
    #   3. Relative Last Date — "12 gün önce" insancıl format
    #   4. Açık Pozisyon Sayısı — Faz 4 (açıktan bağlama) tamamlanana kadar 0
    full_balance = s.balance_summary()  # tüm para birimleri için net özet

    # Dominant: HS net pozisyonu
    karar_hs_net = Decimal(str(full_balance.get('HS', {}).get('net', 0)))
    if karar_hs_net > 0:
        karar_dominant_direction = 'RECEIVABLE'   # biz alacaklıyız
    elif karar_hs_net < 0:
        karar_dominant_direction = 'PAYABLE'      # biz borçluyuz
    else:
        karar_dominant_direction = 'BALANCED'

    # Currency Pills — HS dışındaki sıfırdan farklı bakiyeler.
    # Sıralama: HG önce (metal), sonra fiat para birimleri standart sırada.
    PILL_ORDER = {'HG': 1, 'TRY': 2, 'USD': 3, 'EUR': 4, 'GBP': 5, 'CAD': 6, 'QAR': 7}
    karar_currency_pills = []
    for cur, amounts in (full_balance or {}).items():
        if cur == 'HS':
            continue  # dominant'ta zaten gösteriliyor
        cur_net = Decimal(str(amounts.get('net', 0)))
        if cur_net == 0:
            continue
        karar_currency_pills.append({
            'currency': cur,
            'net': cur_net,
            'abs_net': abs(cur_net),
            'direction': 'RECEIVABLE' if cur_net > 0 else 'PAYABLE',
        })
    karar_currency_pills.sort(key=lambda p: PILL_ORDER.get(p['currency'], 99))

    # Relative Last Date — insancıl format ("12 gün önce", "Dün", "3 saat önce")
    karar_relative_last = None
    if last_date:
        now = timezone.now()
        # last_date naive olabilir (eski Process kayıtlarında); güvenli şekilde aware'e çevir
        if timezone.is_naive(last_date):
            try:
                last_date_aware = timezone.make_aware(last_date)
            except Exception:
                last_date_aware = last_date
        else:
            last_date_aware = last_date
        try:
            delta = now - last_date_aware
            days = delta.days
            if days < 0:
                karar_relative_last = 'Yakın zaman'
            elif days == 0:
                hours = delta.seconds // 3600
                if hours == 0:
                    mins = (delta.seconds % 3600) // 60
                    karar_relative_last = f'{mins} dk önce' if mins > 0 else 'Az önce'
                else:
                    karar_relative_last = f'{hours} saat önce'
            elif days == 1:
                karar_relative_last = 'Dün'
            elif days < 30:
                karar_relative_last = f'{days} gün önce'
            elif days < 365:
                karar_relative_last = f'{days // 30} ay önce'
            else:
                karar_relative_last = f'{days // 365} yıl önce'
        except Exception:
            # Tarih hesaplama hatası UI'ı çökertmesin
            karar_relative_last = None

    # Açık Pozisyon Sayısı — Faz 4'te SupplierOpenPosition modeli eklendiğinde
    # buraya gerçek count gelecek. Şimdilik 0 placeholder; template > 0 ise
    # uyarı badge'ini gösterir.
    karar_open_positions = 0
    # TODO FAZ 4: karar_open_positions = SupplierOpenPosition.objects.filter(
    #     supplier=s, status='ACIK'
    # ).count()

    # ════════════════════════════════════════════════════════════════════
    # FAZ 2 / KARAR-02 — Analitik Panel Verileri (2026-04-29)
    # ════════════════════════════════════════════════════════════════════
    # Karar Şeridi'nin altında yer alan, esnafın bu tedarikçiyle olan iş
    # ilişkisini "trend" olarak okuduğu panel için server-side metrikler.
    # Sol kolon: iş bağlamı (hacim trendleri, milyem kalitesi, kategori).
    # Sağ kolon: iletişim & ilişki (record üzerinden direkt okunur, ek
    #           hesaplama gerekmez).
    now_dt = timezone.now()
    date_30_ago = now_dt - timedelta(days=30)
    date_90_ago = now_dt - timedelta(days=90)
    month_start = now_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Son 30 gün hacim (Process PURCHASE/STOCK_IN gramı + GoldPurchases ürün gramı)
    proc_30 = process_qs.filter(
        transaction_type__in=['PURCHASE', 'STOCK_IN'],
        date__gte=date_30_ago,
    ).aggregate(t_gram=Coalesce(Sum('gram'), Decimal('0')))
    bar_30 = barcode_qs.filter(created_on__gte=date_30_ago).aggregate(
        t_gram=Coalesce(Sum('product__gram'), Decimal('0'))
    )
    analitik_volume_30 = float(proc_30['t_gram']) + float(bar_30['t_gram'])

    # Son 90 gün hacim
    proc_90 = process_qs.filter(
        transaction_type__in=['PURCHASE', 'STOCK_IN'],
        date__gte=date_90_ago,
    ).aggregate(t_gram=Coalesce(Sum('gram'), Decimal('0')))
    bar_90 = barcode_qs.filter(created_on__gte=date_90_ago).aggregate(
        t_gram=Coalesce(Sum('product__gram'), Decimal('0'))
    )
    analitik_volume_90 = float(proc_90['t_gram']) + float(bar_90['t_gram'])

    # Ortalama milyem — GoldPurchases.product_mileage > 0 olan kayıtlar
    # üzerinden. 0 milyem genelde "girişi yapılmamış" anlamı taşır,
    # ortalamayı bozmaması için filtrelenir.
    avg_mileage_row = barcode_qs.filter(product__product_mileage__gt=0).aggregate(
        avg_m=Avg('product__product_mileage')
    )
    analitik_avg_mileage = avg_mileage_row['avg_m']  # Decimal veya None

    # Bu ay en çok alınan kategori (Process bazında, gram ağırlıklı top 1)
    top_cat_row = (process_qs
                   .filter(transaction_type__in=['PURCHASE', 'STOCK_IN'],
                           product__isnull=False,
                           product__category__isnull=False,
                           date__gte=month_start)
                   .values('product__category__name')
                   .annotate(total_gr=Sum('gram'))
                   .order_by('-total_gr')
                   .first())
    if top_cat_row and top_cat_row.get('product__category__name'):
        analitik_top_category_name = top_cat_row['product__category__name']
        analitik_top_category_gram = float(top_cat_row['total_gr'] or 0)
    else:
        analitik_top_category_name = None
        analitik_top_category_gram = 0.0

    # İade oranı uyarı eşiği (raporun 1.2'sinde "yüksekse sarı arka plan").
    # %5 üzeri "ortalamanın üstünde" sayılır (mevcut UX-01 ile tutarlı).
    analitik_return_rate_high = round(return_rate, 2) > 5

    ctx = {
        'title': f'{s.company_name} - Detay',
        'record': s,
        # Bakiye
        'receivable_hs': bal['receivable'],
        'payable_hs': bal['payable'],
        'balance_hs': bal['net'],
        # KPI
        'kpi_volume_gram': total_volume_gram,
        'kpi_return_rate': round(return_rate, 2),
        'kpi_total_tx': total_interaction,
        'kpi_last_date': last_date,
        # Karar Şeridi (FAZ 2 / KARAR-01)
        'karar_hs_net': karar_hs_net,
        'karar_hs_abs': abs(karar_hs_net),
        'karar_dominant_direction': karar_dominant_direction,
        'karar_currency_pills': karar_currency_pills,
        'karar_relative_last': karar_relative_last,
        'karar_open_positions': karar_open_positions,
        # Analitik Panel (FAZ 2 / KARAR-02)
        'analitik_volume_30': analitik_volume_30,
        'analitik_volume_90': analitik_volume_90,
        'analitik_avg_mileage': analitik_avg_mileage,
        'analitik_top_category_name': analitik_top_category_name,
        'analitik_top_category_gram': analitik_top_category_gram,
        'analitik_return_rate_high': analitik_return_rate_high,
        # Tablo Verisi
        'top_products': top_products,
    }
    return render(request, 'management/suppliers/detail.html', ctx)


# --- AJAX DATATABLE ---

@login_required(login_url='login')
def get_supplier_process_history(request):
    draw = int(request.GET.get('draw', 1))
    length = int(request.GET.get('length', 10))
    start = int(request.GET.get('start', 0))
    search = request.GET.get('search[value]', '').strip()

    # Sıralama
    order_col_idx = int(request.GET.get('order[0][column]', 0))
    order_dir = '-' if request.GET.get('order[0][dir]') == 'desc' else ''

    col_map = {0: 'date', 1: 'process_no', 2: 'transaction_type', 3: 'product__name', 5: 'price_hs'}
    order_col = col_map.get(order_col_idx, 'date')

    # Filtreler
    sup_id = request.GET.get('supplier_id')
    tx_type = request.GET.get('transaction_type')
    currency = request.GET.get('currency')

    # FAZ 11 / UX-04: Genişletilmiş filtreler (2026-04-24)
    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()
    material_type = (request.GET.get('material_type') or '').strip().upper()
    min_amount_raw = (request.GET.get('min_amount') or '').strip()
    max_amount_raw = (request.GET.get('max_amount') or '').strip()

    supplier = get_object_or_404(Suppliers, id=sup_id)

    # Ana Sorgu: Sadece Process tablosu (Finansal hareketler burada)
    qs = Process.objects.filter(supplier=supplier, is_deleted=False).select_related('product')

    if tx_type:
        if tx_type == 'ENTRY':
            qs = qs.filter(transaction_type__in=['PURCHASE', 'STOCK_IN', 'ORDER_IN'])
        elif tx_type == 'EXIT':
            qs = qs.filter(transaction_type__in=['RETURN', 'PAYMENT', 'SALE'])

    if currency:
        if currency == 'HS':
            qs = qs.filter(Q(product__currency='HS') | Q(price_hs__gt=0))
        elif currency == 'TRY':
            qs = qs.filter(Q(product__currency='TRY') | Q(amount__gt=0))
        else:
            qs = qs.filter(product__currency=currency)

    # FAZ 11 / UX-04: Tarih aralığı (YYYY-MM-DD formatı bekleniyor)
    if date_from:
        try:
            qs = qs.filter(date__date__gte=date_from)
        except (ValueError, TypeError):
            pass  # Geçersiz tarih sessizce yok sayılır
    if date_to:
        try:
            qs = qs.filter(date__date__lte=date_to)
        except (ValueError, TypeError):
            pass

    # FAZ 11 / UX-04: Kategori filtresi (material_type bazlı)
    # Geçerli değerler: GOLD, SILVER, WATCH, DIAMOND, SCRAP, BRACELET
    # SCRAP ve BRACELET, material_type=GOLD + product flag bazlı
    if material_type:
        if material_type == 'SCRAP':
            qs = qs.filter(product__is_scrap=True)
        elif material_type == 'BRACELET':
            qs = qs.filter(product__category__name__icontains='Bilezik')
        elif material_type in ('GOLD', 'SILVER', 'WATCH', 'DIAMOND'):
            qs = qs.filter(product__material_type=material_type)
        # Tanınmayan değer sessizce yok sayılır

    # FAZ 11 / UX-04: Tutar aralığı (HS veya TL bazlı, en yüksek olanı kullanır)
    if min_amount_raw:
        try:
            min_amount = Decimal(min_amount_raw.replace(',', '.'))
            qs = qs.filter(Q(price_hs__gte=min_amount) | Q(amount__gte=min_amount))
        except (Exception,):
            pass
    if max_amount_raw:
        try:
            max_amount = Decimal(max_amount_raw.replace(',', '.'))
            qs = qs.filter(Q(price_hs__lte=max_amount) & Q(amount__lte=max_amount))
        except (Exception,):
            pass

    if search:
        qs = qs.filter(
            Q(process_no__icontains=search) |
            Q(product__name__icontains=search)
        )

    total = qs.count()
    qs = qs.order_by(f'{order_dir}{order_col}')[start:start + length]

    data = []
    for p in qs:
        # Ürün Adı veya İşlem Tipi
        p_name = p.product.name if p.product else p.get_transaction_type_display()

        # Miktar ve Gram Kontrolü
        qty = 0
        is_gram = False
        unit_str = "Ad"

        if p.product:
            if p.product.is_gram_bullion:
                qty = p.gram
                is_gram = True
                unit_str = "Gr"
            else:
                qty = p.piece

        # Birim Fiyat (Basit Hesaplama)
        unit_p = 0
        divider = float(qty) if qty > 0 else 1.0

        # Eğer HS tutarı varsa öncelik HS
        if p.price_hs and p.price_hs > 0:
            unit_p = float(p.price_hs) / divider
        elif p.amount and p.amount > 0:
            unit_p = float(p.amount) / divider

        # ════════════════════════════════════════════════════════════════
        # FAZ 11 / CAT-01 — FİŞ KALEMİ TİPİ ETİKETİ (2026-04-24)
        # ════════════════════════════════════════════════════════════════
        # Esnafa "bu satır neye ait?" cevabını veren read-only computed
        # alan. Process modeli değiştirilmiyor — sadece response zenginleşiyor.
        # ════════════════════════════════════════════════════════════════
        line_item_type = 'OTHER'
        line_item_label = '-'

        if p.bank_account_id is not None and p.product is None:
            line_item_type = 'CASH_PAYMENT'
            line_item_label = 'Nakit / Kasa'
        elif p.product:
            prod = p.product
            mat = (getattr(prod, 'material_type', '') or '').upper()
            is_scrap = bool(getattr(prod, 'is_scrap', False))
            cat_name = (getattr(getattr(prod, 'category', None), 'name', '') or '').lower()

            if is_scrap:
                line_item_type = 'SCRAP_GRAM'
                line_item_label = 'Hurda'
            elif 'bilezik' in cat_name:
                if prod.is_gram_bullion:
                    line_item_type = 'BRACELET_GRAM'
                    line_item_label = 'Bilezik (Gram)'
                else:
                    line_item_type = 'BRACELET_PIECE'
                    line_item_label = 'Bilezik (Adet)'
            elif mat == 'GOLD':
                if prod.is_gram_bullion:
                    line_item_type = 'GOLD_GRAM'
                    line_item_label = 'Altın (Gram)'
                else:
                    line_item_type = 'GOLD_PIECE'
                    line_item_label = 'Altın (Adet)'
            elif mat == 'SILVER':
                line_item_type = 'SILVER_GRAM' if prod.is_gram_bullion else 'SILVER_PIECE'
                line_item_label = 'Gümüş'
            elif mat == 'DIAMOND':
                line_item_type = 'DIAMOND_PIECE'
                line_item_label = 'Elmas / Pırlanta'
            elif mat == 'WATCH':
                line_item_type = 'WATCH_PIECE'
                line_item_label = 'Saat'

        data.append({
            # FAZ 11 / SYNC-04: TR locale tarih formatı (gg.aa.yyyy SS:DD)
            # Tüm uygulama genelinde standart format. Backend ORM sort'u
            # yine timestamp üzerinden yaptığı için sıralama doğru çalışır.
            'date': timezone.localtime(p.date).strftime('%d.%m.%Y %H:%M'),
            'process_no': p.process_no,
            'transaction_type': p.transaction_type,
            'product__name': p_name,
            'quantity': float(qty),
            'unit_str': unit_str,
            'amount_paid_hs': float(p.price_hs or 0),
            'amount_paid_tl': float(p.amount or 0),
            'unit_price_calc': float(unit_p),
            # FAZ 11 / CAT-01
            'line_item_type': line_item_type,
            'line_item_label': line_item_label,
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total,
        'recordsFiltered': total,
        'data': data
    })


# --- DİĞER STANDART VIEWLAR ---

@login_required(login_url='login')
@role_required('SUPPLIERS_SUPPLIERS_VIEW')
def suppliers_view(request):
    write_log(request, 'Tedarikçiler', 'Tedarikçiler Görüntülendi.')
    return render(request, 'management/suppliers/index.html', {'title': 'Tedarikçiler'})


@login_required(login_url='login')
def add_supplier(request):
    sid = request.POST.get('supplier_id')
    if sid:
        rec = get_object_or_404(Suppliers, id=sid)
    else:
        rec = Suppliers(store_id=request.user.store_id)

    for f in ('company_name', 'person_name', 'person_surname', 'email', 'phone', 'company_address', 'post_code'):
        val = request.POST.get(f)
        setattr(rec, f, val if val else None)

    city_id = request.POST.get('city')
    rec.city_id = city_id if city_id else None

    district_id = request.POST.get('district')
    rec.district_id = district_id if district_id else None

    tax_office_id = request.POST.get('tax_office')
    rec.tax_office_id = tax_office_id if tax_office_id else None

    rec.tax_payer_type = request.POST.get('tax_payer_type', 'CORPORATE')
    rec.tax_number = request.POST.get('tax_number')

    # Hesap Türü (Tedarikçi / Çantacı)
    account_type = request.POST.get('account_type')
    if account_type in ('SUPPLIER', 'CANTACI'):
        rec.account_type = account_type

    is_einvoice = request.POST.get('is_einvoice_user')
    rec.is_einvoice_user = True if is_einvoice == 'on' else False

    rec.save()

    if not sid:
        write_log(request, 'Tedarikçiler', f'Tedarikçi Firma Eklendi. ID={rec.id}'.upper())

    return JsonResponse({'result': True, 'supplier_id': rec.id})


@login_required(login_url='login')
@role_required('SUPPLIERS_DELETE')
def delete(request):
    """
    Tedarikçi/Çantacı silme endpoint'i (FAZ TS-2 ile servis tabanlı).

    POST parametreleri:
        ids[]    : Silinecek tedarikçi ID'leri.
        mode     : 'preflight' verilirse silme yapılmaz, sadece rapor döner.
                   Aksi halde silme akışı çalışır.
        confirm  : 'true' verilirse warnings (bakiye, anonimleşme) bypass
                   edilir. Bloklayıcılar (açık process vb.) yine engellenir.

    Yanıtlar:
        200 {result: True}                                — başarılı silme
        200 {result: True, report: {...}}                 — preflight modu
        409 {result: False, blocked: True, errors, warnings} — bloklayıcı / onay bekleniyor

    Geriye dönük uyumluluk:
        Mevcut UI sadece ids[] gönderdiğinde:
          - Bloklayıcı yoksa ve uyarı yoksa: silme yapılır, 200 döner.
          - Uyarı varsa (bakiye / GP) ve confirm=false ise: 409 döner.
            UI tarafı bu durumu yakalayıp kullanıcıya onay modalı göstermeli
            (FAZ TS-3 UI işidir; geçici olarak hata olarak da düşebilir).
    """
    ids = request.POST.getlist('ids[]')
    mode = (request.POST.get('mode') or 'execute').lower()
    confirm = (request.POST.get('confirm') or '').lower() == 'true'

    if not ids:
        return JsonResponse({'result': False, 'error': 'NO_IDS'}, status=400)

    service = SupplierDeleteService(user=request.user)

    if mode == 'preflight':
        report = service.preflight(supplier_ids=ids)
        return JsonResponse({'result': True, 'report': report})

    try:
        result = service.execute(supplier_ids=ids, force=confirm)
    except SupplierDeleteBlocked as exc:
        return JsonResponse(
            {
                'result': False,
                'blocked': True,
                'errors': exc.errors,
                'warnings': exc.warnings,
            },
            status=409,
        )

    # Audit log — view katmanında, request elimizdeyken.
    for d in result.get('details', []):
        write_log(
            request,
            'Tedarikçiler',
            (
                f"TEDARİKÇİ SİLİNDİ. ID={d['supplier_id']} "
                f"NAME={d['supplier_name']} "
                f"LEDGERS_DEACTIVATED={d['ledgers_deactivated']} "
                f"GOLD_PURCHASES_COUNT={d['gold_purchases_count']}"
            ).upper(),
        )

    return JsonResponse({'result': True, 'detail': result})


@login_required(login_url='login')
@role_required('SUPPLIERS_CHANGE_STATUS')
def change_status(request):
    ids = request.POST.getlist('ids[]')
    qs = Suppliers.objects.filter(id__in=ids)
    for s in qs:
        s.is_active = not s.is_active
        s.save(update_fields=['is_active'])
    return JsonResponse({'result': True})


@login_required(login_url='login')
@role_required('SUPPLIERS_GET_ALL')
def get_all(request):
    draw = int(request.GET.get('draw', 1))
    length = int(request.GET.get('length', 10))
    start = int(request.GET.get('start', 0))
    search = request.GET.get('search[value]', '').strip()
    order_col_idx = request.GET.get('order[0][column]')
    order_col_name = request.GET.get(f'columns[{order_col_idx}][data]')

    valid_columns = ['company_name', 'phone', 'email', 'city__name', 'district__name']
    order_col = order_col_name if order_col_name in valid_columns else 'company_name'
    order_dir = '-' if request.GET.get('order[0][dir]') == 'desc' else ''

    store = request.user.store_id

    agg_pay = Sum('ledgers__amount_value',
                  filter=Q(ledgers__transaction_type='ENTRY',
                           ledgers__currency='HS',
                           ledgers__is_active=True),
                  output_field=DecimalField(max_digits=14, decimal_places=3))

    agg_rec = Sum('ledgers__amount_value',
                  filter=Q(ledgers__transaction_type='EXIT',
                           ledgers__currency='HS',
                           ledgers__is_active=True),
                  output_field=DecimalField(max_digits=14, decimal_places=3))

    last_act = Max('ledgers__created_on')

    qs = (Suppliers.objects
          .filter(is_deleted=False, store_id=store)
          .annotate(payable_hs=agg_pay, receivable_hs=agg_rec, last_activity=last_act))

    if search:
        qs = qs.filter(
            Q(company_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )

    is_active_filter = request.GET.get('is_active')
    if is_active_filter:
        qs = qs.filter(is_active=(is_active_filter == 'true' or is_active_filter == 'True'))

    total = qs.count()

    data = (qs.order_by(f'{order_dir}{order_col}')
            [start: None if length == -1 else start + length]
            .values('id', 'company_name', 'person_name', 'person_surname',
                    'email', 'phone', 'company_address', 'is_active',
                    'payable_hs', 'receivable_hs', 'last_activity',
                    'tax_payer_type', 'tax_number', 'tax_office', 'is_einvoice_user', 'post_code',
                    'city__id', 'city__name', 'district__id', 'district__name',
                    'account_type'))

    return JsonResponse({'draw': draw, 'recordsFiltered': total,
                         'recordsTotal': total, 'data': list(data)})


@login_required(login_url='login')
def get_suppliers(request):
    store = request.user.store
    agg_pay = Sum('ledgers__amount_value',
                  filter=Q(ledgers__transaction_type='ENTRY',
                           ledgers__currency='HS',
                           ledgers__is_active=True),
                  output_field=DecimalField(max_digits=14, decimal_places=3))
    agg_rec = Sum('ledgers__amount_value',
                  filter=Q(ledgers__transaction_type='EXIT',
                           ledgers__currency='HS',
                           ledgers__is_active=True),
                  output_field=DecimalField(max_digits=14, decimal_places=3))

    qs = (Suppliers.objects.filter(is_deleted=False, is_active=True, store=store)
          .annotate(payable_hs=agg_pay, receivable_hs=agg_rec)
          .values('id', 'company_name', 'receivable_hs', 'payable_hs'))
    return JsonResponse(list(qs), safe=False)


@login_required(login_url='login')
def get_balances(request):
    sid = request.GET.get('supplier_id')
    if not sid:
        return JsonResponse({'data': []})

    s = get_object_or_404(Suppliers, id=sid, is_active=True, is_deleted=False)
    zero = Value(Decimal('0'), output_field=DecimalField(max_digits=20, decimal_places=6))

    rows = (
        SupplierLedger.objects
        .filter(supplier_id=sid, is_active=True)
        .values('currency')
        .annotate(
            receivable=Coalesce(
                Sum('amount_value', filter=Q(transaction_type='EXIT'),
                    output_field=DecimalField(max_digits=20, decimal_places=6)),
                zero, output_field=DecimalField(max_digits=20, decimal_places=6)
            ),
            payable=Coalesce(
                Sum('amount_value', filter=Q(transaction_type='ENTRY'),
                    output_field=DecimalField(max_digits=20, decimal_places=6)),
                zero, output_field=DecimalField(max_digits=20, decimal_places=6)
            ),
        )
        .order_by('currency')
    )

    data = []
    for r in rows:
        net = r['receivable'] - r['payable']
        receivable = net if net >= 0 else Decimal('0')
        payable = -net if net < 0 else Decimal('0')

        data.append({
            'currency': r['currency'],
            'receivable': receivable,
            'payable': payable,
            'net': net
        })

    return JsonResponse({'data': data})


@login_required
def get_balances_all(request):
    sid = request.GET.get('supplier_id')
    if not sid:
        return JsonResponse({'data': []})
    s = get_object_or_404(Suppliers, id=sid, is_active=True, is_deleted=False)

    all_cur = ('HS', 'TRY', 'USD', 'EUR', 'CAD', 'QAR')
    summary = s.balance_summary()
    rows = [{'currency': c,
             'receivable': float(summary.get(c, {}).get('receivable', 0)),
             'payable': float(summary.get(c, {}).get('payable', 0)),
             'net': float(summary.get(c, {}).get('net', 0))}
            for c in all_cur]

    return JsonResponse({'data': rows})



@login_required(login_url='login')
def download_supplier_report(request, record_id):
    # 1. Tedarikçi
    s = get_object_or_404(Suppliers, id=record_id)

    # 2. Temel Sorgular
    process_qs = Process.objects.filter(supplier=s, is_deleted=False).exclude(is_status='CANCELED')
    barcode_qs = GoldPurchases.objects.filter(supplier=s, is_deleted=False, is_active=True)

    # 3. KPI ve Hacim Hesaplamaları
    DEC_FIELD = DecimalField(max_digits=18, decimal_places=3)
    ZERO = Value(Decimal("0.000"), output_field=DEC_FIELD)

    proc_aggr = process_qs.filter(transaction_type__in=['PURCHASE', 'STOCK_IN']).aggregate(
        total_gram=Coalesce(Sum('gram', output_field=DEC_FIELD), ZERO)
    )
    bar_aggr = barcode_qs.aggregate(
        total_gram=Coalesce(Sum('product__gram', output_field=DEC_FIELD), ZERO)
    )

    total_volume_gram = proc_aggr['total_gram'] + bar_aggr['total_gram']

    total_tx = process_qs.count()
    return_tx = process_qs.filter(transaction_type='RETURN').count()
    return_rate = (return_tx / total_tx * 100) if total_tx > 0 else 0
    total_interaction = total_tx + barcode_qs.count()

    # Son Hareket Tarihi
    last_proc = process_qs.order_by('-date').first()
    last_bar = barcode_qs.order_by('-created_on').first()
    last_date = None
    if last_proc and last_bar:
        last_date = last_proc.date if last_proc.date > last_bar.created_on else last_bar.created_on
    elif last_proc: last_date = last_proc.date
    elif last_bar: last_date = last_bar.created_on

    # 4. Ürün Analizi (Top 5)
    top_products = (process_qs
    .filter(transaction_type__in=['PURCHASE', 'STOCK_IN'], product__isnull=False)
    .values('product__name', 'product__category__name')
    .annotate(
        total_qty=Sum('piece'),
        total_gr=Sum('gram', output_field=DEC_FIELD)
    )
    .order_by('-total_gr')[:5]
    )

    # 5. Son Hareketler (Limit: 50)
    last_transactions = process_qs.select_related('product').order_by('-date')[:50]

    # FAZ 11 / BL-04: Mutabakat Düzeltmeleri (ADJ- prefix'li SupplierLedger kayıtları)
    # Adjustment kayıtları Process tablosunda DEĞİL, doğrudan SupplierLedger'da
    # tutulur. Bu yüzden ayrı bir sorgu ile çekilip PDF'te ayrı başlık altında
    # gösterilir.
    adjustment_entries = (
        SupplierLedger.objects
        .filter(supplier=s, process_no__startswith='ADJ-')
        .order_by('-created_on')[:50]
    )

    # 6. BAKİYE DETAYI (Tüm Dövizler)
    raw_balances = s.balance_summary()
    balance_list = []

    priority_currencies = ['HS', 'TRY', 'USD', 'EUR']

    for curr in priority_currencies:
        if curr in raw_balances:
            data = raw_balances.pop(curr)
            # 0.01'den büyük veya küçükse listeye ekle (Sıfır bakiyeleri gösterme)
            if abs(data['net']) > Decimal('0.01'):
                balance_list.append({'currency': curr, 'net': data['net']})

    for curr, data in raw_balances.items():
        if abs(data['net']) > Decimal('0.01'):
            balance_list.append({'currency': curr, 'net': data['net']})

    # Ana Bakiye (HS)
    bal_hs = s.balance_summary().get('HS', {'net': 0})['net']

    # 7. Context
    context = {
        'company_name': "Kuyum Plus",
        'report_date': timezone.localtime(timezone.now()).strftime("%d.%m.%Y %H:%M"),
        'report_no': f"RPT-{s.id.hex[:6].upper()}",
        'authorized': request.user.get_full_name() or request.user.username,
        'record': s,
        'balance_hs': bal_hs,
        'balance_list': balance_list,
        'kpi_volume': total_volume_gram,
        'kpi_tx_count': total_interaction,
        'kpi_return_rate': round(return_rate, 2),
        'kpi_last_date': last_date,
        'top_products': top_products,
        'transactions': last_transactions,
        # FAZ 11 / BL-04: Mutabakat Düzeltmeleri ayrı bölümde
        'adjustment_entries': adjustment_entries,
    }

    # 8. Render PDF
    html_string = render_to_string('management/suppliers/supplier_report_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    filename = f"TedarikciEkstre_{s.company_name}_{timezone.localtime(timezone.now()).strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html_string, dest=response, encoding='utf-8')

    if pisa_status.err:
        return HttpResponse('PDF hatası.', status=500)

    return response


# ============================================================================
# ANLIK TEDARİKÇİ İŞLEM FİŞİ
# ============================================================================

@login_required(login_url='login')
def supplier_fis_data(request):
    """
    Toptan ekranında seçili tedarikçi için anlık fiş verisi döndürür.
    Hem klasik tedarikçi hem Çantacı hesaplarını destekler.
    Tek AJAX çağrısıyla bakiye + son hareketleri getirir.
    """
    sid = request.GET.get('supplier_id')
    if not sid:
        return JsonResponse({'result': False, 'error_msg': 'Tedarikçi ID gerekli'}, status=400)

    s = get_object_or_404(Suppliers, id=sid, is_deleted=False)
    is_cantaci = s.account_type == 'CANTACI'

    # ── 1. BAKİYELER ──
    raw = s.balance_summary()
    priority = ['HS', 'TRY', 'USD', 'EUR']
    balances = []

    for cur in priority:
        if cur in raw:
            d = raw.pop(cur)
            if abs(d['net']) > Decimal('0.01'):
                balances.append({
                    'currency': cur,
                    'receivable': float(d['receivable']),
                    'payable': float(d['payable']),
                    'net': float(d['net']),
                })
    for cur, d in raw.items():
        if abs(d['net']) > Decimal('0.01'):
            balances.append({
                'currency': cur,
                'receivable': float(d['receivable']),
                'payable': float(d['payable']),
                'net': float(d['net']),
            })

    # ── 2. SON HAREKETLER (8 kayıt) ──
    movements = []

    if is_cantaci:
        # ────────────────────────────────────────────────────────────────
        # FAZ 21 / Bug 4 — Çantacı için TÜM aktif SupplierLedger satırları.
        # ────────────────────────────────────────────────────────────────
        # Eski davranış: yalnızca cantaci_tx_type__isnull=False satırlar
        # gösteriliyordu. Toptan ekranından çantacıya yapılan alış/satış
        # işlemleri book_supplier_product_tx ile yazılıyor ve cantaci_tx_type
        # alanı NULL kalıyor → fişte hiç görünmüyordu (bakiye doğru, hareket
        # boş). Artık tüm aktif ledger satırları (toptan + manuel çantacı
        # işlemleri birlikte) gösteriliyor; tip etiketi bağlama göre türetiliyor.
        # ────────────────────────────────────────────────────────────────
        ct_labels = {
            'HURDA_VERILDI': 'Hurda Verildi',
            'URUN_ALINDI': 'Ürün Alındı',
            'ODEME': 'Ödeme',
            'TAHSILAT': 'Tahsilat',
        }
        qs = (SupplierLedger.objects
              .filter(supplier=s, is_active=True)
              .select_related('product')
              .order_by('-created_on')[:8])
        for r in qs:
            # Tip etiketi: çantacı-spesifik → manuel etiket; yoksa toptan
            # alım/satış olarak tip + ürün adı.
            if r.cantaci_tx_type:
                tip = ct_labels.get(r.cantaci_tx_type, r.cantaci_tx_type or '-')
            else:
                _base = 'Toptan Alış' if r.transaction_type == 'ENTRY' else 'Toptan Satış'
                _prod = r.product.name if r.product else ''
                tip = f'{_base} — {_prod}' if _prod else _base
            # Miktar: gram öncelikli, yoksa adet, yoksa '-'
            if r.quantity_gram and r.quantity_gram > 0:
                miktar = f'{r.quantity_gram:.3f} Gr'
            elif r.quantity_piece and r.quantity_piece > 0:
                miktar = f'{r.quantity_piece} Ad'
            else:
                miktar = '-'
            movements.append({
                'tarih': timezone.localtime(r.created_on).strftime('%d.%m.%Y %H:%M'),
                'tip': tip,
                'miktar': miktar,
                'tutar': f'{r.amount_value:.3f} {r.currency}',
                'yon': r.transaction_type,
            })
    else:
        # Klasik tedarikçi: Process tablosu
        tx_labels = {
            'PURCHASE': 'Alış', 'STOCK_IN': 'Stok Girişi', 'ORDER_IN': 'Sipariş',
            'SALE': 'Satış', 'RETURN': 'İade', 'PAYMENT': 'Ödeme',
        }
        qs = (Process.objects
              .filter(supplier=s, is_deleted=False)
              .exclude(is_status='CANCELED')
              .select_related('product')
              .order_by('-date')[:8])
        for p in qs:
            prod_name = p.product.name if p.product else ''
            tip = tx_labels.get(p.transaction_type, p.transaction_type)
            if prod_name:
                tip = f'{tip} — {prod_name}'

            miktar = '-'
            if p.gram and p.gram > 0:
                miktar = f'{p.gram:.3f} Gr'
            elif p.piece and p.piece > 0:
                miktar = f'{p.piece} Ad'

            tutar_parts = []
            if p.price_hs and p.price_hs > 0:
                tutar_parts.append(f'{p.price_hs:.3f} HS')
            if p.amount and p.amount > 0:
                tutar_parts.append(f'{p.amount:.2f} TL')

            movements.append({
                'tarih': timezone.localtime(p.date).strftime('%d.%m.%Y %H:%M') if p.date else '-',
                'tip': tip,
                'miktar': miktar,
                'tutar': ' / '.join(tutar_parts) if tutar_parts else '-',
                'yon': 'ENTRY' if p.transaction_type in ('PURCHASE', 'STOCK_IN', 'ORDER_IN') else 'EXIT',
            })

    # ── 3. FİŞ META ──
    fis_no = f'FIS-{s.id.hex[:6].upper()}-{timezone.localtime(timezone.now()).strftime("%d%m%y")}'

    # ════════════════════════════════════════════════════════════════════════
    # FAZ 11 / SYNC-01 + SYNC-02 — TASLAK SEPET ÖZETİ (2026-04-24)
    # ════════════════════════════════════════════════════════════════════════
    # Toptan ekranında esnafın "şu an sepete ne koydum, onaylarsam carisi
    # nasıl olacak?" sorusuna cevap verir. Frontend opsiyonel olarak
    # wholesale_process_no ve/veya 'all_user_basket=1' gönderebilir.
    # ════════════════════════════════════════════════════════════════════════
    pending_basket = {
        'has_pending': False,
        'process_no': None,
        'rows_count': 0,
        'totals': [],  # [{'currency': 'HS', 'entry': 50.0, 'exit': 0.0, 'net': 50.0}, ...]
        'status_label': 'TAMAMLANDI',  # TASLAK | TAMAMLANDI
    }
    proc_no_param = (request.GET.get('wholesale_process_no') or '').strip()
    all_user_basket = request.GET.get('all_user_basket') == '1'

    pending_qs = Process.objects.filter(
        process_type='WHOLESALE',
        is_status='IN_PROGRESS',
        is_deleted=False,
        employee=request.user,
    ).select_related('product')

    if proc_no_param:
        pending_qs = pending_qs.filter(process_no=proc_no_param)
    elif not all_user_basket:
        # Hiçbir filtre verilmediyse boş — yanıltmamak adına
        pending_qs = pending_qs.none()

    if pending_qs.exists():
        pending_basket['has_pending'] = True
        pending_basket['process_no'] = proc_no_param or None
        pending_basket['rows_count'] = pending_qs.count()
        pending_basket['status_label'] = 'TASLAK'

        # Currency bazlı toplama
        totals_map = {}  # {currency: {'entry': 0, 'exit': 0}}

        for p in pending_qs:
            # Yön: complete_process_wholesale ile aynı mantık
            is_entry = p.transaction_type in ('PURCHASE', 'STOCK_IN', 'RETURN', 'ORDER_IN')

            # Kasa kalemi (ürün yok, banka var) → payment_currency
            if p.bank_account_id is not None:
                cur = (p.payment_currency or 'TRY').upper()
                amt = float(p.amount or 0)
            elif p.product:
                # Ürün kalemi: book_supplier_product_tx mantığı
                prod_cur = (p.product.currency or '').upper()
                if prod_cur == 'TRY':
                    cur = 'TRY'
                    amt = float(p.amount or 0)
                elif prod_cur in ('', 'HS'):
                    cur = 'HS'
                    amt = float(p.price_hs or 0)
                else:
                    cur = prod_cur
                    amt = float(p.gram or p.piece or 0)
            else:
                continue

            if amt <= 0:
                continue

            row = totals_map.setdefault(cur, {'entry': 0.0, 'exit': 0.0})
            if is_entry:
                row['entry'] += amt
            else:
                row['exit'] += amt

        pending_basket['totals'] = [
            {
                'currency': cur,
                'entry': round(t['entry'], 3),
                'exit': round(t['exit'], 3),
                'net': round(t['entry'] - t['exit'], 3),
            }
            for cur, t in totals_map.items()
        ]

    return JsonResponse({
        'result': True,
        'supplier': {
            'id': str(s.id),
            'company_name': s.company_name,
            'person_name': s.person_name or '',
            'person_surname': s.person_surname or '',
            'phone': s.phone or '-',
            'account_type': s.account_type or 'SUPPLIER',
        },
        'balances': balances,
        'movements': movements,
        'is_cantaci': is_cantaci,
        'fis_no': fis_no,
        'fis_tarih': timezone.localtime(timezone.now()).strftime('%d.%m.%Y %H:%M'),
        # FAZ 11 / SYNC-01 + SYNC-02
        'pending_basket': pending_basket,
    })


# ============================================================================
# ÇANTACI MODÜLÜ VIEW'LARI
# ============================================================================

def _cantaci_kpi(supplier):
    """Çantacı KPI verilerini hesaplar (view + AJAX response için ortak).

    FAZ 31 / BUG-2: net_bakiye_tl alanı eklendi — ENTRY/EXIT yönüne göre
    aktif TL satırlarının net bakiyesi. Mevcut hurda/ürün HAS hesabı
    DEĞİŞMEDİ; ek alan olarak net_bakiye_tl döner.
    """
    ledgers = SupplierLedger.objects.filter(supplier=supplier, is_active=True)

    hurda_agg = ledgers.filter(cantaci_tx_type='HURDA_VERILDI').aggregate(
        gram=Coalesce(Sum('quantity_gram'), Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=3))),
        has=Coalesce(Sum('amount_value'), Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=3))),
    )
    urun_agg = ledgers.filter(cantaci_tx_type='URUN_ALINDI').aggregate(
        gram=Coalesce(Sum('quantity_gram'), Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=3))),
        has=Coalesce(Sum('amount_value'), Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=3))),
    )

    # Ödeme/Tahsilat (TL cinsinden)
    odeme_tl = ledgers.filter(
        cantaci_tx_type__in=['ODEME', 'TAHSILAT'], currency='TRY'
    ).aggregate(
        toplam=Coalesce(Sum('amount_value'), Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=3)))
    )['toplam']

    # Net Bakiye HAS: Hurda Verildi (EXIT/Alacak) - Ürün Alındı (ENTRY/Borç)
    net_bakiye_has = hurda_agg['has'] - urun_agg['has']

    toplam_islem = ledgers.filter(cantaci_tx_type__isnull=False).count()

    # ────────────────────────────────────────────────────────────────
    # FAZ 31 / BUG-2: TL bakiye — ENTRY/EXIT yönüne göre net hesap
    # ────────────────────────────────────────────────────────────────
    # ENTRY = biz çantacıya borçluyuz (URUN_ALINDI + TAHSILAT yönü)
    # EXIT  = çantacı bize borçlu (HURDA_VERILDI + ODEME yönü)
    #
    # Ödeme yönü düzeltmesinden sonra:
    #   ODEME ya da TAHSILAT alındığında book_supplier_tx FIFO ile karşı
    #   yöndeki satırları kapatıyor (is_active=False). Aktif satırların
    #   ENTRY toplamı - EXIT toplamı = mevcut TL borç bakiyesi.
    # ────────────────────────────────────────────────────────────────
    tl_entry = ledgers.filter(
        currency='TRY', transaction_type='ENTRY',
    ).aggregate(
        t=Coalesce(Sum('amount_value'), Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=2))),
    )['t']
    tl_exit = ledgers.filter(
        currency='TRY', transaction_type='EXIT',
    ).aggregate(
        t=Coalesce(Sum('amount_value'), Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=2))),
    )['t']
    # Pozitif net_bakiye_tl → biz çantacıya borçluyuz
    # Negatif → çantacı bize borçlu
    net_bakiye_tl = (tl_entry or Decimal('0')) - (tl_exit or Decimal('0'))

    return {
        'toplam_hurda_gram': hurda_agg['gram'],
        'toplam_hurda_has': hurda_agg['has'],
        'toplam_urun_gram': urun_agg['gram'],
        'toplam_urun_has': urun_agg['has'],
        'net_bakiye_has': net_bakiye_has,
        'toplam_odeme_tl': odeme_tl,
        'toplam_islem': toplam_islem,
        # FAZ 31 / BUG-2: ek alan (mevcut KPI'lar bozulmadı)
        'net_bakiye_tl': net_bakiye_tl,
    }


@login_required(login_url='login')
def cantaci_detail_view(request, record_id):
    """Çantacı hesabı detay sayfası."""
    s = get_object_or_404(Suppliers, id=record_id, account_type='CANTACI')

    kpi = _cantaci_kpi(s)

    ctx = {
        'title': f'{s.company_name} - Çantacı Detay',
        'record': s,
        **kpi,
    }
    return render(request, 'management/suppliers/cantaci_detail.html', ctx)


@login_required(login_url='login')
def cantaci_hareketler(request):
    """Çantacı hareket tablosu için DataTables AJAX endpoint."""
    draw = int(request.GET.get('draw', 1))
    length = int(request.GET.get('length', 10))
    start = int(request.GET.get('start', 0))

    supplier_id = request.GET.get('supplier_id')
    cantaci_tx_type = request.GET.get('cantaci_tx_type', '').strip()

    # ────────────────────────────────────────────────────────────────────
    # FAZ 21 / Bug 4 — Çantacı hareket listesi: tüm aktif SupplierLedger
    # satırları (toptan + manuel çantacı işlemleri). Filtre verilirse
    # cantaci_tx_type ile daraltılır (eski davranış korunur).
    # Toptan satırları için cantaci_tx_type alanı NULL → frontend tarafında
    # 'TOPTAN' pseudo-etiketi gönderilir; description bağlamı ürün adıyla
    # zenginleştirilir.
    # ────────────────────────────────────────────────────────────────────
    qs = SupplierLedger.objects.filter(
        supplier_id=supplier_id,
        is_active=True,
    ).select_related('product').order_by('-created_on')

    if cantaci_tx_type:
        if cantaci_tx_type == 'TOPTAN':
            qs = qs.filter(cantaci_tx_type__isnull=True)
        else:
            qs = qs.filter(cantaci_tx_type=cantaci_tx_type)

    total = qs.count()
    rows = qs[start:start + length]

    data = []
    for r in rows:
        # HAS ve TL ayrımı
        amount_hs = float(r.amount_value) if r.currency == 'HS' else 0
        amount_eur = float(r.amount_value) if r.currency == 'TRY' else 0

        # Tip sınıflaması (frontend için):
        # - cantaci_tx_type doluysa onu kullan
        # - Boşsa toptan alış/satış olarak göster ('TOPTAN_ENTRY' / 'TOPTAN_EXIT')
        if r.cantaci_tx_type:
            _tx = r.cantaci_tx_type
        else:
            _tx = 'TOPTAN_ENTRY' if r.transaction_type == 'ENTRY' else 'TOPTAN_EXIT'

        # Açıklama: orijinal description varsa onu kullan; yoksa
        # ürün adıyla bağlam ekle.
        _desc = r.description or ''
        if not _desc and r.product is not None:
            _desc = f'Toptan: {r.product.name}'

        data.append({
            'created_on': timezone.localtime(r.created_on).strftime('%Y-%m-%d %H:%M'),
            'cantaci_tx_type': _tx,
            'description': _desc,
            'quantity_gram': float(r.quantity_gram),
            'amount_hs': amount_hs,
            'amount_eur': amount_eur,
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total,
        'recordsFiltered': total,
        'data': data,
    })


# ════════════════════════════════════════════════════════════════════
# FAZ 32 — `cantaci_islem_ekle` view fonksiyonu kaldırıldı.
# ════════════════════════════════════════════════════════════════════
# Önceki yapı: çantacı detay sayfasındaki "Yeni İşlem" modalı
# `SupplierLedger.objects.create(...)` ile DOĞRUDAN kayıt yazıyor;
# `book_supplier_tx`'i, FIFO auto_setoff'u ve StockLedger akışını
# tamamen baypas ediyordu. Saha geri bildirimi: işlemler ne ürün
# stoğuna ne de hareket listesine yansımıyordu.
#
# Karar: Çantacı işlemleri tek bir akıştan (Toptan / Wholesale modülü)
# yürütülecek. İzole endpoint kaldırıldı; `cantaci_hareketler` view'ı
# zaten toptan kaynaklı `TOPTAN_ENTRY` / `TOPTAN_EXIT` pseudo-tiplerini
# gösterdiği için listede veri kaybı yaşanmaz.
# ════════════════════════════════════════════════════════════════════


@login_required(login_url='login')
def cantaci_export(request, record_id):
    """Çantacı hesabı PDF/Excel export."""
    s = get_object_or_404(Suppliers, id=record_id, account_type='CANTACI')
    export_format = request.GET.get('format', 'pdf')

    kpi = _cantaci_kpi(s)

    # Tüm hareketler
    hareketler = SupplierLedger.objects.filter(
        supplier=s, cantaci_tx_type__isnull=False
    ).order_by('-created_on')

    if export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return HttpResponse('openpyxl kütüphanesi bulunamadı.', status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Çantacı Ekstre"

        # Başlık
        header_font = Font(bold=True, size=14)
        ws.append([f"Çantacı Ekstre - {s.company_name}"])
        ws['A1'].font = header_font
        ws.append([f"Tarih: {timezone.localtime(timezone.now()).strftime('%d.%m.%Y %H:%M')}"])
        ws.append([])

        # KPI Özet
        ws.append(["Toplam Hurda Verildi (HAS)", float(kpi['toplam_hurda_has'])])
        ws.append(["Toplam Ürün Alındı (HAS)", float(kpi['toplam_urun_has'])])
        ws.append(["Net Bakiye (HAS)", float(kpi['net_bakiye_has'])])
        ws.append([])

        # Tablo Başlıkları
        headers = ["Tarih", "İşlem Tipi", "Açıklama", "Gram", "HAS Değeri", "TL Tutar"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", fill_type="solid")

        # Veriler
        tx_labels = {
            'HURDA_VERILDI': 'Hurda Verildi',
            'URUN_ALINDI': 'Ürün Alındı',
            'ODEME': 'Ödeme',
            'TAHSILAT': 'Tahsilat',
        }
        for h in hareketler:
            ws.append([
                timezone.localtime(h.created_on).strftime('%d.%m.%Y %H:%M'),
                tx_labels.get(h.cantaci_tx_type, h.cantaci_tx_type or '-'),
                h.description or '-',
                float(h.quantity_gram),
                float(h.amount_value) if h.currency == 'HS' else 0,
                float(h.amount_value) if h.currency == 'TRY' else 0,
            ])

        # Sütun genişlikleri
        for col_idx, width in enumerate([20, 16, 30, 12, 14, 14], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"CantaciEkstre_{s.company_name}_{timezone.localtime(timezone.now()).strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    else:
        # PDF Export
        tx_labels = {
            'HURDA_VERILDI': 'Hurda Verildi',
            'URUN_ALINDI': 'Ürün Alındı',
            'ODEME': 'Ödeme',
            'TAHSILAT': 'Tahsilat',
        }
        context = {
            'record': s,
            'report_date': timezone.localtime(timezone.now()).strftime("%d.%m.%Y %H:%M"),
            'authorized': request.user.get_full_name() or request.user.username,
            'hareketler': hareketler,
            'tx_labels': tx_labels,
            **kpi,
        }
        html_string = render_to_string('management/suppliers/cantaci_report_pdf.html', context)
        response = HttpResponse(content_type='application/pdf')
        filename = f"CantaciEkstre_{s.company_name}_{timezone.localtime(timezone.now()).strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        pisa_status = pisa.CreatePDF(html_string, dest=response, encoding='utf-8')
        if pisa_status.err:
            return HttpResponse('PDF hatası.', status=500)
        return response


# ════════════════════════════════════════════════════════════════════════════
# FAZ 11 / BL-01 — TEDARİKÇİ CARİ SIFIRLAMA / MANUEL DÜZELTME (2026-04-24)
# ════════════════════════════════════════════════════════════════════════════
# Kuyumcu esnafının "Şu cariyi sıfırla" / "Şu tutara eşitle" ihtiyacını,
# Immutable Ledger mimarisine uygun olarak "Adjustment Entry" (düzeltme fişi)
# yazarak karşılar. Mevcut SupplierLedger kayıtlarına UPDATE/DELETE YAPILMAZ;
# bunun yerine ters yönde yeni kayıt açılır ve FIFO auto-setoff ile karşı
# taraftaki açık satırlar kapatılır.
#
# Modlar:
#   ALL         — Tüm para birimlerini sıfırla (her currency için ayrı ADJ kaydı)
#   BY_CURRENCY — Belirli bir para birimini sıfırla
#   CUSTOM      — Belirli bir para biriminde net bakiyeyi target_amount'a eşitle
#
# Yetki: Bu view BL-03'te '@role_required(SUPPLIERS_ADJUSTMENT_CREATE)' ile
# sarılacaktır. Şu aşamada inline superuser/role_id kontrolü yapılır.
# ════════════════════════════════════════════════════════════════════════════

def _parse_adjustment_amount(raw):
    """
    TR locale destekli Decimal parser (yerel helper).
    Örn: '1.500,50' -> Decimal('1500.50'), '1500.50' -> Decimal('1500.50').
    Geçersiz veri için None döner.
    """
    from decimal import InvalidOperation
    if raw is None:
        return None
    txt = str(raw).strip()
    if not txt:
        return None
    if '.' in txt and ',' in txt:
        txt = txt.replace('.', '').replace(',', '.')
    elif ',' in txt:
        txt = txt.replace(',', '.')
    try:
        return Decimal(txt)
    except (InvalidOperation, ValueError):
        return None


def _write_adjustment_ledger(*, supplier, currency, net_balance, target_balance,
                             process_no, user_description, user):
    """
    Tek bir para birimi için Adjustment Entry yazar.

    Algoritma:
        diff = net_balance - target_balance
        diff > 0  → Biz alacaklıyız; ENTRY yazarak alacağı sıfırla
                    (auto_setoff karşı yöndeki EXIT satırlarını kapatır)
        diff < 0  → Biz borçluyuz; EXIT yazarak borcu sıfırla
        diff == 0 → İşlem gerekmez, None döner

    Args:
        supplier: Suppliers instance
        currency: 'HS' | 'HG' | 'TRY' | 'USD' | ...
        net_balance: mevcut net (receivable - payable)
        target_balance: hedef net bakiye (sıfırlama için 0)
        process_no: Ortak ADJ- process_no
        user_description: Kullanıcının girdiği açıklama (zorunlu)
        user: request.user (log için)

    Returns:
        dict: {'currency', 'tx_type', 'amount', 'ledger_id'} veya None.
    """
    # Döngüsel import'u engellemek için lazy import.
    from apps.process.wholesale_views import book_supplier_tx

    diff = net_balance - target_balance
    if diff == 0:
        return None

    if diff > 0:
        tx_type = 'ENTRY'  # Alacağımızı sıfırlamak için kendi hesabımıza borç
        amount = diff
    else:
        tx_type = 'EXIT'   # Borcumuzu sıfırlamak için kendi hesabımıza alacak
        amount = abs(diff)

    # ── FAZ 11 / BL-01-HARDEN-02: Defensive Positive-Amount Guard ──
    # amount mantıksal olarak her zaman > 0 olmalıdır (diff != 0 yukarıda
    # kontrol edildi + abs() ile pozitiflik garanti). Yine de SupplierLedger
    # tablosuna negatif veya sıfır amount_value yazılmasını mutlak olarak
    # engellemek için son savunma katmanı. Aşağıdaki senaryolardan herhangi
    # birinde tespit edilemeyen bir bug oluşursa burada görülür:
    #   - net_balance veya target_balance yanlış tipte (str -> Decimal kaybı)
    #   - Decimal precision overflow
    #   - book_supplier_tx tarafındaki future-refactor regresyon riski
    # ValueError fırlatmak @transaction.atomic dekoratörünü tetikler ve
    # transaction otomatik rollback edilir → veri bütünlüğü korunur.
    if amount is None or amount <= 0:
        raise ValueError(
            f'[ADJUSTMENT-GUARD] Geçersiz amount={amount} '
            f'(supplier={getattr(supplier, "id", None)}, currency={currency}, '
            f'net_balance={net_balance}, target_balance={target_balance}). '
            f'SupplierLedger\'a sıfır/negatif tutar yazımı engellendi; '
            f'transaction rollback edildi.'
        )

    # Açıklama: Kullanıcı notu + sistem etiketi (raporlarda parse edilebilir)
    ledger_description = f'[ADJUSTMENT] {user_description}'

    ledger = book_supplier_tx(
        supplier=supplier,
        transaction_type=tx_type,
        amount_value=amount,
        currency=currency,
        process_no=process_no,
        description=ledger_description,
        auto_setoff=True,
    )

    if ledger is None:
        # book_supplier_tx guard'ı (amount <= 0) — teorik olarak buraya düşmez
        # çünkü diff != 0 olduğunu yukarıda kontrol ettik.
        return None

    return {
        'currency': currency,
        'tx_type': tx_type,
        'amount': str(amount),
        'previous_net': str(net_balance),
        'target_net': str(target_balance),
        'ledger_id': str(ledger.id),
    }


@login_required(login_url='login')
@role_required('SUPPLIERS_ADJUSTMENT_CREATE')
@transaction.atomic
def supplier_adjustment_create(request):
    """
    FAZ 11 / BL-01: Tedarikçi cari sıfırlama / manuel düzeltme endpoint'i.

    POST Parametreleri:
        supplier_id    (zorunlu) — Suppliers UUID
        mode           (zorunlu) — 'ALL' | 'BY_CURRENCY' | 'CUSTOM'
        currency       (mode != 'ALL' ise zorunlu) — 'HS', 'HG', 'TRY', 'USD' ...
        target_amount  (mode == 'CUSTOM' ise zorunlu) — Hedef net bakiye
                       (TR locale: '1.500,50' veya EN: '1500.50')
        description    (zorunlu) — Denetim notu (min 3 karakter)

    Response (başarılı):
        {
            'result': True,
            'process_no': 'ADJ-20260424-143512-a1b2c3d4-9f8e',
            'entries': [{'currency': 'HS', 'tx_type': 'ENTRY', 'amount': '50.000', ...}, ...],
            'message': 'Cari düzeltme başarıyla uygulandı.'
        }

    Response (hata): {'error': True, 'error_msg': '...'} + uygun HTTP status

    Güvenlik:
        - @login_required — Oturum zorunlu
        - @role_required('SUPPLIERS_ADJUSTMENT_CREATE') — FAZ 11 / BL-03
          (3 katmanlı SaaS Hybrid-Gate: Konasoft personeli RoleDetail,
          mağaza personeli mağaza efektif yetki havuzu, superuser bypass)
        - @transaction.atomic — Bir para birimi yazılıp diğeri başarısız
          olursa tamamı rollback edilir
        - description alanı zorunlu — her düzeltme denetlenebilir

    Permission Seed (Manuel Adım — Geliştirici Sorumluluğu):
        Bu yetki kodunun (`SUPPLIERS_ADJUSTMENT_CREATE`) Permission tablosuna
        ve uygun rolün/mağaza paketinin yetki havuzuna eklenmesi gerekir.
        Aksi takdirde superuser dışında hiçbir kullanıcı erişemez.

    Immutable Ledger Garantisi:
        - Mevcut SupplierLedger satırlarına UPDATE atılır fakat bu yalnızca
          book_supplier_tx içindeki auto_setoff mekanizmasıdır (amount_value
          azaltma + is_active=False). Bu yıllardır kullanılan standart akış
          olup Immutable Ledger prensibini ihlal etmez — yeni ADJ kaydı
          ters yönde eklenir ve FIFO ile kapatma yapılır.
        - Hiçbir satır silinmez. Audit trail korunur.
    """
    if request.method != 'POST':
        return JsonResponse(
            {'error': True, 'error_msg': 'Geçersiz istek.'}, status=405
        )

    u = request.user

    # ── Parametre parse ──
    supplier_id = (request.POST.get('supplier_id') or '').strip()
    mode = (request.POST.get('mode') or '').strip().upper()
    currency = (request.POST.get('currency') or '').strip().upper() or None
    target_amount_raw = (request.POST.get('target_amount') or '').strip()
    description = (request.POST.get('description') or '').strip()

    # ── Validasyon ──
    if not supplier_id:
        return JsonResponse(
            {'error': True, 'error_msg': 'Tedarikçi seçilmedi.'}, status=400
        )
    if mode not in ('ALL', 'BY_CURRENCY', 'CUSTOM'):
        return JsonResponse(
            {'error': True, 'error_msg': 'Geçersiz mod. ALL / BY_CURRENCY / CUSTOM bekleniyor.'},
            status=400,
        )
    if len(description) < 3:
        return JsonResponse(
            {'error': True, 'error_msg': 'Açıklama zorunludur (en az 3 karakter). Denetim için gereklidir.'},
            status=400,
        )
    if mode in ('BY_CURRENCY', 'CUSTOM') and not currency:
        return JsonResponse(
            {'error': True, 'error_msg': 'Para birimi seçilmelidir.'}, status=400
        )

    # CUSTOM modunda target_amount parse
    target_amount = None
    if mode == 'CUSTOM':
        target_amount = _parse_adjustment_amount(target_amount_raw)
        if target_amount is None:
            return JsonResponse(
                {'error': True, 'error_msg': 'Hedef tutar geçersiz formatta.'},
                status=400,
            )
        if target_amount < 0:
            return JsonResponse(
                {'error': True, 'error_msg': 'Hedef tutar negatif olamaz.'},
                status=400,
            )

    # ── Supplier ve bakiye (FAZ 11 / BL-01-HARDEN-01: Race Condition Koruması) ──
    # Aynı tedarikçiye ait eşzamanlı yazma operasyonlarının ADJ matematiğini
    # bozmaması için Supplier satırını transaction süresince kilitliyoruz.
    # @transaction.atomic dekoratörü tek başına SELECT'leri serialize etmez;
    # PostgreSQL READ COMMITTED izolasyonunda T1 SELECT (balance_summary) ile
    # T2 INSERT (eşzamanlı toptan/perakende işlem) arası pencere açıktır.
    # select_for_update() Supplier satırını kilitleyerek bu pencereyi kapatır;
    # diğer transaction'lar bu commit'e kadar bekler. Böylece balance_summary()
    # ile ADJ yazımı arasındaki "stale read" sorunu ortadan kalkar.
    try:
        supplier = (Suppliers.objects
                    .select_for_update()
                    .get(pk=supplier_id))
    except Suppliers.DoesNotExist:
        return JsonResponse(
            {'error': True, 'error_msg': 'Tedarikçi bulunamadı.'}, status=404
        )
    # Kilit alındıktan sonra bakiye okunur — bu noktadan transaction commit'ine
    # kadar başka bir thread bu supplier'a SupplierLedger satırı ekleyemez
    # (book_supplier_tx çağrıları aynı kilidi beklemek zorunda kalır).
    balance = supplier.balance_summary()  # {currency: {'receivable', 'payable', 'net'}, ...}

    # ── FAZ 11 / BL-01-HARDEN-03: Optimistic Balance Snapshot Doğrulaması ──
    # Kullanıcı modalı açtığı anda gördüğü bakiyeyi frontend POST ile geri
    # gönderir (`expected_balance` JSON parametresi). Lock alındıktan sonra
    # taze okunan bakiye ile karşılaştırırız. Aralarında fark varsa, modal
    # açıldıktan sonra başka bir kullanıcı işlem yapmış demektir → 409
    # Conflict ile geri çevrilir, kullanıcıdan sayfayı yenilemesi istenir.
    #
    # GERİYE DÖNÜK UYUM:
    #   Bu parametre OPSİYONELDIR. Frontend henüz göndermiyorsa (eski client),
    #   doğrulama atlanır ve eski davranış korunur. Frontend güncellendikten
    #   sonra parametre gönderildiğinde otomatik aktif olur.
    #
    # Beklenen format (POST):
    #   expected_balance = '{"HS": "100.000", "TRY": "5000.00"}'
    #   Anahtar: para birimi kodu (büyük harf), Değer: net bakiye (string).
    expected_balance_raw = (request.POST.get('expected_balance') or '').strip()
    if expected_balance_raw:
        import json
        from decimal import InvalidOperation
        try:
            expected_balance = json.loads(expected_balance_raw)
            if not isinstance(expected_balance, dict):
                raise ValueError('expected_balance bir dict olmalı.')
        except (json.JSONDecodeError, ValueError) as exc:
            return JsonResponse(
                {'error': True, 'error_msg': f'expected_balance JSON parse hatası: {exc}'},
                status=400,
            )

        # Her gönderilen para birimi için fresh balance ile karşılaştır.
        # Eksik para birimleri (frontend göndermemişse) doğrulanmaz; bu kasıtlı
        # — frontend yalnızca kullanıcının düzeltme yapmaya çalıştığı para
        # birim(ler)i için snapshot gönderebilir.
        mismatches = []
        for cur_key, expected_net_raw in expected_balance.items():
            cur_norm = str(cur_key).upper()
            try:
                expected_net = Decimal(str(expected_net_raw))
            except (InvalidOperation, ValueError):
                return JsonResponse(
                    {
                        'error': True,
                        'error_msg': (
                            f'expected_balance[{cur_norm}] geçersiz Decimal: '
                            f'{expected_net_raw!r}'
                        ),
                    },
                    status=400,
                )
            actual_net = Decimal(str(
                balance.get(cur_norm, {}).get('net', Decimal('0'))
            ))
            if expected_net != actual_net:
                mismatches.append({
                    'currency': cur_norm,
                    'expected': str(expected_net),
                    'actual': str(actual_net),
                })

        if mismatches:
            return JsonResponse(
                {
                    'error': True,
                    'error_code': 'BALANCE_CHANGED',
                    'error_msg': (
                        'Bakiye, modal açıldıktan sonra değişti. Başka bir '
                        'kullanıcı bu tedarikçiye işlem yapmış olabilir. '
                        'Lütfen sayfayı yenileyip tekrar deneyin.'
                    ),
                    'mismatches': mismatches,
                },
                status=409,  # HTTP 409 Conflict — optimistic locking standardı
            )

    # ── ADJ- prefix'li ortak process_no üret (BL-02 servisi) ──
    from apps.suppliers.services import SupplierLedgerService

    # ════════════════════════════════════════════════════════════════════
    # FAZ 32 — DEFENSIVE EXCEPTION WRAP (2026-05-01)
    # ════════════════════════════════════════════════════════════════════
    # Saha geri bildirimi: "Düzeltmeyi Uygula" butonu 500 Internal Server
    # Error döndürüyor. Kök neden analizi:
    #   1) generate_adjustment_process_no()  → çakışma retry tükenince
    #      RuntimeError fırlatır (services.py:330 civarı).
    #   2) _write_adjustment_ledger()        → defensive amount-guard
    #      (views.py:1546–1553) ValueError fırlatır (Decimal precision
    #      veya beklenmedik tip durumlarında).
    # Her iki exception da @transaction.atomic dekoratörü tarafından
    # yakalanmıyor; Django default 500 sayfası dönerek frontend'in JSON
    # parse'ı kırılıyor → kullanıcıya gerçek hata sebebi yansımıyor.
    # Çözüm: kontrollü exception'ları yakala, transaction'ı rollback et
    # (raise yok → atomic bloğu commit etmez çünkü exception yakalanmadan
    # önce çıkıyor ama biz kontrolü ele aldıktan sonra raise yeniden;
    # JSON 500 gönderiyoruz). Bilinmeyen Exception'lar bilinçli olarak
    # propagate edilir (Django default 500 + sentry/tracing gözükür).
    # ════════════════════════════════════════════════════════════════════
    try:
        process_no = SupplierLedgerService.generate_adjustment_process_no(supplier.id)

        entries = []

        # ── Moda göre Adjustment Entry yazımı ──
        if mode == 'ALL':
            # Tüm para birimlerinin netini 0'a çek.
            if not balance:
                return JsonResponse(
                    {'error': True, 'error_msg': 'Bu tedarikçi için düzeltilecek bakiye bulunmuyor.'},
                    status=400,
                )
            for cur, amounts in balance.items():
                net = amounts.get('net', Decimal('0'))
                result = _write_adjustment_ledger(
                    supplier=supplier,
                    currency=cur,
                    net_balance=Decimal(str(net)),
                    target_balance=Decimal('0'),
                    process_no=process_no,
                    user_description=description,
                    user=u,
                )
                if result:
                    entries.append(result)

            if not entries:
                return JsonResponse(
                    {'error': True, 'error_msg': 'Tüm para birimlerinin net bakiyesi zaten sıfır.'},
                    status=400,
                )

        elif mode == 'BY_CURRENCY':
            cur_data = balance.get(currency)
            if not cur_data:
                return JsonResponse(
                    {'error': True, 'error_msg': f'{currency} para biriminde işlem bulunmuyor.'},
                    status=400,
                )
            net = Decimal(str(cur_data.get('net', Decimal('0'))))
            if net == 0:
                return JsonResponse(
                    {'error': True, 'error_msg': f'{currency} bakiyesi zaten sıfır.'},
                    status=400,
                )
            result = _write_adjustment_ledger(
                supplier=supplier,
                currency=currency,
                net_balance=net,
                target_balance=Decimal('0'),
                process_no=process_no,
                user_description=description,
                user=u,
            )
            if result:
                entries.append(result)

        else:  # CUSTOM
            cur_data = balance.get(currency, {})
            net = Decimal(str(cur_data.get('net', Decimal('0'))))
            if net == target_amount:
                return JsonResponse(
                    {
                        'error': True,
                        'error_msg': f'{currency} bakiyesi zaten {target_amount}. Düzeltme gereksiz.',
                    },
                    status=400,
                )
            result = _write_adjustment_ledger(
                supplier=supplier,
                currency=currency,
                net_balance=net,
                target_balance=target_amount,
                process_no=process_no,
                user_description=description,
                user=u,
            )
            if result:
                entries.append(result)

    except RuntimeError as exc:
        # generate_adjustment_process_no çakışma retry tükendi.
        try:
            write_log(
                request,
                'SUPPLIER_ADJUSTMENT_FAIL',
                f'process_no üretimi başarısız — supplier_id={supplier.id}, mode={mode}, exc={exc}',
            )
        except Exception:
            pass
        # @transaction.atomic — exception bloğu içinde JsonResponse döndürmek
        # transaction'ı rollback ETMEZ; bu yüzden açıkça raise edip otomatik
        # rollback'i tetikliyoruz, ardından dış katmanda yakalayıp temiz
        # JSON dönüyoruz. Ancak Django'da view bir kez return ettiğinde
        # transaction commit edilir — bu yüzden burada doğrudan JSON döner
        # ve aynı transaction'da hiç INSERT olmadığı için rollback gereksiz
        # (process_no üretimi sırf cache.add+incr ile yapılır, DB yazımı yok).
        return JsonResponse(
            {
                'error': True,
                'error_code': 'ADJ_PROCESS_NO_COLLISION',
                'error_msg': (
                    'Düzeltme fişi numarası üretilemedi (eşzamanlı işlem '
                    'çakışması). Lütfen birkaç saniye bekleyip tekrar deneyin.'
                ),
            },
            status=500,
        )
    except ValueError as exc:
        # _write_adjustment_ledger amount-guard'ı (Decimal precision /
        # beklenmedik tip kaybı). Transaction'ı rollback edebilmek için
        # raise edip dış @transaction.atomic'in algılamasını sağlamak
        # gerekirdi; ancak burada JsonResponse döndürdüğümüz için Django
        # transaction'ı commit eder. Bu noktaya gelindiğinde hiçbir
        # SupplierLedger satırı yazılmamış olduğunu ispatlamak için
        # transaction.set_rollback(True) çağrısı yapıyoruz — atomic
        # blok sona erdiğinde rollback garanti edilir.
        try:
            transaction.set_rollback(True)
        except Exception:
            pass
        try:
            write_log(
                request,
                'SUPPLIER_ADJUSTMENT_FAIL',
                f'amount-guard tetiklendi — supplier_id={supplier.id}, mode={mode}, exc={exc}',
            )
        except Exception:
            pass
        return JsonResponse(
            {
                'error': True,
                'error_code': 'ADJ_AMOUNT_GUARD',
                'error_msg': (
                    'Düzeltme tutarı hesaplanamadı (sıfır veya negatif). '
                    'Bakiyenin değişmiş olabileceğini kontrol edip sayfayı '
                    'yenileyin. Sorun devam ederse destek ile iletişime geçin.'
                ),
            },
            status=500,
        )

    # ── Aktivite logu ──
    # write_log imzası: (request, title, description)
    try:
        write_log(
            request,
            'SUPPLIER_ADJUSTMENT',
            (
                f'Cari düzeltme — supplier_id={supplier.id}, mode={mode}, '
                f'process_no={process_no}, entries={len(entries)}, '
                f'note={description[:100]}'
            ),
        )
    except Exception:
        # Log başarısızlığı iş akışını durdurmaz
        pass

    return JsonResponse({
        'result': True,
        'process_no': process_no,
        'mode': mode,
        'entries': entries,
        'entries_count': len(entries),
        'message': 'Cari düzeltme başarıyla uygulandı.',
    })


# ════════════════════════════════════════════════════════════════════════════
# FAZ 52 — TEDARİKÇİ AÇILIŞ CARİSİ (OPENING BALANCE) (2026-05-05)
# ════════════════════════════════════════════════════════════════════════════
# Mağaza programa yeni geçtiğinde mevcut tedarikçilerle olan tarihsel
# borç/alacak kalemlerini stok hareketsiz olarak ledger'a girmek için
# kullanılır. process_no 'OB-' prefix'i ile yazılır:
#   - book_supplier_tx içindeki FIFO setoff filtresi tarafından muaf tutulur.
#   - Sayfada "Açılış Bakiyesi" badge'i ile ayrı gösterilir.
#   - Idempotent: bir tedarikçi için sadece BİR KEZ girilebilir; sonraki
#     denemelerde 409 Conflict döner ("Cari Düzelt" akışına yönlendirme).
#
# Mimari Garantiler:
#   1) Mevcut SupplierLedger şeması KORUNUR — yeni alan/transaction_type yok.
#   2) book_supplier_tx 'OB-' prefix exclusion'u FAZ 11'den beri rezerve
#      edilmişti; FAZ 52 bu altyapıyı aktive eder.
#   3) ADJ akışı (cari düzelt) zaten 'OB-' satırlarını kapatamaz; iki
#      mekanizma ortogonal kalır.
# ════════════════════════════════════════════════════════════════════════════

def _parse_opening_balance_entries(raw_json):
    """
    Frontend'den gelen entries JSON'unu parse eder ve normalize eder.

    Kabul edilen format (POST `entries` parametresi):
        '[{"currency": "HS", "direction": "RECEIVABLE", "amount": "50.000",
           "exchange_rate_eur": null},
          {"currency": "TRY", "direction": "PAYABLE", "amount": "1500.50",
           "exchange_rate_eur": "1"}]'

    Returns:
        list[dict] — her dict: {currency, direction, amount, exchange_rate_eur}.
                     Tutarlar 0 olan satırlar UI'da gösterilse bile bu fonksiyon
                     filter etmez; servis katmanı `amount > 0` dayatımını yapar.

    Raises:
        ValueError: JSON parse hatası veya yapı uyumsuzluğu.
    """
    import json

    if not raw_json:
        raise ValueError('entries parametresi boş.')

    try:
        data = json.loads(raw_json)
    except (json.JSONDecodeError, ValueError) as exc:
        raise ValueError(f'entries JSON parse hatası: {exc}')

    if not isinstance(data, list):
        raise ValueError('entries bir liste olmalı.')

    cleaned = []
    for idx, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f'entries[{idx}] dict olmalı.')
        cleaned.append({
            'currency':         item.get('currency'),
            'direction':        item.get('direction'),
            'amount':           item.get('amount'),
            'exchange_rate_eur': item.get('exchange_rate_eur'),
        })

    return cleaned


@login_required(login_url='login')
@role_required('SUPPLIERS_OPENING_BALANCE_CREATE')
@transaction.atomic
def supplier_opening_balance_create(request):
    """
    Tedarikçi için açılış carisi (opening balance) yazımı endpoint'i.

    POST Parametreleri:
        supplier_id  (zorunlu) — Suppliers UUID
        entries      (zorunlu) — JSON string. Liste, her eleman:
                                 {currency, direction, amount, exchange_rate_eur}
        description  (zorunlu) — Denetim notu (min 3 karakter)

    Response (başarılı):
        {
            'result': True,
            'process_no': 'OB-20260505-...',
            'entries': [...],
            'entries_count': N,
            'message': 'Açılış carisi başarıyla kaydedildi.',
        }

    Response (idempotent ihlali):
        409 Conflict, error_code='OB_ALREADY_EXISTS'

    Güvenlik:
        - @login_required, @role_required('SUPPLIERS_OPENING_BALANCE_CREATE')
        - @transaction.atomic — kısmi yazım yok (tüm satırlar veya hiçbiri).
        - select_for_update — eşzamanlı OB yazımı serileştirilir.
    """
    if request.method != 'POST':
        return JsonResponse(
            {'error': True, 'error_msg': 'Geçersiz istek.'}, status=405
        )

    supplier_id = (request.POST.get('supplier_id') or '').strip()
    description = (request.POST.get('description') or '').strip()
    entries_raw = (request.POST.get('entries') or '').strip()

    if not supplier_id:
        return JsonResponse(
            {'error': True, 'error_msg': 'Tedarikçi seçilmedi.'}, status=400
        )
    if len(description) < 3:
        return JsonResponse(
            {
                'error': True,
                'error_msg': (
                    'Açıklama zorunludur (en az 3 karakter). '
                    'Açılış kaynağı için referans/not girin.'
                ),
            },
            status=400,
        )

    try:
        entries = _parse_opening_balance_entries(entries_raw)
    except ValueError as exc:
        return JsonResponse(
            {'error': True, 'error_msg': str(exc)}, status=400
        )

    if not entries:
        return JsonResponse(
            {
                'error': True,
                'error_msg': 'En az bir açılış satırı girmelisiniz.',
            },
            status=400,
        )

    try:
        supplier = (Suppliers.objects
                    .select_for_update()
                    .get(pk=supplier_id))
    except Suppliers.DoesNotExist:
        return JsonResponse(
            {'error': True, 'error_msg': 'Tedarikçi bulunamadı.'}, status=404
        )

    # İdempotent kontrol — kilitten sonra okuma (race-safe).
    existing_ob = SupplierLedger.objects.filter(
        supplier=supplier,
        process_no__startswith='OB-',
    ).first()
    if existing_ob:
        return JsonResponse(
            {
                'error': True,
                'error_code': 'OB_ALREADY_EXISTS',
                'error_msg': (
                    'Bu tedarikçi için açılış carisi daha önce girilmiş '
                    f'(Fiş No: {existing_ob.process_no}). Mevcut bakiyeyi '
                    'değiştirmek için "Cariyi Düzelt" akışını kullanın.'
                ),
                'existing_process_no': existing_ob.process_no,
            },
            status=409,
        )

    # ── Servis çağrısı ──
    from apps.suppliers.services import SupplierLedgerService

    try:
        result = SupplierLedgerService.write_opening_balance(
            supplier=supplier,
            entries=entries,
            description=description,
            user=request.user,
        )
    except ValueError as exc:
        try:
            transaction.set_rollback(True)
        except Exception:
            pass
        return JsonResponse(
            {'error': True, 'error_msg': str(exc)}, status=400
        )
    except Exception as exc:
        try:
            transaction.set_rollback(True)
        except Exception:
            pass
        try:
            write_log(
                request,
                'SUPPLIER_OPENING_BALANCE_FAIL',
                f'Beklenmedik hata — supplier_id={supplier.id}, exc={exc!r}',
            )
        except Exception:
            pass
        return JsonResponse(
            {
                'error': True,
                'error_code': 'OB_UNEXPECTED',
                'error_msg': (
                    'Açılış carisi yazımında beklenmedik hata. '
                    'Lütfen tekrar deneyin; sorun sürerse destek ile '
                    'iletişime geçin.'
                ),
            },
            status=500,
        )

    # ── Aktivite logu ──
    try:
        write_log(
            request,
            'SUPPLIER_OPENING_BALANCE',
            (
                f'Açılış carisi yazıldı — supplier_id={supplier.id}, '
                f'process_no={result["process_no"]}, '
                f'entries={result["entries_count"]}, '
                f'note={description[:100]}'
            ),
        )
    except Exception:
        pass

    return JsonResponse({
        'result': True,
        'process_no': result['process_no'],
        'entries': result['entries'],
        'entries_count': result['entries_count'],
        'message': 'Açılış carisi başarıyla kaydedildi.',
    })


@login_required(login_url='login')
def supplier_opening_balance_status(request):
    """
    Tedarikçi için açılış carisinin daha önce girilip girilmediğini sorgular.

    Modal açılırken çağrılır; mevcutsa UI "Tek seferlik işlem zaten yapılmış"
    bilgilendirmesi gösterir ve form alanlarını disable eder.

    GET Parametreleri:
        supplier_id (zorunlu)

    Response:
        { 'exists': bool, 'process_no': str|null, 'created_on': iso|null }
    """
    supplier_id = (request.GET.get('supplier_id') or '').strip()
    if not supplier_id:
        return JsonResponse(
            {'error': True, 'error_msg': 'supplier_id zorunlu.'}, status=400
        )

    row = (SupplierLedger.objects
           .filter(supplier_id=supplier_id, process_no__startswith='OB-')
           .order_by('-created_on')
           .values('process_no', 'created_on')
           .first())

    if not row:
        return JsonResponse({'exists': False})

    return JsonResponse({
        'exists': True,
        'process_no': row['process_no'],
        'created_on': row['created_on'].isoformat() if row['created_on'] else None,
    })


# ════════════════════════════════════════════════════════════════════════════
# FAZ 2 / KARAR-04 — İŞLEM DETAY ENDPOINT (2026-04-29)
# ════════════════════════════════════════════════════════════════════════════
# Tedarikçi detay sayfasındaki "process_no" badge'ine tıklandığında açılan
# zengin işlem detay modalı için JSON veri sağlar. Mimari raporun 1.3 bölümü
# bu modalın içermesi gereken tüm bilgileri tanımladı:
#   - Fatura/Fiş Kimliği: process_no, tarih, kullanıcı, line_item_type
#   - Mal Hareketi: ürün, gram, milyem, price_hs/tl
#   - Finansal Hareket: SupplierLedger satırları (aktif + FIFO ile kapanmış)
#                       + exchange_rate_eur (frozen kur) + ADJ etiketi
#   - Bağlı İşlemler: aynı process_no ile bağlantılı tüm satırlar
#
# Yetki: SUPPLIERS_SUPPLIERS_DETAIL — detay sayfasını görme yetkisi olan
# kullanıcı bu endpoint'i de kullanabilir (mantıklı genişletme).
# ════════════════════════════════════════════════════════════════════════════

@login_required(login_url='login')
@role_required('SUPPLIERS_SUPPLIERS_DETAIL')
def supplier_process_detail(request):
    """
    GET parametreleri:
        supplier_id (zorunlu) — Suppliers UUID
        process_no  (zorunlu) — İşlem fiş numarası

    Response (başarılı):
        {
            'result': True,
            'process_no': '...',
            'is_adjustment': bool,            # ADJ- prefix'li mi
            'process_rows': [                  # Process tablosu satırları
                {date, transaction_type, transaction_label, employee_name,
                 product_name, piece, gram, process_mileage, amount_eur,
                 price_hs, hs_rate_buy_eur, is_status, ...}
            ],
            'ledger_rows': [                   # SupplierLedger satırları
                {transaction_type, transaction_label, currency,
                 amount_value, exchange_rate_eur, description,
                 created_on, is_active, cantaci_tx_type}
                # is_active=False olanlar (FIFO kapanmış) da dahil
            ],
            'summary': {
                'currencies': [...],
                'active_count', 'closed_count',
                'process_count', 'ledger_count',
            }
        }

    Response (hata): {'error': True, 'error_msg': '...'} + uygun HTTP status
    """
    supplier_id = (request.GET.get('supplier_id') or '').strip()
    process_no = (request.GET.get('process_no') or '').strip()

    if not supplier_id or not process_no:
        return JsonResponse(
            {'error': True, 'error_msg': 'supplier_id ve process_no zorunludur.'},
            status=400,
        )

    try:
        supplier = Suppliers.objects.get(pk=supplier_id)
    except (Suppliers.DoesNotExist, ValueError):
        return JsonResponse(
            {'error': True, 'error_msg': 'Tedarikçi bulunamadı.'}, status=404
        )

    # ADJ- prefix kontrolü (SupplierLedgerService üzerinden)
    from apps.suppliers.services import SupplierLedgerService
    is_adjustment = SupplierLedgerService.is_adjustment_process_no(process_no)

    # ── İnsan okunabilir transaction_type etiketleri ──
    TX_LABELS = {
        'PURCHASE':   'Alış (Toptan)',
        'STOCK_IN':   'Stok Girişi',
        'SALE':       'Satış',
        'RETURN':     'İade',
        'PAYMENT':    'Ödeme',
        'COLLECTION': 'Tahsilat',
        'ORDER_IN':   'Sipariş Girişi',
        'TRANSFER':   'Transfer',
    }

    # ── Process satırları (mal hareketleri) ──
    proc_rows = (Process.objects
                 .filter(supplier=supplier, process_no=process_no, is_deleted=False)
                 .select_related('employee', 'product', 'product__category')
                 .order_by('date'))

    process_data = []
    for p in proc_rows:
        # Çalışan adı — first_name + last_name birleşik, yoksa username
        emp_name = '-'
        if p.employee:
            full = ' '.join(filter(None, [
                getattr(p.employee, 'first_name', '') or '',
                getattr(p.employee, 'last_name', '') or '',
            ])).strip()
            emp_name = full or getattr(p.employee, 'username', '') or '-'

        # Tarih güvenli formatla
        try:
            date_str = (
                timezone.localtime(p.date).strftime('%d.%m.%Y %H:%M:%S')
                if p.date else '-'
            )
        except Exception:
            date_str = '-'

        process_data.append({
            'id': str(p.id),
            'date': date_str,
            'transaction_type': p.transaction_type or '',
            'transaction_label': TX_LABELS.get(
                p.transaction_type or '', p.transaction_type or '-'
            ),
            'employee_name': emp_name,
            'product_name': p.product.name if p.product else '-',
            'product_category': (
                p.product.category.name
                if (p.product and p.product.category) else None
            ),
            'piece': int(p.piece or 0),
            'gram': str(p.gram or 0),
            'process_mileage': str(p.process_mileage) if p.process_mileage else '-',
            'amount_eur': str(p.amount or 0),
            'price_hs': str(p.price_hs or 0),
            'hs_rate_buy_eur': str(p.hs_rate_buy_eur) if p.hs_rate_buy_eur else None,
            'hs_rate_sale_eur': str(p.hs_rate_sale_eur) if p.hs_rate_sale_eur else None,
            'karat': p.karat,
            'is_status': p.is_status,
        })

    # ── SupplierLedger satırları (aktif + kapanmış FIFO satırları dahil) ──
    ledger_rows = list(SupplierLedger.objects
                       .filter(supplier=supplier, process_no=process_no)
                       .select_related('product')
                       .order_by('created_on'))

    ledger_data = []
    for l in ledger_rows:
        try:
            cr_str = (
                timezone.localtime(l.created_on).strftime('%d.%m.%Y %H:%M:%S')
                if l.created_on else '-'
            )
        except Exception:
            cr_str = '-'

        # description'dan ADJUSTMENT etiketini ayır (raporlarda parse edilebilir)
        desc = l.description or ''
        is_adj_row = desc.startswith('[ADJUSTMENT]') or is_adjustment

        ledger_data.append({
            'id': str(l.id),
            'transaction_type': l.transaction_type,
            'transaction_label': (
                'Giriş (Borç)' if l.transaction_type == 'ENTRY'
                else 'Çıkış (Alacak)'
            ),
            'currency': l.currency or '',
            'amount_value': str(l.amount_value or 0),
            'exchange_rate_eur': (
                str(l.exchange_rate_eur)
                if l.exchange_rate_eur is not None else None
            ),
            'description': desc,
            'created_on': cr_str,
            'is_active': bool(l.is_active),
            'is_adjustment_row': is_adj_row,
            'cantaci_tx_type': l.cantaci_tx_type,
            'product_name': l.product.name if l.product else None,
            'quantity_piece': int(l.quantity_piece or 0),
            'quantity_gram': str(l.quantity_gram or 0),
        })

    # ── Currency-bazlı özet (yalnızca aktif satırlar) ──
    currency_summary_dict = {}
    active_count = 0
    closed_count = 0
    for l in ledger_rows:
        if l.is_active:
            active_count += 1
            cur = (l.currency or '').upper()
            entry = currency_summary_dict.setdefault(
                cur, {'entry': Decimal('0'), 'exit': Decimal('0')}
            )
            if l.transaction_type == 'ENTRY':
                entry['entry'] += (l.amount_value or Decimal('0'))
            else:
                entry['exit'] += (l.amount_value or Decimal('0'))
        else:
            closed_count += 1

    summary_currencies = []
    for cur, vals in currency_summary_dict.items():
        net = vals['entry'] - vals['exit']
        summary_currencies.append({
            'currency': cur,
            'entry_total': str(vals['entry']),
            'exit_total': str(vals['exit']),
            'net': str(net),
            'direction': (
                'RECEIVABLE' if net < 0  # ledger'da EXIT alacaktır → biz alacaklı
                else 'PAYABLE' if net > 0
                else 'BALANCED'
            ),
        })
    # Currencies'i metal-önce-fiat-sonra olarak sırala
    PILL_ORDER = {'HS': 0, 'HG': 1, 'TRY': 2, 'USD': 3, 'EUR': 4,
                  'GBP': 5, 'CAD': 6, 'QAR': 7}
    summary_currencies.sort(key=lambda c: PILL_ORDER.get(c['currency'], 99))

    return JsonResponse({
        'result': True,
        'process_no': process_no,
        'is_adjustment': is_adjustment,
        'process_rows': process_data,
        'ledger_rows': ledger_data,
        'summary': {
            'currencies': summary_currencies,
            'active_count': active_count,
            'closed_count': closed_count,
            'process_count': len(process_data),
            'ledger_count': len(ledger_data),
        },
    })